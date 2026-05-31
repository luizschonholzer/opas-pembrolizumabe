"""
Fixes the output spreadsheet (sets Sheet1 as the active/first tab)
and generates the final comprehensive report.
"""
import json, shutil, openpyxl
from pathlib import Path
from datetime import datetime
from openpyxl.styles import Alignment, Font, PatternFill, GradientFill
from openpyxl.utils import get_column_letter
from openpyxl.styles.colors import Color

BASE_DIR    = Path(__file__).parent
INPUT_FILE  = BASE_DIR / "OPAS_PDF.xlsx"
OUTPUT_FILE = BASE_DIR / "OPAS_PDF_CLAUDE_ANALISE.xlsx"
REPORT_FILE = BASE_DIR / "RELATORIO_CLAUDE_SDK_FINAL.txt"

COL_NEW    = 28
NEW_HEADER = "Comprehensive Clinical and Judicial Analysis (Claude SDK)"

FIELD_LABELS = {
    "PRIMARY_DIAGNOSIS":      "01. Primary Diagnosis",
    "SECONDARY_DIAGNOSES":    "02. Secondary Diagnoses",
    "ICD_CODE":               "03. ICD-10 Code",
    "PATIENT_AGE":            "04. Patient Age",
    "CLINICAL_HISTORY":       "05. Clinical History",
    "SYMPTOMS":               "06. Symptoms",
    "DISEASE_SEVERITY":       "07. Disease Severity and Performance Status",
    "FUNCTIONAL_LIMITATIONS": "08. Functional Limitations",
    "COMORBIDITIES":          "09. Comorbidities",
    "PRIOR_TREATMENTS":       "10. Prior Treatments",
    "THERAPEUTIC_FAILURE":    "11. Therapeutic Failure",
    "BIOMARKERS":             "12. Biomarkers",
    "METASTASIS_SITES":       "13. Metastasis Sites",
    "MEDICAL_JUSTIFICATION":  "14. Medical Justification",
    "URGENCY":                "15. Clinical Urgency",
    "RISK_DETERIORATION":     "16. Risk of Clinical Deterioration",
    "RISK_HOSPITALIZATION":   "17. Risk of Hospitalization",
    "RISK_DEATH":             "18. Risk of Death",
    "REQUESTED_TREATMENT":    "19. Requested Treatment",
    "JUDICIAL_DECISION":      "20. Judicial Decision",
    "DECISION_GROUNDS":       "21. Decision Grounds",
    "COURT_CITATIONS":        "22. Court Citations",
}

def format_cell(fields, missing, incorrect):
    lines = []
    for key, label in FIELD_LABELS.items():
        raw = fields.get(key) or "NOT FOUND"
        val = str(raw).strip() or "NOT FOUND"
        lines.append(f"{label}: {val}")
    lines.append("")
    lines.append("-- COMPARISON WITH EXISTING SPREADSHEET DATA (Columns N and O) --")
    miss = str(missing or "").strip()
    if miss and "none" not in miss.lower() and miss.upper() != "NOT FOUND":
        lines.append(f"[MISSING FROM N/O] Information present in PDF but absent from columns N/O:")
        lines.append(f"  {miss}")
    else:
        lines.append("Missing from N/O: None detected — columns appear adequately complete")
    incorr = str(incorrect or "").strip()
    if incorr and "no conflicts" not in incorr.lower() and incorr.upper() != "NOT FOUND":
        lines.append(f"[POSSIBLY INCORRECT IN N/O] Potential conflicts with PDF content:")
        lines.append(f"  {incorr}")
    else:
        lines.append("Possibly incorrect in N/O: No conflicts detected")
    lines.append("")
    lines.append(f"[Analyzed by Claude Agent SDK | claude-opus-4-7 | {datetime.now().strftime('%Y-%m-%d')}]")
    return "\n".join(lines)

def main():
    # Load all results
    result_files = sorted(BASE_DIR.glob("resultados_batch_txt_*.json"))
    print(f"Loading {len(result_files)} result files...")

    row_data = {}
    for rf in result_files:
        results = json.loads(rf.read_text(encoding="utf-8"))
        for item in results:
            if item.get("status") == "not_found":
                continue
            for row_info in item.get("rows", []):
                row_idx = row_info["row"]
                row_data[row_idx] = {
                    "fields":    item.get("fields", {}),
                    "missing":   item.get("missing_from_no", ""),
                    "incorrect": item.get("possibly_incorrect", ""),
                    "n":         row_info.get("N", ""),
                    "o":         row_info.get("O", ""),
                    "pdf":       item.get("pdf_name", ""),
                    "country":   item.get("country", ""),
                    "error":     item.get("error", ""),
                }

    print(f"  -> {len(row_data)} rows mapped from PDFs.")

    # Fresh copy from original
    print(f"Copying original to output file...")
    shutil.copy2(INPUT_FILE, OUTPUT_FILE)

    wb = openpyxl.load_workbook(str(OUTPUT_FILE))

    # Make Sheet1 the active sheet AND move it to first position
    if "Sheet1" in wb.sheetnames:
        wb.active = wb["Sheet1"]
        # Move Sheet1 to first position
        wb.move_sheet("Sheet1", offset=-wb.sheetnames.index("Sheet1"))

    ws = wb["Sheet1"]

    # Write column header
    hdr = ws.cell(row=1, column=COL_NEW)
    hdr.value = NEW_HEADER
    hdr.font = Font(bold=True, color="FFFFFF")
    hdr.fill = PatternFill("solid", fgColor="2E5FA3")
    hdr.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    # Write data
    stats = {"written": 0, "errors": 0}
    flags_missing = 0
    flags_incorrect = 0

    report_entries = []

    for row_idx, data in sorted(row_data.items()):
        cell = ws.cell(row=row_idx, column=COL_NEW)
        if data.get("error"):
            cell.value = f"[ANALYSIS ERROR: {data['error'][:400]}]"
            stats["errors"] += 1
        else:
            text = format_cell(data["fields"], data["missing"], data["incorrect"])
            cell.value = text
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            stats["written"] += 1

            has_miss = (str(data["missing"]) and
                        "none" not in str(data["missing"]).lower() and
                        str(data["missing"]).upper() != "NOT FOUND")
            has_incorr = (str(data["incorrect"]) and
                          "no conflicts" not in str(data["incorrect"]).lower() and
                          str(data["incorrect"]).upper() != "NOT FOUND")
            if has_miss:
                flags_missing += 1
            if has_incorr:
                flags_incorrect += 1

            # Build report entry
            fields = data["fields"]
            entry = {
                "row":       row_idx,
                "country":   data["country"],
                "case":      ws.cell(row=row_idx, column=6).value,
                "pdf":       data["pdf"],
                "diagnosis": str(fields.get("PRIMARY_DIAGNOSIS", "NOT FOUND"))[:200],
                "decision":  str(fields.get("JUDICIAL_DECISION", "NOT FOUND"))[:100],
                "grounds":   str(fields.get("DECISION_GROUNDS", "NOT FOUND"))[:300],
                "missing":   str(data["missing"])[:300],
                "incorrect": str(data["incorrect"])[:300],
                "has_miss":  has_miss,
                "has_incorr":has_incorr,
            }
            report_entries.append(entry)

    # Set column width for AB
    ws.column_dimensions[get_column_letter(COL_NEW)].width = 80

    # Set row height to auto (approximate)
    ws.row_dimensions[1].height = 30

    wb.save(str(OUTPUT_FILE))
    print(f"Spreadsheet saved: {OUTPUT_FILE.name}")
    print(f"  Rows written: {stats['written']} | Errors: {stats['errors']}")

    # ----------------------------------------------------------------
    # Generate final report
    # ----------------------------------------------------------------
    lines = []
    lines.append("=" * 80)
    lines.append("OPAS PEMBROLIZUMAB JUDICIAL CASES — FINAL ANALYSIS REPORT")
    lines.append("Generated by Claude Agent SDK (claude-opus-4-7)")
    lines.append(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lines.append("=" * 80)
    lines.append("")
    lines.append("SUMMARY")
    lines.append("-" * 40)
    lines.append(f"Total spreadsheet rows:              197")
    lines.append(f"Rows with PDF analyzed:              {stats['written']}")
    lines.append(f"Rows without PDF (skipped):          {197 - stats['written'] - stats['errors']}")
    lines.append(f"Rows with analysis errors:           {stats['errors']}")
    lines.append(f"Rows with data MISSING from N/O:     {flags_missing}")
    lines.append(f"Rows with POSSIBLY INCORRECT N/O:    {flags_incorrect}")
    lines.append("")

    # Group by country
    by_country = {}
    for e in report_entries:
        by_country.setdefault(e["country"], []).append(e)

    lines.append("RESULTS BY COUNTRY")
    lines.append("-" * 40)
    for country, entries in sorted(by_country.items()):
        miss_c = sum(1 for e in entries if e["has_miss"])
        incorr_c = sum(1 for e in entries if e["has_incorr"])
        lines.append(f"\n{country.upper()} ({len(entries)} rows analyzed | "
                      f"{miss_c} missing flags | {incorr_c} incorrect flags)")
        lines.append("-" * 60)
        for e in entries:
            lines.append(f"\n  ROW {e['row']} | Case: {e['case']} | PDF: {e['pdf']}")
            lines.append(f"  Primary Diagnosis: {e['diagnosis']}")
            lines.append(f"  Judicial Decision: {e['decision']}")
            lines.append(f"  Decision Grounds:  {e['grounds']}")
            if e["has_miss"]:
                lines.append(f"  [MISSING FROM N/O]: {e['missing']}")
            if e["has_incorr"]:
                lines.append(f"  [POSSIBLY INCORRECT IN N/O]: {e['incorrect']}")

    lines.append("")
    lines.append("=" * 80)
    lines.append("KEY INCONSISTENCIES FOUND ACROSS ALL CASES")
    lines.append("=" * 80)
    incorrect_entries = [e for e in report_entries if e["has_incorr"]]
    for e in incorrect_entries:
        lines.append(f"\n  ROW {e['row']} | {e['country']} | Case: {e['case']}")
        lines.append(f"  {e['incorrect']}")

    lines.append("")
    lines.append("=" * 80)
    lines.append("END OF REPORT")
    lines.append("=" * 80)

    REPORT_FILE.write_text("\n".join(lines), encoding="utf-8")
    print(f"Final report saved: {REPORT_FILE.name}")

    print(f"\n  [MISSING FROM N/O] flags:     {flags_missing} rows")
    print(f"  [POSSIBLY INCORRECT] flags:   {flags_incorrect} rows")
    print("\nDone.")

if __name__ == "__main__":
    main()
