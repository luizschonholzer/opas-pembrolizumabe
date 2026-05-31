import re
import openpyxl
from pathlib import Path

XLSX_PATH = Path(r"C:\Users\Claudio\Downloads\Processos - OPAS\OPAS_PDF_CLAUDE_ANALISE.xlsx")


def strip_topics_20_to_22(text):
    if not text:
        return text

    # Localiza o início do tópico 20 (no começo de uma linha)
    match_20 = re.search(r'\n\s*20\.[ \t]', text)

    # Localiza a seção de comparação (-- COMPARISON ... ou --- COMPARISON ...)
    match_comp = re.search(r'\n\s*-{2,}\s*COMPARISON', text, re.IGNORECASE)

    if match_20 is None:
        return text  # tópico 20 não existe nesta célula, nada a fazer

    # Texto antes do tópico 20 (sem espaços/quebras sobrando no final)
    before_20 = text[:match_20.start()].rstrip()

    if match_comp is not None:
        # Reconecta: tópicos 01-19 + quebra dupla + seção de comparação
        comparison_block = text[match_comp.start():].lstrip('\n')
        return before_20 + '\n\n' + comparison_block
    else:
        # Sem seção de comparação — apenas trunca após o tópico 19
        return before_20


def main():
    print(f"Abrindo: {XLSX_PATH}")
    wb = openpyxl.load_workbook(str(XLSX_PATH))
    ws = wb.active

    total = 0
    modified = 0
    already_clean = 0
    empty = 0

    for row_idx in range(2, ws.max_row + 1):
        # Para se a primeira coluna estiver vazia (fim dos dados)
        if not ws.cell(row=row_idx, column=1).value:
            break

        cell = ws.cell(row=row_idx, column=28)  # Coluna AB
        original = cell.value

        if not original or len(str(original).strip()) < 50:
            empty += 1
            continue

        total += 1
        new_value = strip_topics_20_to_22(str(original))

        if new_value != str(original):
            cell.value = new_value
            modified += 1
        else:
            already_clean += 1

    print(f"\n=== RELATÓRIO ===")
    print(f"Linhas com conteúdo em AB   : {total}")
    print(f"Linhas modificadas          : {modified}")
    print(f"Linhas já sem tópico 20     : {already_clean}")
    print(f"Linhas sem conteúdo em AB   : {empty}")

    if modified > 0:
        wb.save(str(XLSX_PATH))
        print(f"\nArquivo salvo: {XLSX_PATH}")
    else:
        print("\nNenhuma alteração necessária — arquivo não foi regravado.")

    # Verificação rápida: mostra o final do conteúdo da linha 2
    wb2 = openpyxl.load_workbook(str(XLSX_PATH), read_only=True)
    ws2 = wb2.active
    sample = ws2.cell(row=2, column=28).value or ""
    print("\n=== AMOSTRA (linha 2, últimos 400 chars da col AB) ===")
    print(repr(sample[-400:]) if len(sample) > 400 else repr(sample))
    wb2.close()


if __name__ == "__main__":
    main()
