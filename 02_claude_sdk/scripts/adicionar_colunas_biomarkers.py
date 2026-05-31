import openpyxl
from openpyxl.utils import get_column_letter

FILEPATH = r"C:\Users\Claudio\Downloads\Processos - OPAS\OPAS_PDF_CLAUDE_ANALISE.xlsx"

MSI_H_DATA = {
    16: "Patient status: POSITIVE (MSI-H). FDA tumor-agnostic approval (2017): pembrolizumab indicated for any unresectable or metastatic MSI-H/dMMR solid tumor regardless of histological type. KEYNOTE-177 (CRC 1st line): median OS 77.5 vs. 36.7 months with chemotherapy. Strong legal and clinical basis for access.",
    74: "Patient status: NEGATIVE (Microsatellite Stable — MSS). MSS tumors do not respond to anti-PD-1 immunotherapy. Pembrolizumab is not indicated for MSS/pMMR tumors. Absence of MSI-H provides clinical grounds for judicial denial.",
    111: "Patient status: POSITIVE (MSI-H confirmed by IHC — loss of MLH-1, MSH-6, and PMS2). FDA tumor-agnostic approval (2017): indicated for any MSI-H/dMMR solid tumor post-progression, regardless of histology. KEYNOTE-158: ORR 30.8% in non-colorectal MSI-H/dMMR solid tumors; median duration of response 47.5 months.",
    167: "Patient status: POSITIVE (MSI-H confirmed by hepatic biopsy). FDA tumor-agnostic approval: pembrolizumab indicated for MSI-H/dMMR metastatic sigmoid adenocarcinoma. KEYNOTE-177: median OS 77.5 months in 1st line vs. 36.7 months with chemotherapy. KRAS/NRAS/BRAF negative — no targeted therapy alternative available.",
    181: "Patient status: POSITIVE (MSI-H — loss of MLH1 and PMS2, dMMR confirmed). Rare tumor (intrahepatic cholangiocarcinoma). KEYNOTE-158: ORR 30.8% in advanced non-colorectal MSI-H/dMMR solid tumors; median duration of response 47.5 months. FDA tumor-agnostic approval applies regardless of histological type.",
    186: "Patient status: POSITIVE (dMMR — MLH1 loss; MSH2/MSH6/PMS2 positive). Recurrent metastatic rectal adenocarcinoma. FDA tumor-agnostic approval. KEYNOTE-177: median OS 77.5 vs. 36.7 months with chemotherapy. Prior adjuvant chemotherapy discontinued due to toxicity — pembrolizumab as 1st systemic line.",
    187: "Patient status: POSITIVE (dMMR — MLH1 loss; MSH2/MSH6/PMS2 positive). Identical molecular profile to case at row 186 (same defendants, same hospital). FDA tumor-agnostic approval applies. Clinical and legal indication equivalent to row 186.",
}

PDL1_DATA = {
    42: "Patient PD-L1: >50% (TPS). NSCLC: TPS ≥50% threshold enables pembrolizumab as 1st-line monotherapy (KEYNOTE-024). Median OS 26.3 months; 5-year survival rate 32%; ORR 52%. No driver mutations (EGFR/ALK). Maximum level of evidence for 1st-line indication.",
    44: "Patient PD-L1: >50% (TPS) — lung adenocarcinoma without driver mutations. KEYNOTE-024: 1st-line pembrolizumab vs. chemotherapy in TPS ≥50% — median OS 26.3 months, 5-year survival 32%. NATJUS opinion confirms indication. No EGFR/ALK/ROS1 — pembrolizumab is the appropriate 1st-line therapy.",
    51: "Patient PD-L1: positive (exact percentage not specified). EGFR negative, ALK negative. For NSCLC with PD-L1 ≥1%: benefit from pembrolizumab + chemotherapy combination (KEYNOTE-189/407). Absence of exact value limits direct application of 50% monotherapy threshold.",
    62: "Patient PD-L1: not measured. FNR (Uruguay) coverage criterion: PD-L1 ≥50%. KEYNOTE-024: maximum benefit at ≥50%; between 1–49% benefit exists but is smaller. Absence of testing prevents threshold confirmation — legal argument: right to diagnostic testing before denial.",
    65: "Patient PD-L1: negative (<50%). FNR (Uruguay) requires ≥50% for pembrolizumab coverage in NSCLC. Evidence: below 50%, pembrolizumab monotherapy does not demonstrate superior survival over chemotherapy (KEYNOTE-024 — subgroup <50% excluded from benefit). FNR denial is clinically supported.",
    73: "Patient PD-L1: <1% (negative). FNR criterion: ≥50% for coverage. With PD-L1 <1%, anti-PD-1 immunotherapy alone has very limited efficacy. Alternative: pembrolizumab + chemotherapy combination (KEYNOTE-189) does not depend on PD-L1 threshold, but falls outside FNR protocol.",
    75: "Patient PD-L1: 40% — positive but below FNR threshold (≥50%). EGFR/ALK negative. For TPS 1–49%: pembrolizumab + chemotherapy combination has demonstrated benefit (KEYNOTE-189/407); monotherapy does not show superiority. Patient may benefit from combination regimen — argument for protocol revision.",
    79: "Patient PD-L1: 10% — positive, below ≥50% threshold. EGFR/ALK negative. For TPS 1–49%: pembrolizumab + chemotherapy benefit demonstrated (KEYNOTE-189). Monotherapy at PD-L1 10% lacks robust evidence of superiority over chemotherapy alone. FNR 50% criterion excludes this patient from coverage.",
    88: "Patient PD-L1: 30%. Hypopharyngeal cancer (HNSCC). For HNSCC: CPS ≥1 enables pembrolizumab 1st line (KEYNOTE-048); CPS ≥20 shows greater OS benefit. Note: 30% is TPS — CPS conversion often yields higher values (includes immune cells). May reach CPS ≥20 — CPS assessment is indicated.",
    89: "Patient PD-L1: 30% (same patient as case at row 88). Hypopharyngeal cancer. CPS ≥1: enables pembrolizumab; CPS ≥20: greater benefit (KEYNOTE-048). TPS 30% may correspond to CPS ≥20. Same clinical and legal context as rows 88, 90, 91.",
    90: "Patient PD-L1: 30% (same patient as cases at rows 88–89). Hypopharyngeal cancer. CPS ≥1 FDA-approved for HNSCC. TPS 30% potentially equivalent to CPS ≥20 (threshold with greatest demonstrated benefit). Four separate judicial proceedings for the same patient.",
    91: "Patient PD-L1: 30% (same patient as cases at rows 88–90). Hypopharyngeal cancer. Same clinical profile. CPS ≥1 is the approved threshold for HNSCC — distinct from the 50% TPS criterion used for lung cancer.",
    175: "Patient PD-L1: ≥10% referenced by ESMO for triple-negative breast cancer; exact patient value not documented. For TNBC: pembrolizumab approved with PD-L1 CPS ≥10 (KEYNOTE-522/355). Absence of exact patient CPS value is a relevant gap — CPS assessment recommended to confirm eligibility.",
}

print("Loading workbook...")
wb = openpyxl.load_workbook(FILEPATH)
ws = wb.active

# Find last column with data (should be AB = 28)
last_col = ws.max_column
print(f"Current max column: {get_column_letter(last_col)} ({last_col})")

# New columns: AC = 29, AD = 30
col_msi = 29
col_pdl1 = 30

ws.cell(row=1, column=col_msi).value = "MSI-H / dMMR Status — Clinical Evidence"
ws.cell(row=1, column=col_pdl1).value = "PD-L1 Expression — Clinical Evidence"

msi_count = 0
pdl1_count = 0

for spreadsheet_row, content in MSI_H_DATA.items():
    # spreadsheet_row is already the actual Excel row number (data row = row number)
    ws.cell(row=spreadsheet_row, column=col_msi).value = content
    msi_count += 1

for spreadsheet_row, content in PDL1_DATA.items():
    ws.cell(row=spreadsheet_row, column=col_pdl1).value = content
    pdl1_count += 1

print(f"MSI-H cells written: {msi_count}")
print(f"PD-L1 cells written: {pdl1_count}")

print("Saving workbook...")
wb.save(FILEPATH)
print("Done. File saved successfully.")
