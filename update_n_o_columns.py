# -*- coding: utf-8 -*-
"""
Atualiza colunas N (Disease) e O (Further clinical details) da planilha
OPAS_com_links_PDF.xlsx com base nos dados da coluna 28 (Additional Clinical
Findings) e nas correções identificadas na revisão manual.

Regras:
- Não apaga dados existentes
- Apenas corrige erros evidentes e adiciona informações ausentes
- Todo texto em inglês
"""

import openpyxl
from openpyxl.styles import Alignment
from copy import copy

BASE = r"C:\Users\Claudio\Downloads\Processos - OPAS"
XLSX = BASE + r"\OPAS_com_links_PDF.xlsx"

wb = openpyxl.load_workbook(XLSX)
ws = wb.active

COL_N   = 14
COL_O   = 15
COL_PDF = 27
COL_28  = 28

def cell_val(row, col):
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v is not None else ""

def is_empty(v):
    return v.upper() in ("", "NI", "NONE", "NULL", "NOT FOUND", "NONE\n")

def set_n(row, value):
    ws.cell(row=row, column=COL_N).value = value

def set_o(row, value):
    cell = ws.cell(row=row, column=COL_O)
    cell.value = value
    cell.alignment = Alignment(wrap_text=True)

def append_o(row, addition):
    """Append additional info to O; avoid duplicating existing content."""
    current = cell_val(row, COL_O)
    if is_empty(current):
        set_o(row, addition)
        return
    # Avoid duplicate
    if addition.lower()[:40] in current.lower():
        return
    set_o(row, current + " | " + addition)

log = []

def upd(row, col_label, old, new):
    log.append(f"Row {row:3d} | Col {col_label} | {old[:60]!r} → {new[:80]!r}")

# ─── COLUMN N CORRECTIONS (translation / factual errors) ────────────────────

# Row 2: Portuguese → English
r, old = 2, cell_val(2, COL_N)
new = "Lung cancer"
set_n(r, new); upd(r, "N", old, new)

# Row 92: contradictory BRAF info in N — correct to reflect test result in O
r, old = 92, cell_val(92, COL_N)
new = "Invasive nodular melanoma (BRAF V600E mutation: negative per molecular testing)"
set_n(r, new); upd(r, "N", old, new)

# Row 128: Portuguese → English
r, old = 128, cell_val(128, COL_N)
new = "Not about pembrolizumab"
set_n(r, new); upd(r, "N", old, new)

# Row 181: Portuguese portions in N → full English translation
r, old = 181, cell_val(181, COL_N)
new = ("Recurrent intrahepatic cholangiocarcinoma with IDH1 gene mutation "
       "and dMMR (loss of MLH1 and PMS2 expression); MSI-H "
       "(high microsatellite instability)")
set_n(r, new); upd(r, "N", old, new)

# ─── COLUMN O UPDATES ────────────────────────────────────────────────────────
# Format: (row, new_content_if_empty_OR_addition_to_existing)
# append_o handles both cases.

updates_o = {
    # Argentina
    2:  "Prior procedures: surgical procedure",
    5:  "Disease staging: Stage IV (metastatic/advanced disease)",
    7:  "64-year-old patient | Prior treatment: Lenvatinib | Clinical urgency: treatment marked as urgent by the treating physician",
    8:  "Prior treatments: Carboplatin, Paclitaxel, Cyclophosphamide | Clinical urgency: treatment marked as urgent by the treating physician",
    9:  "Prior treatments: Carboplatin, radiotherapy | Patient in serious/severe clinical condition",
    10: "Prior treatment: Axitinib | Patient in serious/severe clinical condition",

    # Brazil
    16: ("MSI-H (high microsatellite instability — strongly predicts response to Pembrolizumab) "
         "| Patient in serious/severe clinical condition"),
    17: "Active disease progression — condition worsening despite treatment",
    38: "Patient in serious/severe clinical condition",
    40: ("MSS/pMMR (microsatellite stable — lower predicted response to immunotherapy) "
         "| Patient in serious/severe clinical condition"),
    41: "Patient in serious/severe clinical condition",
    42: ("MSS/pMMR (microsatellite stable) "
         "| Clinical urgency: risk of death if treatment is not provided"),
    44: "Clinical urgency: risk of death if treatment is not provided",

    # Chile
    47: "15-year-old patient | Clinical urgency: treatment marked as urgent by the treating physician",

    # Uruguay
    59: "34-year-old patient | Clinical urgency: treatment marked as urgent by the treating physician",
    61: "73-year-old patient | Clinical urgency: treatment marked as urgent by the treating physician",
    62: ("68-year-old patient | PD-L1 expression: 50% | EGFR: negative | ALK: negative "
         "| Clinical urgency: treatment marked as urgent by the treating physician"),
    63: ("64-year-old patient | Prior treatments: Carboplatin, Paclitaxel, Lenvatinib "
         "| Clinical urgency: treatment marked as urgent by the treating physician"),
    64: "Prior treatment: Trastuzumab (suggests HER2-positive or gastric/breast cancer — disease not further specified in PDF)",
    65: "70-year-old patient | Active disease progression — condition worsening despite treatment",
    66: "72-year-old patient | Prior treatments: Carboplatin, Paclitaxel, radiotherapy",
    67: "Patient in serious/severe clinical condition",
    69: "35-year-old patient",
    70: ("66-year-old patient | Prior procedures: nephrectomy "
         "| Clinical urgency: treatment marked as urgent by the treating physician"),
    71: "61-year-old patient | Clinical urgency: treatment marked as urgent by the treating physician",
    72: "72-year-old patient | Prior treatment: Axitinib (tyrosine kinase inhibitor — targeted therapy)",
    73: "73-year-old patient | Clinical urgency: treatment marked as urgent by the treating physician",
    74: "80-year-old patient | Clinical urgency: treatment marked as urgent by the treating physician",
    75: ("67-year-old patient | Prior treatment: Docetaxel "
         "| Palliative treatment context (disease is incurable; goal is quality of life)"),
    76: "20-year-old patient | Disease staging: Stage III | Prior treatment: Cetuximab",
    79: ("72-year-old patient | EGFR: negative | ALK: negative "
         "| Clinical urgency: treatment marked as urgent by the treating physician"),
    82: "65-year-old patient",
    83: "Clinical urgency: treatment marked as urgent by the treating physician",
    85: ("61-year-old patient | Prior treatment: Axitinib "
         "| Active disease progression — condition worsening despite treatment"),
    87: "77-year-old patient | Patient in serious/severe clinical condition",
    89: ("66-year-old patient | Prior treatments: Carboplatin, Paclitaxel, Rituximab, radiotherapy"),
    92: "57-year-old patient | Patient in serious/severe clinical condition",
    93: ("Note: prior treatment history (DHAP, Cisplatin, bone marrow transplant) is consistent "
         "with prior Hodgkin lymphoma — patient may have a second primary lung cancer or there may "
         "be a case file discrepancy requiring verification | "
         "Clinical urgency: treatment marked as urgent by the treating physician"),
    94: ("56-year-old patient | Prior treatment: Docetaxel "
         "| Clinical urgency: treatment marked as urgent by the treating physician"),
    95: "65-year-old patient",

    # Guatemala
    99:  ("37-year-old patient | Prior treatments: Cisplatin, radiotherapy "
          "| Active disease progression — condition worsening despite treatment"),
    100: ("Metastasis to: lung | Life-threatening condition — treatment urgently needed"),
    102: ("60-year-old patient | Prior treatment: Paclitaxel "
          "| Patient in serious/severe clinical condition"),
    103: ("Prior treatment: Axitinib | Patient in serious/severe clinical condition"),
    104: "Patient in serious/severe clinical condition",
    105: "Active disease progression — condition worsening despite treatment",
    106: "Prior treatment: Capecitabine",
    107: ("Metastasis to: lung | Prior procedures: surgical procedure "
          "| Active disease progression — condition worsening despite treatment"),
    108: ("Prior procedures: surgical procedure "
          "| Clinical urgency: treatment marked as urgent by the treating physician"),
    109: "Active disease progression — condition worsening despite treatment",
    110: ("65-year-old patient | Prior treatment: radiotherapy "
          "| Patient in serious/severe clinical condition"),
    111: "Prior treatments: XELOX, Capecitabine, Oxaliplatin, XELIRI | Metastasis to: peritoneum",
    113: ("70-year-old patient | Prior treatments: Cisplatin, radiotherapy, subtotal glossectomy "
          "| Clinical urgency: treatment marked as urgent by the treating physician"),
    114: ("74-year-old patient | Metastasis to: lung "
          "| Prior treatments: Axitinib, nephrectomy, radiotherapy "
          "| Active disease progression — condition worsening despite treatment"),
    115: "Prior treatment: Capecitabine",
    116: ("Prior treatment: Axitinib "
          "| Clinical urgency: treatment marked as urgent by the treating physician"),
    117: "54-year-old patient | Clinical urgency: treatment marked as urgent by the treating physician",
    119: "Prior treatment: Axitinib",
    120: "Patient in serious/severe clinical condition",
    121: ("Prior treatments: Axitinib, nephrectomy "
          "| Patient in serious/severe clinical condition"),
    122: ("Prior treatment: radiotherapy "
          "| Patient in serious/severe clinical condition"),
    123: ("Metastasis to: bone | Prior treatments: Axitinib, radiotherapy "
          "| Patient in serious/severe clinical condition"),
    124: ("Prior treatment: Lenvatinib "
          "| Clinical urgency: treatment marked as urgent by the treating physician"),
    125: ("Prior treatments: Carboplatin, Paclitaxel "
          "| Life-threatening condition — treatment urgently needed"),
    126: "Metastasis to: bone | Patient in serious/severe clinical condition",
    127: "Patient in serious/severe clinical condition",
    129: ("61-year-old patient "
          "| Prior treatments: Gemcitabine, Bevacizumab, Doxorubicin, radiotherapy "
          "| Clinical urgency: treatment marked as urgent by the treating physician"),
    130: "Life-threatening condition — treatment urgently needed",

    # Costa Rica
    162: ("60-year-old patient | Prior treatments: Axitinib, nephrectomy "
          "| Patient in serious/severe clinical condition"),
    163: ("Prior treatments: Axitinib, nephrectomy "
          "| Clinical urgency: risk of death if treatment is not provided"),
    164: "Prior treatment: Axitinib",
    166: "74-year-old patient",
    167: ("57-year-old patient | Prior treatments: CAPOX, Capecitabine, Oxaliplatin "
          "| KRAS/NRAS/BRAF: data mentioned but incomplete (not detected in PDF extraction)"),
    170: "Prior procedures: surgical procedure | Patient in serious/severe clinical condition",
    172: ("70-year-old patient | Metastasis to: lung | Prior treatment: radiotherapy "
          "| Patient in serious/severe clinical condition"),
    173: ("49-year-old patient | Prior treatment: radiotherapy "
          "| Patient in serious/severe clinical condition"),
    174: "61-year-old patient | Prior treatments: Capecitabine, radiotherapy",
    175: ("44-year-old patient | PD-L1: positive "
          "| Prior treatments: Carboplatin, Paclitaxel, Cyclophosphamide "
          "| Clinical urgency: treatment marked as urgent by the treating physician"),
    176: ("67-year-old patient | Prior treatments: Docetaxel, Cetuximab "
          "| Patient in serious/severe clinical condition"),
    177: ("88-year-old patient | ECOG 1 (symptomatic but fully ambulatory) | PD-L1: 10% "
          "| Metastasis to: lung | Prior treatments: Cisplatin, Docetaxel, radiotherapy"),
    178: "62-year-old patient | Patient in serious/severe clinical condition",
    179: "60-year-old patient | Patient in serious/severe clinical condition",
    180: ("46-year-old patient "
          "| Prior treatments: Cisplatin, Doxorubicin, Gemcitabine, Oxaliplatin, bone marrow transplant "
          "| Patient in serious/severe clinical condition"),
    181: ("50-year-old patient | PD-L1 expression noted (value not extracted) "
          "| Prior treatments: Oxaliplatin, Capecitabine, radiotherapy "
          "| Patient in serious/severe clinical condition"),
    182: ("66-year-old patient "
          "| Prior treatments: Paclitaxel, Cisplatin, Gemcitabine, Docetaxel "
          "| Clinical urgency: treatment marked as urgent by the treating physician"),
    185: ("31-year-old patient | ECOG 0 (fully active, no symptoms) "
          "| Prior treatments: Cisplatin, Doxorubicin"),
    186: ("63-year-old patient | Pathological staging: pT3N0M0 "
          "| MSI-H (high microsatellite instability — strongly predicts response to Pembrolizumab) "
          "| Metastasis to: liver "
          "| Prior treatments: Bevacizumab, Cetuximab, FOLFOX, radiotherapy "
          "| Patient in serious/severe clinical condition"),
    191: ("49-year-old patient | ECOG 0 (fully active, no symptoms) "
          "| PD-L1: negative (low predicted response to immunotherapy) "
          "| HER2: equivocal (2+) "
          "| Prior treatments: Paclitaxel, Cyclophosphamide "
          "| Clinical urgency: treatment marked as urgent by the treating physician"),
    192: ("46-year-old patient | ECOG 0 (fully active, no symptoms) "
          "| Prior treatment: radiotherapy "
          "| Patient in serious/severe clinical condition"),
    193: ("68-year-old patient | Prior treatments: Carboplatin, Paclitaxel, Cyclophosphamide "
          "| Clinical urgency: treatment marked as urgent by the treating physician"),
    198: "74-year-old patient | Prior treatment: Oxaliplatin",
}

for row, addition in sorted(updates_o.items()):
    old = cell_val(row, COL_O)
    append_o(row, addition)
    new = cell_val(row, COL_O)
    if new != old:
        upd(row, "O", old, new)

# ─── SPECIAL NOTE for Row 80 (Uruguay 153/2022): N and O both NI ─────────────
r80_n = cell_val(80, COL_N)
r80_o = cell_val(80, COL_O)
if is_empty(r80_n):
    set_n(80, "Not identified in PDF (prior surgical procedure noted)")
    upd(80, "N", r80_n, "Not identified in PDF (prior surgical procedure noted)")
if is_empty(r80_o):
    set_o(80, "Prior procedures: surgical procedure (specific diagnosis not identified in PDF)")
    upd(80, "O", r80_o, "Prior procedures: surgical procedure (specific diagnosis not identified in PDF)")

# ─── SPECIAL NOTE for Row 64 (Uruguay 16/2025): N still NI ──────────────────
r64_n = cell_val(64, COL_N)
if is_empty(r64_n):
    new_n = "Not identified in PDF (Trastuzumab mentioned — suggests HER2-positive cancer)"
    set_n(64, new_n)
    upd(64, "N", r64_n, new_n)

# ─── Row 197 (Costa Rica 05332-2025): col28 data appears unreliable ──────────
# The PDF is shared across multiple rows; extracted age (10 yrs) and prior
# treatments (Axitinib, BMT) are clinically inconsistent with breast cancer.
# Do NOT update N or O for this row.

# ─── Save ─────────────────────────────────────────────────────────────────────
wb.save(XLSX)

print(f"Saved: {XLSX}")
print(f"\nTotal updates applied: {len(log)}")
print("\nChange log:")
for entry in log:
    print(" ", entry)
