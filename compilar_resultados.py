"""
Compiles all agent result JSON files and writes the new column
"Comprehensive Clinical and Judicial Analysis (Claude SDK)" to the spreadsheet.

Run after all batch agents have finished:
    python compilar_resultados.py
"""
import json, shutil, os
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR    = Path(__file__).parent
INPUT_FILE  = BASE_DIR / "OPAS_PDF.xlsx"
OUTPUT_FILE = BASE_DIR / "OPAS_PDF_CLAUDE_ANALISE.xlsx"
REPORT_FILE = BASE_DIR / "RELATORIO_CLAUDE_SDK.txt"

COL_NEW    = 28   # AB
NEW_HEADER = "Comprehensive Clinical and Judicial Analysis (Claude SDK)"

FIELD_LABELS = {
    "PRIMARY_DIAGNOSIS":      "01. Primary Diagnosis",
    "SECONDARY_DIAGNOSES":    "02. Secondary Diagnoses",
    "ICD_CODE":               "03. ICD-10 Code",
    "PATIENT_AGE":            "04. Patient Age",
    "CLINICAL_HISTORY":       "05. Clinical History",
    "SYMPTOMS":               "06. Symptoms",
    "DISEASE_SEVERITY":       "07. Disease Severity and Performance Status",
    "FUNCTIONAL_LIMITATIONS": "08. Functional Limitations",
    "COMORBIDITIES":          "09. Comorbidities",
    "PRIOR_TREATMENTS":       "10. Prior Treatments",
    "THERAPEUTIC_FAILURE":    "11. Therapeutic Failure",
    "BIOMARKERS":             "12. Biomarkers",
    "METASTASIS_SITES":       "13. Metastasis Sites",
    "MEDICAL_JUSTIFICATION":  "14. Medical Justification",
    "URGENCY":                "15. Clinical Urgency",
    "RISK_DETERIORATION":     "16. Risk of Clinical Deterioration",
    "RISK_HOSPITALIZATION":   "17. Risk of Hospitalization",
    "RISK_DEATH":             "18. Risk of Death",
    "REQUESTED_TREATMENT":    "19. Requested Treatment",
    "JUDICIAL_DECISION":      "20. Judicial Decision",
    "DECISION_GROUNDS":       "21. Decision Grounds",
    "COURT_CITATIONS":        "22. Court Citations",
}

def format_cell(fields: dict, row_n: str, row_o: str, missing: str, incorrect: str) -> str:
    lines = ["=== COMPREHENSIVE CLINICAL AND JUDICIAL ANALYSIS (Claude SDK) ===", ""]
    for key, label in FIELD_LABELS.items():
        raw = fields.get(key) or "NOT FOUND"
        val = str(raw).strip() if not isinstance(raw, str) else raw.strip()
        val = val or "NOT FOUND"
        lines.append(f"{label}: {val}")
    lines.append("")
    lines.append("--- COMPARISON WITH EXISTING SPREADSHEET DATA ---")
    if missing and missing.upper() != "NOT FOUND" and "none" not in missing.lower():
        lines.append(f"[MISSING FROM N/O] Information present in the PDF but absent from columns N/O:")
        lines.append(f"  {missing}")
    else:
        lines.append("Missing from N/O: None — columns appear adequately complete")
    if incorrect and "no conflicts" not in incorrect.lower() and incorrect.upper() != "NOT FOUND":
        lines.append(f"[POSSIBLY INCORRECT IN N/O] Potential conflicts with PDF content:")
        lines.append(f"  {incorrect}")
    else:
        lines.append("Possibly incorrect in N/O: No conflicts detected")
    lines.append("")
    lines.append(f"[Analyzed by Claude Agent SDK on {datetime.now().strftime('%Y-%m-%d')}]")
    return "\n".join(lines)

def main():
    # Collect all result files (txt-based analysis)
    result_files = sorted(BASE_DIR.glob("resultados_batch_txt_*.json"))
    if not result_files:
        print("ERROR: No result files found (resultados_batch_*.json).")
        print("Run the batch agents first, then re-run this script.")
        return

    print(f"Found {len(result_files)} result file(s).")

    # Build row→data mapping
    row_data: dict[int, dict] = {}
    rows_missing_pdf: list[int] = []

    for rf in result_files:
        results = json.loads(rf.read_text(encoding="utf-8"))
        for item in results:
            if item.get("status") == "not_found":
                rows_missing_pdf.extend(r["row"] for r in item.get("rows", []))
                continue
            for row_info in item.get("rows", []):
                row_idx = row_info["row"]
                row_data[row_idx] = {
                    "fields":    item.get("fields", {}),
                    "missing":   item.get("missing_from_no", ""),
                    "incorrect": item.get("possibly_incorrect", ""),
                    "n":         row_info.get("N", ""),
                    "o":         row_info.get("O", ""),
                    "pdf":       item.get("pdf_name", ""),
                    "country":   item.get("country", ""),
                    "error":     item.get("error", ""),
                }

    # Copy original → output
    print(f"Copying {INPUT_FILE.name} to {OUTPUT_FILE.name} ...")
    shutil.copy2(INPUT_FILE, OUTPUT_FILE)
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    ws = wb.active

    # Write header
    ws.cell(row=1, column=COL_NEW).value = NEW_HEADER
    ws.cell(row=1, column=COL_NEW).font = Font(bold=True)

    stats = {"written": 0, "errors": 0, "not_found": 0}
    report_lines = [
        "OPAS PDF ANALYSIS REPORT — Claude Agent SDK",
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "=" * 65, ""
    ]

    for row_idx, data in sorted(row_data.items()):
        cell = ws.cell(row=row_idx, column=COL_NEW)
        if data.get("error"):
            cell.value = f"[ANALYSIS ERROR: {data['error'][:300]}]"
            stats["errors"] += 1
            report_lines.append(
                f"ROW {row_idx} | {data['country']} | {data['pdf']}\n"
                f"  Status: Error — {data['error'][:100]}\n"
            )
        else:
            text = format_cell(
                data["fields"], data["n"], data["o"],
                data["missing"], data["incorrect"]
            )
            cell.value = text
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            stats["written"] += 1
            flags = []
            if data["missing"] and "none" not in data["missing"].lower():
                flags.append("MISSING")
            if data["incorrect"] and "no conflicts" not in data["incorrect"].lower():
                flags.append("POSSIBLY_INCORRECT")
            report_lines.append(
                f"ROW {row_idx} | {data['country']} | {data['pdf']}\n"
                f"  Status: Analyzed | Flags: {', '.join(flags) or 'none'}\n"
                f"  Primary diagnosis: {data['fields'].get('PRIMARY_DIAGNOSIS', 'NOT FOUND')[:120]}\n"
            )

    for row_idx in rows_missing_pdf:
        ws.cell(row=row_idx, column=COL_NEW).value = (
            "[PDF file not found on disk at the time of analysis]"
        )
        stats["not_found"] += 1

    wb.save(OUTPUT_FILE)

    # Report
    report_lines += [
        "", "=" * 65, "FINAL SUMMARY", "=" * 65,
        f"Rows written with full analysis: {stats['written']}",
        f"Rows with analysis errors:       {stats['errors']}",
        f"Rows where PDF was not found:    {stats['not_found']}",
        f"Total rows updated:              {stats['written'] + stats['errors'] + stats['not_found']}",
    ]
    REPORT_FILE.write_text("\n".join(report_lines), encoding="utf-8")

    print(f"\nSpreadsheet saved: {OUTPUT_FILE.name}")
    print(f"Report saved:      {REPORT_FILE.name}")
    print(f"\nRows with full analysis: {stats['written']}")
    print(f"Rows with errors:        {stats['errors']}")
    print(f"Rows - PDF not found:    {stats['not_found']}")

if __name__ == "__main__":
    main()
