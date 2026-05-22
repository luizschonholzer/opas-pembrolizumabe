#!/usr/bin/env python3
"""
OPAS Judicial Process Analysis using Claude Agent SDK
------------------------------------------------------
Reads PDFs linked in OPAS_PDF.xlsx, sends each PDF to the Claude API for
comprehensive clinical and judicial extraction, and writes results to a new
column in a copy of the spreadsheet.

Rules:
  - Columns N and O are NEVER modified (read-only for comparison)
  - Colored cells are NEVER touched (only new column AB is written)
  - No abbreviations in output
  - All new findings are clearly flagged
  - Output language: English (names/case numbers kept as-is)

Usage:
  1. Set your API key:
       $env:ANTHROPIC_API_KEY = "sk-ant-..."
     OR create a file named ".env" in this folder with:
       ANTHROPIC_API_KEY=sk-ant-...
  2. Run:
       python analise_com_claude_sdk.py
"""

import os
import sys
import base64
import time
import json
import shutil
from pathlib import Path
from datetime import datetime

import anthropic
import openpyxl
from openpyxl.styles import Alignment

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

BASE_DIR = Path(__file__).parent
INPUT_FILE  = BASE_DIR / "OPAS_PDF.xlsx"
OUTPUT_FILE = BASE_DIR / "OPAS_PDF_CLAUDE_ANALISE.xlsx"
REPORT_FILE = BASE_DIR / "RELATORIO_CLAUDE_SDK.txt"
ERROR_FILE  = BASE_DIR / "erros_claude_sdk.txt"
PROGRESS_FILE = BASE_DIR / "progresso_claude_sdk.json"

# Spreadsheet column indices (1-based)
COL_COUNTRY  = 1   # A
COL_CASE     = 6   # F
COL_DISEASE  = 14  # N
COL_DETAILS  = 15  # O
COL_PDF_LINK = 27  # AA
COL_NEW      = 28  # AB  ← new column written by this script

NEW_COL_HEADER = "Comprehensive Clinical and Judicial Analysis (Claude SDK)"

COUNTRY_FOLDER = {
    "Argentina":  "Argentina",
    "Brazil":     "Brasil",
    "Chile":      "Chile",
    "Colombia":   "Colombia",
    "Ecuador":    "Equador",
    "Guatemala":  "Guatemala",
    "Costa Rica": "Costa Rica",
    "Uruguay":    "Uruguai",
    "Paraguay":   "Paraguay",
    "Peru":       "Peru",
    "Bolivia":    "Bolivia",
    "Venezuela":  "Venezuela",
}

MAX_PDF_SIZE_BYTES = 30 * 1024 * 1024  # 30 MB — above this use pdfplumber text fallback
SLEEP_BETWEEN_CALLS = 2.0   # seconds between API calls
MAX_RETRIES = 3


# ---------------------------------------------------------------------------
# Extraction prompt
# ---------------------------------------------------------------------------

def build_prompt(n_value: str, o_value: str) -> str:
    context = ""
    if n_value.strip():
        context += f'\nCurrent value in column "Disease" (N): {n_value.strip()}'
    if o_value.strip():
        context += f'\nCurrent value in column "Further clinical details" (O): {o_value.strip()}'

    return f"""You are analyzing a judicial document related to Pembrolizumab from a Latin American court case.
{context}

Extract all information listed below from this document.
Write every answer in English. Do NOT use abbreviations — always write the full, unabbreviated form.
(Examples: write "Stage Four" not "St. IV", write "Microsatellite Instability-High" not "MSI-H",
write "ECOG Performance Status 2" not "ECOG 2", write "non-small cell lung carcinoma" not "NSCLC".)
Keep original proper names, medication names, institution names, and case numbers unchanged.
If a field is not present anywhere in the document, write exactly: NOT FOUND

After the 22 fields, add a COMPARISON section that notes:
  - What information from the PDF is MISSING from the current N/O column values shown above.
  - Any information in the current N/O values that appears POSSIBLY INCORRECT based on the PDF.
  If both columns are empty or the PDF provides much richer detail, note that as well.

Use this exact format for each field (one per line):

FIELD_01_PRIMARY_DIAGNOSIS: <full text>
FIELD_02_SECONDARY_DIAGNOSES: <full text>
FIELD_03_ICD_CODE: <full text>
FIELD_04_PATIENT_AGE: <full text>
FIELD_05_CLINICAL_HISTORY: <full text>
FIELD_06_SYMPTOMS: <full text>
FIELD_07_DISEASE_SEVERITY: <full text>
FIELD_08_FUNCTIONAL_LIMITATIONS: <full text>
FIELD_09_COMORBIDITIES: <full text>
FIELD_10_PRIOR_TREATMENTS: <full text>
FIELD_11_THERAPEUTIC_FAILURE: <full text>
FIELD_12_BIOMARKERS: <full text>
FIELD_13_METASTASIS_SITES: <full text>
FIELD_14_MEDICAL_JUSTIFICATION: <full text>
FIELD_15_URGENCY: <full text>
FIELD_16_RISK_DETERIORATION: <full text>
FIELD_17_RISK_HOSPITALIZATION: <full text>
FIELD_18_RISK_DEATH: <full text>
FIELD_19_REQUESTED_TREATMENT: <full text>
FIELD_20_JUDICIAL_DECISION: <Granted / Partially Granted / Denied / Other — specify>
FIELD_21_DECISION_GROUNDS: <full text — minimum 3 complete sentences>
FIELD_22_COURT_CITATIONS: <full text>
COMPARISON_MISSING_FROM_N_O: <list every piece of clinically relevant information found in the PDF that is absent from the current N/O values, or write "None — columns appear complete" if nothing is missing>
COMPARISON_POSSIBLY_INCORRECT: <list any value in the current N/O columns that appears to conflict with what the PDF says, including the current value and the correct value from the PDF, or write "No conflicts detected">

Field descriptions (read before answering):

FIELD_01_PRIMARY_DIAGNOSIS:
  Full name of the main oncological or medical diagnosis. Include cancer type, histological subtype,
  anatomical location, laterality, and disease stage — all written in full words.
  Example: "Stage Four metastatic non-small cell lung adenocarcinoma of the right upper lobe"

FIELD_02_SECONDARY_DIAGNOSES:
  All other diagnoses mentioned in the document, each written in full. Separate with semicolons.

FIELD_03_ICD_CODE:
  ICD-10 code if explicitly stated in the document. Include the code and its full description.
  Example: "C34.1 — Malignant neoplasm of upper lobe, bronchus or lung"

FIELD_04_PATIENT_AGE:
  Patient age written in full. Example: "64 years old"

FIELD_05_CLINICAL_HISTORY:
  Complete oncological history written in chronological order. Include: initial diagnosis date,
  all surgical procedures undergone, all prior treatment cycles, and relevant clinical events.

FIELD_06_SYMPTOMS:
  All symptoms described in the document, written in full. Separate with semicolons.

FIELD_07_DISEASE_SEVERITY:
  Full staging description (including tumor-node-metastasis classification if present),
  performance status written in full (e.g., "ECOG Performance Status 2 — ambulatory, capable of
  self-care but unable to carry out work activities"), and all other severity indicators.

FIELD_08_FUNCTIONAL_LIMITATIONS:
  Full description of any functional, physical, or daily-life impairments mentioned.

FIELD_09_COMORBIDITIES:
  All comorbid medical conditions, written in full. Separate with semicolons.

FIELD_10_PRIOR_TREATMENTS:
  Complete list of all prior treatments. For each treatment include:
  - Full generic drug name (not trade name unless only trade name is given)
  - Route and dose if mentioned
  - Duration or number of cycles if mentioned
  - Documented outcome or reason for discontinuation

FIELD_11_THERAPEUTIC_FAILURE:
  Full description of all documented evidence that prior treatments failed, including disease
  progression events, resistance development, or intolerable toxicity.

FIELD_12_BIOMARKERS:
  All biomarkers mentioned, each written in full. Include:
  - Programmed Death-Ligand 1 (PD-L1) expression percentage and assay used if stated
  - Microsatellite Instability status (write in full: "Microsatellite Instability-High" or
    "Microsatellite Stable" or "Mismatch Repair Deficient" or "Mismatch Repair Proficient")
  - Tumor Mutational Burden value if stated
  - Human Epidermal Growth Factor Receptor 2 (HER2) status
  - Epidermal Growth Factor Receptor (EGFR) mutation status
  - Anaplastic Lymphoma Kinase (ALK) rearrangement status
  - ROS Proto-Oncogene 1 (ROS1) rearrangement status
  - Kirsten Rat Sarcoma Viral Proto-Oncogene (KRAS) mutation status
  - B-Raf Proto-Oncogene (BRAF) mutation status
  - Any other biomarkers mentioned

FIELD_13_METASTASIS_SITES:
  Full list of all sites of metastatic disease mentioned. Separate with semicolons.

FIELD_14_MEDICAL_JUSTIFICATION:
  Complete medical justification provided by the treating physician, as described in the document.
  Quote relevant parts if present.

FIELD_15_URGENCY:
  Full description of any urgency, emergency, or time-sensitive indicators mentioned in the document.

FIELD_16_RISK_DETERIORATION:
  Full description of the risk of clinical deterioration or disease worsening as stated in the
  document.

FIELD_17_RISK_HOSPITALIZATION:
  Full description of the risk of hospitalization as stated in the document.

FIELD_18_RISK_DEATH:
  Full description of the risk of death or any life-threatening condition as stated in the
  document.

FIELD_19_REQUESTED_TREATMENT:
  Full name and all details of the medication, procedure, or treatment requested by the plaintiff.
  For Pembrolizumab: include dose, administration route, frequency, and line of treatment if stated.

FIELD_20_JUDICIAL_DECISION:
  The outcome of the judicial decision. Write exactly one of:
  Granted / Partially Granted / Denied / Other — (specify in full)

FIELD_21_DECISION_GROUNDS:
  Complete summary of the legal and clinical grounds on which the court based its decision.
  Write in full sentences. Minimum 3 sentences. Include whether the court applied the right to
  health, constitutional provisions, or technical/scientific arguments.

FIELD_22_COURT_CITATIONS:
  Full list of all legal articles, regulations, constitutional provisions, international treaties,
  scientific studies, or clinical guidelines explicitly cited by the court. Separate with semicolons.
"""


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------

def load_api_key() -> str:
    """Load Anthropic API key from environment or .env file."""
    key = os.environ.get("ANTHROPIC_API_KEY", "")
    if key:
        return key
    env_path = BASE_DIR / ".env"
    if env_path.exists():
        with open(env_path, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line.startswith("ANTHROPIC_API_KEY="):
                    return line.split("=", 1)[1].strip().strip('"').strip("'")
    return ""


def find_pdf(pdf_name: str, country: str) -> Path | None:
    """Find PDF file by name inside the country subfolder (recursive search)."""
    folder_name = COUNTRY_FOLDER.get(country, country)
    country_dir = BASE_DIR / folder_name
    if not country_dir.exists():
        # Try case-insensitive search
        for d in BASE_DIR.iterdir():
            if d.is_dir() and d.name.lower() == folder_name.lower():
                country_dir = d
                break
        else:
            return None

    pdf_name_clean = pdf_name.strip()
    for root, _, files in os.walk(country_dir):
        for fname in files:
            if fname.strip() == pdf_name_clean:
                return Path(root) / fname
    # Fallback: look in BASE_DIR directly
    direct = BASE_DIR / pdf_name_clean
    if direct.exists():
        return direct
    return None


def read_pdf_as_base64(pdf_path: Path) -> str:
    """Read a PDF file and return it as a base64-encoded string."""
    with open(pdf_path, "rb") as f:
        return base64.standard_b64encode(f.read()).decode("utf-8")


def extract_text_with_pdfplumber(pdf_path: Path) -> str:
    """Fallback: extract raw text from a large PDF using pdfplumber."""
    try:
        import pdfplumber
        texts = []
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages[:40]:
                t = page.extract_text()
                if t:
                    texts.append(t)
        return "\n\n".join(texts) if texts else "[No text extractable from PDF]"
    except Exception as e:
        return f"[pdfplumber error: {e}]"


def call_claude(client: anthropic.Anthropic, content_blocks: list, prompt: str) -> str:
    """Call Claude API with retry logic. Returns response text."""
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            response = client.messages.create(
                model="claude-opus-4-7",
                max_tokens=4096,
                messages=[{
                    "role": "user",
                    "content": content_blocks + [{"type": "text", "text": prompt}]
                }]
            )
            return response.content[0].text
        except anthropic.RateLimitError:
            wait = 30 * attempt
            print(f"    Rate limit hit — waiting {wait}s before retry {attempt}/{MAX_RETRIES}")
            time.sleep(wait)
        except anthropic.APIStatusError as e:
            if attempt == MAX_RETRIES:
                raise
            wait = 5 * attempt
            print(f"    API error ({e.status_code}) — retry {attempt}/{MAX_RETRIES} in {wait}s")
            time.sleep(wait)
        except Exception:
            if attempt == MAX_RETRIES:
                raise
            time.sleep(5 * attempt)
    raise RuntimeError("Max retries exceeded")


def parse_extraction(raw: str) -> dict:
    """Parse Claude's structured response into a dictionary."""
    fields = {}
    current_key = None
    current_lines = []

    for line in raw.splitlines():
        # Check if this line starts a new field
        matched = False
        for prefix in [
            "FIELD_01_", "FIELD_02_", "FIELD_03_", "FIELD_04_", "FIELD_05_",
            "FIELD_06_", "FIELD_07_", "FIELD_08_", "FIELD_09_", "FIELD_10_",
            "FIELD_11_", "FIELD_12_", "FIELD_13_", "FIELD_14_", "FIELD_15_",
            "FIELD_16_", "FIELD_17_", "FIELD_18_", "FIELD_19_", "FIELD_20_",
            "FIELD_21_", "FIELD_22_",
            "COMPARISON_MISSING_FROM_N_O:",
            "COMPARISON_POSSIBLY_INCORRECT:",
        ]:
            if line.startswith(prefix) or (prefix.endswith(":") and line.startswith(prefix[:-1])):
                if current_key:
                    fields[current_key] = " ".join(current_lines).strip()
                if ":" in line:
                    key_part, _, val_part = line.partition(":")
                    current_key = key_part.strip()
                    current_lines = [val_part.strip()]
                else:
                    current_key = line.strip()
                    current_lines = []
                matched = True
                break
        if not matched and current_key:
            current_lines.append(line.strip())

    if current_key:
        fields[current_key] = " ".join(current_lines).strip()

    return fields


# Map field keys to human-readable labels (no abbreviations)
FIELD_LABELS = {
    "FIELD_01_PRIMARY_DIAGNOSIS":      "01. Primary Diagnosis",
    "FIELD_02_SECONDARY_DIAGNOSES":    "02. Secondary Diagnoses",
    "FIELD_03_ICD_CODE":               "03. ICD-10 Code",
    "FIELD_04_PATIENT_AGE":            "04. Patient Age",
    "FIELD_05_CLINICAL_HISTORY":       "05. Clinical History",
    "FIELD_06_SYMPTOMS":               "06. Symptoms",
    "FIELD_07_DISEASE_SEVERITY":       "07. Disease Severity and Performance Status",
    "FIELD_08_FUNCTIONAL_LIMITATIONS": "08. Functional Limitations",
    "FIELD_09_COMORBIDITIES":          "09. Comorbidities",
    "FIELD_10_PRIOR_TREATMENTS":       "10. Prior Treatments",
    "FIELD_11_THERAPEUTIC_FAILURE":    "11. Therapeutic Failure",
    "FIELD_12_BIOMARKERS":             "12. Biomarkers",
    "FIELD_13_METASTASIS_SITES":       "13. Metastasis Sites",
    "FIELD_14_MEDICAL_JUSTIFICATION":  "14. Medical Justification",
    "FIELD_15_URGENCY":                "15. Clinical Urgency",
    "FIELD_16_RISK_DETERIORATION":     "16. Risk of Clinical Deterioration",
    "FIELD_17_RISK_HOSPITALIZATION":   "17. Risk of Hospitalization",
    "FIELD_18_RISK_DEATH":             "18. Risk of Death",
    "FIELD_19_REQUESTED_TREATMENT":    "19. Requested Treatment",
    "FIELD_20_JUDICIAL_DECISION":      "20. Judicial Decision",
    "FIELD_21_DECISION_GROUNDS":       "21. Decision Grounds",
    "FIELD_22_COURT_CITATIONS":        "22. Court Citations",
    "COMPARISON_MISSING_FROM_N_O":     "MISSING FROM COLUMNS N/O",
    "COMPARISON_POSSIBLY_INCORRECT":   "POSSIBLY INCORRECT IN COLUMNS N/O",
}


def format_cell_value(fields: dict, raw_response: str) -> tuple[str, list]:
    """
    Build the cell text and a list of flag descriptions.
    Returns (cell_text, flags).
    """
    lines = ["=== COMPREHENSIVE CLINICAL AND JUDICIAL ANALYSIS (Claude SDK) ===", ""]

    flags = []

    for key, label in FIELD_LABELS.items():
        value = fields.get(key, "NOT FOUND").strip() or "NOT FOUND"

        if key == "COMPARISON_MISSING_FROM_N_O":
            lines.append("")
            lines.append("--- COMPARISON WITH EXISTING DATA ---")
            if value and value.upper() != "NOT FOUND" and "none" not in value.lower():
                lines.append(f"[MISSING FROM N/O] {label}:")
                lines.append(f"  {value}")
                flags.append(f"MISSING: {value[:80]}")
            else:
                lines.append(f"{label}: {value}")

        elif key == "COMPARISON_POSSIBLY_INCORRECT":
            if value and value.upper() != "NOT FOUND" and "no conflicts" not in value.lower():
                lines.append(f"[POSSIBLY INCORRECT IN N/O] {label}:")
                lines.append(f"  {value}")
                flags.append(f"INCORRECT: {value[:80]}")
            else:
                lines.append(f"{label}: {value}")

        else:
            lines.append(f"{label}: {value}")

    lines.append("")
    lines.append(f"[Analyzed by Claude claude-opus-4-7 on {datetime.now().strftime('%Y-%m-%d')}]")

    return "\n".join(lines), flags


# ---------------------------------------------------------------------------
# Progress persistence
# ---------------------------------------------------------------------------

def load_progress() -> dict:
    if PROGRESS_FILE.exists():
        with open(PROGRESS_FILE, encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_progress(progress: dict) -> None:
    with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
        json.dump(progress, f, ensure_ascii=False, indent=2)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    print("=" * 65)
    print("  OPAS PDF Analysis — Claude Agent SDK")
    print("=" * 65)

    # --- API key ---
    api_key = load_api_key()
    if not api_key:
        print("\nERROR: ANTHROPIC_API_KEY not found.")
        print("Set it before running:")
        print("  PowerShell:  $env:ANTHROPIC_API_KEY = 'sk-ant-...'")
        print("  Or create a file named .env in this folder:")
        print("    ANTHROPIC_API_KEY=sk-ant-...")
        sys.exit(1)

    client = anthropic.Anthropic(api_key=api_key)
    print(f"API key loaded. Model: claude-opus-4-7\n")

    # --- Prepare output file ---
    print(f"Copying {INPUT_FILE.name} → {OUTPUT_FILE.name} ...")
    shutil.copy2(INPUT_FILE, OUTPUT_FILE)

    wb = openpyxl.load_workbook(OUTPUT_FILE)
    ws = wb.active

    # Write new column header
    header_cell = ws.cell(row=1, column=COL_NEW)
    header_cell.value = NEW_COL_HEADER

    # --- Load progress (for resumability) ---
    progress = load_progress()
    resumed = len(progress)
    if resumed:
        print(f"Resuming from previous run — {resumed} rows already processed.\n")

    # --- Stats ---
    stats = {
        "total": 0, "analyzed": 0, "no_pdf": 0,
        "pdf_not_found": 0, "api_errors": 0, "flags_added": 0,
    }
    report_lines: list[str] = []
    error_lines:  list[str] = []

    report_lines.append("OPAS PDF ANALYSIS REPORT — Claude Agent SDK")
    report_lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    report_lines.append("=" * 65)
    report_lines.append("")

    # --- Row iteration ---
    for row_idx in range(2, ws.max_row + 1):
        country_raw = ws.cell(row=row_idx, column=COL_COUNTRY).value
        if not country_raw or not str(country_raw).strip():
            break  # End of data

        country  = str(country_raw).strip()
        case_num = ws.cell(row=row_idx, column=COL_CASE).value or ""
        pdf_link = ws.cell(row=row_idx, column=COL_PDF_LINK).value or ""
        n_value  = str(ws.cell(row=row_idx, column=COL_DISEASE).value or "").strip()
        o_value  = str(ws.cell(row=row_idx, column=COL_DETAILS).value or "").strip()

        stats["total"] += 1
        row_key = str(row_idx)

        # Skip already processed
        if row_key in progress and progress[row_key] == "done":
            print(f"  Row {row_idx:>3} | {country:<12} | {str(case_num)[:30]:<30} | [ALREADY DONE — skipped]")
            continue

        print(f"  Row {row_idx:>3} | {country:<12} | {str(case_num)[:30]:<30}", end=" | ", flush=True)

        # ---- No PDF ----
        pdf_link_str = str(pdf_link).strip()
        if not pdf_link_str:
            stats["no_pdf"] += 1
            progress[row_key] = "no_pdf"
            report_lines.append(
                f"ROW {row_idx} | Country: {country} | Case: {case_num}\n"
                f"  Status: No PDF available — row skipped\n"
            )
            print("No PDF")
            continue

        # ---- Locate PDF ----
        pdf_path = find_pdf(pdf_link_str, country)
        if not pdf_path:
            stats["pdf_not_found"] += 1
            progress[row_key] = "not_found"
            ws.cell(row=row_idx, column=COL_NEW).value = (
                f"[ERROR: PDF file '{pdf_link_str}' was not found on disk. "
                f"Please verify the file path in column AA.]"
            )
            report_lines.append(
                f"ROW {row_idx} | Country: {country} | Case: {case_num} | PDF: {pdf_link_str}\n"
                f"  Status: PDF file not found on disk\n"
            )
            print(f"PDF NOT FOUND: {pdf_link_str}")
            continue

        # ---- Call Claude ----
        pdf_size = os.path.getsize(pdf_path)
        prompt   = build_prompt(n_value, o_value)

        try:
            if pdf_size > MAX_PDF_SIZE_BYTES:
                # Large file: extract text with pdfplumber, send as text
                print(f"large PDF ({pdf_size/1024/1024:.1f} MB) — using text fallback", end=" | ", flush=True)
                text = extract_text_with_pdfplumber(pdf_path)
                content_blocks = [{"type": "text", "text": f"Document text:\n\n{text}"}]
            else:
                pdf_b64 = read_pdf_as_base64(pdf_path)
                content_blocks = [{
                    "type": "document",
                    "source": {
                        "type": "base64",
                        "media_type": "application/pdf",
                        "data": pdf_b64,
                    }
                }]

            raw_response = call_claude(client, content_blocks, prompt)
            fields       = parse_extraction(raw_response)
            cell_text, flags = format_cell_value(fields, raw_response)

            # Write to new column only
            new_cell = ws.cell(row=row_idx, column=COL_NEW)
            new_cell.value = cell_text
            new_cell.alignment = Alignment(wrap_text=True, vertical="top")

            stats["analyzed"] += 1
            if flags:
                stats["flags_added"] += 1

            progress[row_key] = "done"

            flag_summary = "; ".join(flags[:3]) if flags else "No flags"
            report_lines.append(
                f"ROW {row_idx} | Country: {country} | Case: {case_num} | PDF: {pdf_link_str}\n"
                f"  Status: Analyzed successfully\n"
                f"  Fields extracted: {sum(1 for v in fields.values() if v.upper() != 'NOT FOUND')}/22\n"
                f"  Flags: {flag_summary}\n"
            )
            print(f"OK — {len(flags)} flag(s)")

        except Exception as exc:
            stats["api_errors"] += 1
            progress[row_key] = "error"
            err_msg = f"ROW {row_idx} | {type(exc).__name__}: {exc}"
            error_lines.append(err_msg)
            ws.cell(row=row_idx, column=COL_NEW).value = (
                f"[ANALYSIS ERROR: {type(exc).__name__} — {str(exc)[:300]}. "
                f"Re-run script to retry this row.]"
            )
            report_lines.append(
                f"ROW {row_idx} | Country: {country} | Case: {case_num} | PDF: {pdf_link_str}\n"
                f"  Status: API Error — {type(exc).__name__}\n"
            )
            print(f"ERROR: {exc}")

        # Save periodically
        if stats["total"] % 5 == 0:
            wb.save(OUTPUT_FILE)
            save_progress(progress)

        time.sleep(SLEEP_BETWEEN_CALLS)

    # --- Final save ---
    wb.save(OUTPUT_FILE)
    save_progress(progress)
    print(f"\nSpreadsheet saved: {OUTPUT_FILE}")

    # --- Write report ---
    report_lines.append("")
    report_lines.append("=" * 65)
    report_lines.append("FINAL SUMMARY")
    report_lines.append("=" * 65)
    report_lines.append(f"Total rows in spreadsheet:        {stats['total']}")
    report_lines.append(f"Rows analyzed (PDF read by Claude): {stats['analyzed']}")
    report_lines.append(f"Rows skipped (no PDF link):        {stats['no_pdf']}")
    report_lines.append(f"Rows where PDF was not found:      {stats['pdf_not_found']}")
    report_lines.append(f"Rows with API errors:              {stats['api_errors']}")
    report_lines.append(f"Rows with new flags (missing/incorrect data): {stats['flags_added']}")

    with open(REPORT_FILE, "w", encoding="utf-8") as f:
        f.write("\n".join(report_lines))
    print(f"Report saved:      {REPORT_FILE}")

    if error_lines:
        with open(ERROR_FILE, "w", encoding="utf-8") as f:
            f.write("\n".join(error_lines))
        print(f"Error log saved:   {ERROR_FILE}")

    print("\n=== SUMMARY ===")
    print(f"  Total rows:          {stats['total']}")
    print(f"  Analyzed:            {stats['analyzed']}")
    print(f"  No PDF:              {stats['no_pdf']}")
    print(f"  PDF not found:       {stats['pdf_not_found']}")
    print(f"  API errors:          {stats['api_errors']}")
    print(f"  Rows with flags:     {stats['flags_added']}")
    print(f"\nOutput file: {OUTPUT_FILE.name}")
    print(f"Report:      {REPORT_FILE.name}")
    print("Done.")


if __name__ == "__main__":
    main()
