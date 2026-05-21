#!/usr/bin/env python3
"""
Comprehensive PDF clinical data analysis script.
Reads PDFs linked in the spreadsheet, compares with existing columns N/O,
fills column AB where empty, and generates a detailed report.
"""

import os
import re
import sys
import traceback
from datetime import datetime
import pdfplumber
import openpyxl

# ─── Configuration ───────────────────────────────────────────────────────────
BASE_DIR = r"C:\Users\Claudio\Downloads\Processos - OPAS"
INPUT_XLSX = os.path.join(BASE_DIR, "OPAS_com_links_PDF_BACKUP_20260520_1713.xlsx")
OUTPUT_XLSX = os.path.join(BASE_DIR, "OPAS_ANALISE_ATUALIZADA_20260521.xlsx")
OUTPUT_REPORT = os.path.join(BASE_DIR, "RELATORIO_ANALISE_PDFS_20260521.txt")

# Column indices (1-based)
COL_COUNTRY = 1
COL_CASE = 6
COL_DISEASE = 14       # N
COL_DETAILS = 15       # O
COL_PDF_LINK = 27      # AA
COL_AB = 28            # AB
COL_AC = 29            # AC (overflow if needed)

# Known PDF mismatches from preliminary review
KNOWN_MISMATCHES = {
    3:  "Row 3 (Case 013326/2024): PDF 'CCF 000104_2024.pdf' belongs to case CCF 000104/2024 (row 8)",
    4:  "Row 4 (Case CCF 011768/2025): PDF '14838_2025.pdf' belongs to case 14838/2025 (row 2)",
    6:  "Row 6 (Case CCF 016494/2025): PDF '14838_2025.pdf' belongs to case 14838/2025 (row 2)",
    11: "Row 11 (Case CCF 000453/2016/CA001): PDF 'CCF 004030_2020_CA001.pdf' belongs to case CCF 004030/2020/CA001 (row 10)",
    128:"Row 128 (Case 7104-2025): PDF '1343-2025.pdf' belongs to case 1343-2025 (row 122)",
}

# ─── Build PDF path map ───────────────────────────────────────────────────────
def build_pdf_map(base_dir):
    """Scan all subdirectories and map filename -> full path."""
    pdf_map = {}
    for root, dirs, files in os.walk(base_dir):
        for f in files:
            if f.lower().endswith('.pdf') or f.lower().endswith('.docx'):
                pdf_map[f] = os.path.join(root, f)
    return pdf_map

# ─── PDF text extraction ──────────────────────────────────────────────────────
def extract_pdf_text(pdf_path):
    """Extract all text from a PDF file."""
    try:
        text_pages = []
        with pdfplumber.open(pdf_path) as pdf:
            for i, page in enumerate(pdf.pages):
                t = page.extract_text()
                if t:
                    text_pages.append(f"[PAGE {i+1}]\n{t}")
        return "\n".join(text_pages)
    except Exception as e:
        return f"[ERROR extracting PDF: {e}]"

def extract_docx_text(docx_path):
    """Extract text from a .docx file."""
    try:
        from docx import Document
        doc = Document(docx_path)
        return "\n".join([p.text for p in doc.paragraphs])
    except Exception as e:
        return f"[ERROR extracting DOCX: {e}]"

def extract_text(file_path):
    """Extract text from PDF or DOCX."""
    if file_path.lower().endswith('.docx'):
        return extract_docx_text(file_path)
    return extract_pdf_text(file_path)

# ─── Clinical data extraction ─────────────────────────────────────────────────
def find_case_section(full_text, case_number, country):
    """
    For multi-case PDFs (especially Uruguay), find the section relevant to
    the specific case number.
    """
    if not case_number:
        return full_text

    # Clean case number for searching
    case_str = str(case_number).strip()

    # Try to find the case number in the text
    # Uruguay cases look like "171/2025", "47/2020" etc.
    # Extract just the numeric part before any slash
    match = re.match(r'^(\d+)[/\-](\d+)', case_str)
    if match:
        num_part = match.group(1)
        year_part = match.group(2)

        # Search for the case section
        patterns = [
            rf'\b{re.escape(case_str)}\b',
            rf'\b{num_part}/{year_part}\b',
            rf'\b{num_part}-{year_part}\b',
            rf'[Nn]o\.?\s*{num_part}',
            rf'[Ee]xp\.?\s*{num_part}',
        ]

        for pat in patterns:
            matches = list(re.finditer(pat, full_text))
            if matches:
                # Take the first match and get surrounding context
                start = max(0, matches[0].start() - 200)
                # Try to find the end of this case's section
                # Look for the next case number or end of document
                end = min(len(full_text), matches[0].start() + 5000)
                return full_text[start:end]

    return full_text

def extract_clinical_data(text, case_number=None, country=None):
    """
    Extract key clinical findings from PDF text.
    Returns a dict with clinical data fields.
    """
    text_lower = text.lower()
    findings = {}

    # ── Patient age ──────────────────────────────────────────────────────────
    age_patterns = [
        r'(\d{1,2})[- ]?(?:year[s]?[- ]?old|año[s]?|anos?)',
        r'(?:age|edad|idade)[:\s]+(\d{1,2})',
        r'(?:patient|paciente)[^.]{0,30}(\d{1,2})[- ]?(?:year|año|ano)',
        r'(\d{1,2})[- ]?(?:años|years)[- ]?(?:de edad)?',
    ]
    for pat in age_patterns:
        m = re.search(pat, text_lower)
        if m:
            age = int(m.group(1))
            if 18 <= age <= 100:
                findings['age'] = f"{age}-year-old patient"
                break

    # ── Disease staging ──────────────────────────────────────────────────────
    stage_patterns = [
        r'(?:stage|estadio|estadío|etapa)[:\s]+([IiVv]{1,4}[aAbBcC]?(?:\s*[A-C])?)',
        r'(?:EC|EC\.?|clinical stage|stage)[:\s]*([IiVv]{1,4}[aAbBcC]?)',
        r'\b(stage\s+[IiVv]{1,4}[aAbBcC]?)\b',
        r'\b(estadio\s+[IiVv]{1,4}[aAbBcC]?)\b',
        r'\b([IiVv]{1,4}[aAbBcC]?)\s*(?:stage|estadio)\b',
        # FIGO staging
        r'(?:FIGO|figo)[:\s]+([IiVv]{1,4}[A-C]?\d?)',
    ]
    for pat in stage_patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            stage_raw = m.group(1).strip()
            stage_upper = stage_raw.upper()
            if any(c in stage_upper for c in ['I', 'V']):
                findings['stage'] = f"Stage {stage_upper}"
                break

    # ── PD-L1 expression ─────────────────────────────────────────────────────
    pdl1_patterns = [
        r'PD-?L1[^.]{0,80}?(\d+(?:\.\d+)?)\s*%',
        r'(\d+(?:\.\d+)?)\s*%[^.]{0,50}?PD-?L1',
        r'PD-?L1[^.]{0,50}(?:expression|expressão|expresión)[^.]{0,30}(\d+)',
        r'CPS[:\s]+(\d+(?:\.\d+)?)',
        r'TPS[:\s]+(\d+(?:\.\d+)?)',
        r'PD-?L1[^.]{0,100}?(?:positive|negat|alto|high|low|baixo|elevado)',
    ]
    for pat in pdl1_patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            if m.lastindex and m.group(1):
                findings['pdl1'] = f"PD-L1: {m.group(1)}%"
            else:
                context = m.group(0)
                findings['pdl1'] = f"PD-L1: {context[:60].strip()}"
            break

    # ── MSI status ───────────────────────────────────────────────────────────
    msi_patterns = [
        r'MSI[- ]?H\b',
        r'MSI[- ]?L\b',
        r'(?:high|alta)\s*microsatellite[^.]{0,30}instability',
        r'microsatellite[^.]{0,30}(?:instab|stable)',
        r'deficient\s*(?:MMR|mismatch)',
        r'dMMR\b',
        r'MSS\b',
        r'MMR[^.]{0,50}(?:deficient|proficient)',
        r'inestabilidad\s*(?:de\s*)?microsatélites',
        r'instabilidade\s*(?:de\s*)?microssatélites',
    ]
    for pat in msi_patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            findings['msi'] = f"Microsatellite status: {m.group(0).strip()}"
            break

    # ── ECOG performance status ──────────────────────────────────────────────
    ecog_patterns = [
        r'ECOG[:\s]+(\d)',
        r'performance\s+status[:\s]+(\d)',
        r'ECOG[^.]{0,20}(\d)',
        r'PS[:\s]+(\d)',
    ]
    for pat in ecog_patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            ecog_val = m.group(1)
            findings['ecog'] = f"ECOG PS: {ecog_val}"
            break

    # ── Prior treatments ─────────────────────────────────────────────────────
    prior_treatments = []

    # Chemotherapy
    chemo_patterns = [
        r'(?:chemotherapy|quimioterapia|quimioteraphy|kemoterapi)[^.]{0,100}',
        r'(?:FOLFOX|FOLFIRI|CAPOX|XELOX|FLOT|FOLFIRINOX)[^.]{0,50}',
        r'(?:cisplatin|carboplatin|oxaliplatin|paclitaxel|docetaxel|gemcitabine)[^.]{0,60}',
        r'(?:bevacizumab|cetuximab|trastuzumab|pertuzumab|nivolumab|atezolizumab)[^.]{0,60}',
    ]
    chemo_found = []
    for pat in chemo_patterns:
        matches = re.findall(pat, text, re.IGNORECASE)
        if matches:
            for m in matches[:2]:
                m_clean = m.strip()[:80]
                if m_clean and m_clean not in chemo_found:
                    chemo_found.append(m_clean)
    if chemo_found:
        prior_treatments.append("Chemotherapy: " + "; ".join(chemo_found[:3]))

    # Surgery
    surgery_patterns = [
        r'(?:surgery|surgical|cirurgia|cirurgía|operaci[oó]n|resección|resecção|nephrectomy|mastectomy|hysterectomy|colectomy|gastrectomy)[^.]{0,60}',
    ]
    surgery_found = []
    for pat in surgery_patterns:
        matches = re.findall(pat, text, re.IGNORECASE)
        if matches:
            for m in matches[:2]:
                m_clean = m.strip()[:80]
                if m_clean:
                    surgery_found.append(m_clean)
    if surgery_found:
        prior_treatments.append("Surgery: " + "; ".join(surgery_found[:2]))

    # Radiotherapy
    radio_patterns = [
        r'(?:radiotherapy|radiotherapy|radioterapia|radiation therapy|RT\b)[^.]{0,60}',
    ]
    radio_found = []
    for pat in radio_patterns:
        matches = re.findall(pat, text, re.IGNORECASE)
        if matches:
            for m in matches[:1]:
                radio_found.append(m.strip()[:60])
    if radio_found:
        prior_treatments.append("Radiotherapy: " + "; ".join(radio_found[:1]))

    if prior_treatments:
        findings['prior_treatments'] = " | ".join(prior_treatments)

    # ── Metastasis sites ─────────────────────────────────────────────────────
    mets_patterns = [
        r'metastas[ie][s]?[^.]{0,100}',
        r'(?:liver|hepatic|lung|pulmonary|bone|brain|lymph node|peritoneal|adrenal)[^.]{0,30}metastas',
        r'metastas[^.]{0,30}(?:liver|hepatic|lung|pulmonary|bone|brain|lymph node)',
    ]
    mets_found = []
    for pat in mets_patterns:
        matches = re.findall(pat, text, re.IGNORECASE)
        for m in matches[:3]:
            m_clean = m.strip()[:80]
            if m_clean and m_clean not in mets_found:
                mets_found.append(m_clean)
    if mets_found:
        findings['metastasis'] = "Metastasis: " + "; ".join(mets_found[:2])

    # ── Biomarkers: EGFR, ALK, KRAS, HER2, BRAF ─────────────────────────────
    biomarker_patterns = [
        r'EGFR[^.]{0,60}',
        r'ALK[^.]{0,40}',
        r'KRAS[^.]{0,40}',
        r'HER2[^.]{0,60}',
        r'BRAF[^.]{0,40}',
        r'ROS1[^.]{0,40}',
        r'TMB[^.]{0,40}',
    ]
    biomarkers = []
    for pat in biomarker_patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            clean = m.group(0).strip()[:60]
            if clean:
                biomarkers.append(clean)
    if biomarkers:
        findings['biomarkers'] = "Biomarkers: " + " | ".join(biomarkers[:4])

    # ── Treatment context ─────────────────────────────────────────────────────
    context_terms = {
        'neoadjuvant': r'neoadjuvant|neoadyuvante|neoadjuvante',
        'adjuvant': r'\badjuvant\b|adyuvante|adjuvante',
        'palliative': r'palliative|paliativo|paliativa',
        'first-line': r'first[- ]line|primera[- ]línea|primeira[- ]linha|1[ªa][- ]línea',
        'second-line': r'second[- ]line|segunda[- ]línea|segunda[- ]linha|2[ªa][- ]línea',
        'third-line': r'third[- ]line|tercera[- ]línea|terceira[- ]linha',
    }
    found_contexts = []
    for label, pat in context_terms.items():
        if re.search(pat, text, re.IGNORECASE):
            found_contexts.append(label)
    if found_contexts:
        findings['treatment_context'] = "Treatment context: " + ", ".join(found_contexts)

    # ── Clinical urgency/severity ─────────────────────────────────────────────
    urgency_patterns = [
        r'(?:urgent|urgente|urgência|urgencia|critically ill|critical condition)[^.]{0,60}',
        r'(?:progressive|progressing|progression|progresión|progressão)[^.]{0,60}',
        r'(?:terminal|end[- ]stage|advanced|unresectable|inoperable)[^.]{0,60}',
    ]
    urgency_found = []
    for pat in urgency_patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            urgency_found.append(m.group(0).strip()[:80])
    if urgency_found:
        findings['urgency'] = "Clinical urgency/severity: " + "; ".join(urgency_found[:2])

    return findings

def format_ab_entry(findings, existing_ab=""):
    """Format the clinical findings into the pipe-separated format for column AB."""
    parts = []

    order = ['age', 'stage', 'ecog', 'pdl1', 'msi', 'biomarkers', 'treatment_context',
             'prior_treatments', 'metastasis', 'urgency']

    labels = {
        'age': 'Patient age',
        'stage': 'Disease staging',
        'ecog': 'ECOG performance status',
        'pdl1': None,  # already has label
        'msi': None,   # already has label
        'biomarkers': None,
        'treatment_context': None,
        'prior_treatments': 'Prior treatments',
        'metastasis': None,
        'urgency': None,
    }

    for key in order:
        if key in findings:
            val = findings[key]
            label = labels.get(key)
            if label and not val.startswith(label):
                parts.append(f"{label}: {val}")
            else:
                parts.append(val)

    return " | ".join(parts)

# ─── Main analysis ────────────────────────────────────────────────────────────
def check_disease_match(pdf_text, disease_col_n, details_col_o):
    """Check if the PDF content confirms the disease listed in column N."""
    issues = []

    if not disease_col_n:
        return issues

    disease_lower = str(disease_col_n).lower()
    pdf_lower = pdf_text.lower()

    # Skip procedural/NI cases
    if any(term in disease_lower for term in ['procedural', 'not about', 'ni', 'nao']):
        return issues

    # Key disease terms to check
    disease_keywords = {
        'lung cancer': ['lung', 'pulmón', 'pulmão', 'pulmonar', 'pulmonary', 'nsclc', 'adenocarcinoma'],
        'breast cancer': ['breast', 'mama', 'mamario', 'mamaria'],
        'cervical cancer': ['cervical', 'cervix', 'cuello', 'colo do útero', 'cérvix'],
        'renal cancer': ['renal', 'kidney', 'riñón', 'rim', 'células claras', 'clear cell'],
        'melanoma': ['melanoma'],
        'lymphoma': ['lymphoma', 'linfoma'],
        'colorectal': ['colorectal', 'colon', 'rectal', 'cólon'],
        'endometrial': ['endometrial', 'endometrio', 'uterine', 'uterino'],
        'bladder': ['bladder', 'vejiga', 'bexiga', 'urothelial', 'urotelial'],
        'gastric': ['gastric', 'gástrico', 'estomago', 'estômago'],
    }

    for disease_key, keywords in disease_keywords.items():
        if disease_key in disease_lower:
            found = any(kw in pdf_lower for kw in keywords)
            if not found:
                issues.append(f"WARNING: Disease '{disease_col_n}' not clearly confirmed in PDF text")
            break

    return issues

def analyze_row(row_num, case_num, country, disease, details, pdf_filename, existing_ab, pdf_map):
    """
    Analyze a single row.
    Returns: (new_ab, new_details, issues, notes)
    """
    issues = []
    notes = []
    new_ab = existing_ab or ""
    new_details = details or ""

    # Check for known mismatches
    if row_num in KNOWN_MISMATCHES:
        issues.append(f"KNOWN PDF MISMATCH: {KNOWN_MISMATCHES[row_num]}")

    if not pdf_filename:
        return new_ab, new_details, issues, notes

    # Find file path
    pdf_filename = pdf_filename.strip()
    if pdf_filename not in pdf_map:
        issues.append(f"ERROR: File not found on disk: '{pdf_filename}'")
        return new_ab, new_details, issues, notes

    file_path = pdf_map[pdf_filename]

    # Extract text
    full_text = extract_text(file_path)
    if full_text.startswith('[ERROR'):
        issues.append(full_text)
        return new_ab, new_details, issues, notes

    if not full_text.strip():
        issues.append(f"WARNING: Empty text extracted from '{pdf_filename}'")
        return new_ab, new_details, issues, notes

    # For multi-case PDFs, try to find the relevant section
    pdf_text = find_case_section(full_text, case_num, country)

    # Check disease match (only for non-mismatch rows)
    if row_num not in KNOWN_MISMATCHES:
        disease_issues = check_disease_match(pdf_text, disease, details)
        issues.extend(disease_issues)

    # Extract clinical data if AB is empty or needs enrichment
    if not existing_ab or existing_ab.strip() == "":
        # Skip procedural/NI cases
        disease_str = str(disease or "").lower()
        if any(term in disease_str for term in ['procedural', 'not about pembrolizumab']):
            notes.append(f"Classified as '{disease}' - checking PDF to verify")
            # Still extract to verify classification
            clinical = extract_clinical_data(pdf_text, case_num, country)
            if clinical:
                formatted = format_ab_entry(clinical)
                if formatted:
                    new_ab = f"[PDF checked - classified as {disease}] " + formatted
                    notes.append(f"AB filled with extracted data despite '{disease}' classification")
        elif disease_str == 'ni' or not disease_str:
            notes.append(f"Disease = 'NI' - extracting from PDF")
            clinical = extract_clinical_data(pdf_text, case_num, country)
            if clinical:
                formatted = format_ab_entry(clinical)
                if formatted:
                    new_ab = formatted
                    notes.append(f"AB filled from PDF (was NI/empty)")
        else:
            # Normal case - extract and fill AB
            clinical = extract_clinical_data(pdf_text, case_num, country)
            if clinical:
                formatted = format_ab_entry(clinical)
                if formatted:
                    new_ab = formatted
                    notes.append(f"AB filled from PDF")
            else:
                notes.append(f"Could not extract clinical data from PDF")
    else:
        notes.append(f"AB already has data - skipping extraction")

    return new_ab, new_details, issues, notes

# ─── Run analysis ─────────────────────────────────────────────────────────────
def main():
    print("=" * 70)
    print("COMPREHENSIVE PDF CLINICAL DATA ANALYSIS")
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)

    # Build PDF map
    print("\nScanning PDF files...")
    pdf_map = build_pdf_map(BASE_DIR)
    print(f"Found {len(pdf_map)} files (PDF/DOCX)")

    # Load spreadsheet
    print(f"\nLoading spreadsheet: {INPUT_XLSX}")
    wb = openpyxl.load_workbook(INPUT_XLSX)
    ws = wb.active

    # Collect data rows
    data_rows = []
    for r in range(2, ws.max_row + 1):
        case_num = ws.cell(r, COL_CASE).value
        if case_num:
            data_rows.append(r)
    print(f"Found {len(data_rows)} data rows")

    # ── Analysis results ──────────────────────────────────────────────────────
    results = {
        'filled_ab': [],          # rows where AB was empty and was filled
        'disease_warnings': [],   # rows where N/O might be wrong
        'pdf_mismatches': [],     # rows with PDF mismatch
        'errors': [],             # rows with extraction errors
        'skipped': [],            # rows without PDF link
        'not_about_pembro': [],   # procedural/not-about rows
        'all_rows': []            # all processed rows
    }

    rows_with_pdf = 0
    rows_ab_filled = 0

    for r in data_rows:
        country = ws.cell(r, COL_COUNTRY).value or ""
        case_num = ws.cell(r, COL_CASE).value or ""
        disease = ws.cell(r, COL_DISEASE).value or ""
        details = ws.cell(r, COL_DETAILS).value or ""
        pdf_link = ws.cell(r, COL_PDF_LINK).value
        existing_ab = ws.cell(r, COL_AB).value or ""

        row_summary = {
            'row': r,
            'country': country,
            'case': str(case_num),
            'disease': str(disease),
            'details': str(details)[:100],
            'pdf_link': pdf_link,
            'existing_ab': existing_ab,
            'new_ab': existing_ab,
            'issues': [],
            'notes': [],
            'ab_was_empty': not existing_ab,
            'ab_filled': False,
        }

        if not pdf_link:
            row_summary['notes'].append("No PDF link")
            results['skipped'].append(r)
            results['all_rows'].append(row_summary)
            continue

        rows_with_pdf += 1

        new_ab, new_details, issues, notes = analyze_row(
            r, case_num, country, disease, details,
            pdf_link, existing_ab, pdf_map
        )

        row_summary['new_ab'] = new_ab
        row_summary['issues'] = issues
        row_summary['notes'] = notes

        # Track if AB was filled
        if not existing_ab and new_ab:
            row_summary['ab_filled'] = True
            rows_ab_filled += 1
            results['filled_ab'].append(r)

        # Categorize issues
        if any('MISMATCH' in i for i in issues):
            results['pdf_mismatches'].append(r)
        if any('WARNING' in i for i in issues):
            results['disease_warnings'].append(r)
        if any('ERROR' in i for i in issues):
            results['errors'].append(r)

        # Track not-about-pembrolizumab / procedural
        disease_lower = str(disease).lower()
        if any(term in disease_lower for term in ['not about pembrolizumab', 'procedural']):
            results['not_about_pembro'].append(r)

        # Write to spreadsheet
        if new_ab != existing_ab:
            ws.cell(r, COL_AB).value = new_ab

        results['all_rows'].append(row_summary)

    print(f"\nProcessed {rows_with_pdf} rows with PDF links")
    print(f"Filled AB column: {rows_ab_filled} rows")
    print(f"PDF mismatches: {len(results['pdf_mismatches'])} rows")
    print(f"Disease warnings: {len(results['disease_warnings'])} rows")

    # ── Save updated spreadsheet ──────────────────────────────────────────────
    print(f"\nSaving updated spreadsheet: {OUTPUT_XLSX}")
    wb.save(OUTPUT_XLSX)
    print("Spreadsheet saved.")

    # ── Generate report ───────────────────────────────────────────────────────
    print(f"\nGenerating report: {OUTPUT_REPORT}")
    generate_report(results, OUTPUT_REPORT)
    print("Report saved.")

    return results

def generate_report(results, report_path):
    """Generate the detailed analysis report."""
    lines = []

    def L(s=""):
        lines.append(s)

    L("=" * 80)
    L("REPORT: COMPREHENSIVE PDF CLINICAL DATA ANALYSIS")
    L(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    L("=" * 80)
    L()

    # ── Executive Summary ─────────────────────────────────────────────────────
    L("EXECUTIVE SUMMARY")
    L("-" * 40)

    total_with_pdf = sum(1 for r in results['all_rows'] if r['pdf_link'])
    total_ab_filled = len(results['filled_ab'])
    total_mismatches = len(results['pdf_mismatches'])
    total_warnings = len(results['disease_warnings'])
    total_errors = len(results['errors'])

    L(f"Total rows with PDF links:        {total_with_pdf}")
    L(f"Rows where AB column was filled:  {total_ab_filled}")
    L(f"Confirmed PDF mismatches:         {total_mismatches}")
    L(f"Disease/content warnings:         {total_warnings}")
    L(f"Extraction errors:                {total_errors}")
    L()

    # ── Section 1: Rows where AB was filled ───────────────────────────────────
    L("=" * 80)
    L("SECTION 1: ROWS WHERE COLUMN AB WAS EMPTY AND WAS FILLED")
    L("=" * 80)
    L()

    ab_filled_rows = [r for r in results['all_rows'] if r.get('ab_filled')]
    if not ab_filled_rows:
        L("No rows had empty AB that could be filled.")
    else:
        for row in ab_filled_rows:
            L(f"Row {row['row']} | {row['country']} | Case: {row['case']}")
            L(f"  Disease (N): {row['disease']}")
            L(f"  Details (O): {row['details'][:80]}")
            L(f"  PDF Link:    {row['pdf_link']}")
            L(f"  NEW AB:      {row['new_ab'][:200]}")
            if row['notes']:
                L(f"  Notes:       {' | '.join(row['notes'])}")
            L()

    # ── Section 2: PDF Mismatches ─────────────────────────────────────────────
    L("=" * 80)
    L("SECTION 2: CONFIRMED PDF MISMATCH ERRORS")
    L("=" * 80)
    L()

    mismatch_rows = [r for r in results['all_rows']
                     if any('MISMATCH' in i for i in r.get('issues', []))]
    if not mismatch_rows:
        L("No PDF mismatches detected.")
    else:
        for row in mismatch_rows:
            L(f"Row {row['row']} | {row['country']} | Case: {row['case']}")
            L(f"  Disease (N): {row['disease']}")
            L(f"  PDF Link:    {row['pdf_link']}")
            for issue in row['issues']:
                if 'MISMATCH' in issue:
                    L(f"  ISSUE: {issue}")
            L()

    # ── Section 3: Disease warnings ───────────────────────────────────────────
    L("=" * 80)
    L("SECTION 3: ROWS WHERE DISEASE OR CLINICAL DETAILS MAY BE INCORRECT")
    L("=" * 80)
    L()

    warning_rows = [r for r in results['all_rows']
                    if any('WARNING' in i for i in r.get('issues', []))]
    if not warning_rows:
        L("No disease discrepancies detected.")
    else:
        for row in warning_rows:
            L(f"Row {row['row']} | {row['country']} | Case: {row['case']}")
            L(f"  Disease (N): {row['disease']}")
            L(f"  PDF Link:    {row['pdf_link']}")
            for issue in row['issues']:
                if 'WARNING' in issue:
                    L(f"  WARNING: {issue}")
            L()

    # ── Section 4: Errors ─────────────────────────────────────────────────────
    L("=" * 80)
    L("SECTION 4: EXTRACTION ERRORS")
    L("=" * 80)
    L()

    error_rows = [r for r in results['all_rows']
                  if any('ERROR' in i for i in r.get('issues', []))]
    if not error_rows:
        L("No extraction errors.")
    else:
        for row in error_rows:
            L(f"Row {row['row']} | Case: {row['case']} | PDF: {row['pdf_link']}")
            for issue in row['issues']:
                if 'ERROR' in issue:
                    L(f"  ERROR: {issue}")
            L()

    # ── Section 5: Procedural / Not About Pembrolizumab rows ──────────────────
    L("=" * 80)
    L("SECTION 5: ROWS CLASSIFIED AS 'NOT ABOUT PEMBROLIZUMAB' OR 'PROCEDURAL ISSUE'")
    L("(Verified against PDF content)")
    L("=" * 80)
    L()

    pembro_rows = [r for r in results['all_rows']
                   if r['row'] in results['not_about_pembro'] and r['pdf_link']]
    if not pembro_rows:
        L("No such rows with PDF links.")
    else:
        for row in pembro_rows:
            L(f"Row {row['row']} | {row['country']} | Case: {row['case']}")
            L(f"  Disease (N): {row['disease']}")
            L(f"  PDF Link:    {row['pdf_link']}")
            if row['new_ab']:
                L(f"  PDF Findings: {row['new_ab'][:200]}")
            if row['notes']:
                L(f"  Notes: {' | '.join(row['notes'])}")
            L()

    # ── Section 6: Full per-row analysis ──────────────────────────────────────
    L("=" * 80)
    L("SECTION 6: FULL PER-ROW ANALYSIS LOG")
    L("=" * 80)
    L()

    for row in results['all_rows']:
        if not row['pdf_link']:
            continue  # Skip rows with no PDF

        L(f"--- Row {row['row']} | {row['country']} | Case: {row['case']} ---")
        L(f"  Disease (N):  {row['disease']}")
        L(f"  Details (O):  {row['details'][:120]}")
        L(f"  PDF Link:     {row['pdf_link']}")
        L(f"  AB was empty: {row['ab_was_empty']}")
        if row['ab_filled']:
            L(f"  AB FILLED:    {row['new_ab'][:200]}")
        elif row['new_ab'] and row['new_ab'] == row['existing_ab']:
            L(f"  AB existing:  {str(row['existing_ab'])[:100]}")
        if row['issues']:
            for issue in row['issues']:
                L(f"  [ISSUE] {issue}")
        if row['notes']:
            L(f"  [NOTES] {' | '.join(row['notes'])}")
        L()

    with open(report_path, 'w', encoding='utf-8') as f:
        f.write("\n".join(lines))

if __name__ == "__main__":
    results = main()
