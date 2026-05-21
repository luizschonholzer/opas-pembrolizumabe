import openpyxl
from openpyxl.styles import Font
from pathlib import Path
import re
import os
from datetime import datetime, date

BASE_DIR = Path(r'C:\Users\Claudio\Downloads\Processos - OPAS')

COUNTRY_FOLDER_MAP = {
    'Argentina': 'Argentina',
    'Brazil': 'Brasil',
    'Chile': 'Chile',
    'Colombia': 'Colombia',
    'Ecuador': 'Equador',
    'Guatemala': 'Guatemala',
    'Uruguay': 'Uruguai',
    'Costa Rica': 'Costa Rica',
}

def get_date_str(date_val):
    if date_val is None:
        return None
    if isinstance(date_val, (datetime, date)):
        return date_val.strftime('%Y-%m-%d')
    s = str(date_val).strip()
    if re.match(r'^\d{4}-\d{2}-\d{2}', s):
        return s[:10]
    return None

def get_pdfs(folder):
    result = []
    for root, dirs, files in os.walk(folder):
        for f in files:
            if f.lower().endswith(('.pdf', '.docx')):
                result.append(Path(root) / f)
    return result

def score_match(case_str, date_val, pdf_path):
    pdf_name = pdf_path.name
    case_str = str(case_str) if case_str else ''
    score = 0

    case_nums = re.findall(r'\d+', case_str)
    pdf_nums = re.findall(r'\d+', pdf_name)

    # Date-based match (high priority, used for Uruguay)
    date_str = get_date_str(date_val)
    if date_str and date_str in pdf_name:
        score += 20
        # Penalize extra non-date numbers in PDF that don't appear in the case
        date_parts = set(re.findall(r'\d+', date_str))
        extra_nums = [n for n in pdf_nums if n not in date_parts]
        for n in extra_nums:
            if n not in case_nums:
                score -= 2
            else:
                score += 4

    # Normalized string match
    case_norm = re.sub(r'[^a-z0-9]', '', case_str.lower())
    pdf_base = pdf_name.lower()
    for ext in ['.pdf', '.docx']:
        pdf_base = pdf_base.replace(ext, '')
    pdf_norm = re.sub(r'[^a-z0-9]', '', pdf_base)

    if case_norm and len(case_norm) >= 4:
        if case_norm == pdf_norm:
            score += 15
        elif case_norm in pdf_norm:
            score += 8
        elif pdf_norm in case_norm:
            score += 5

    # Number matching (only for numbers >= 3 digits to avoid false positives)
    if case_nums and pdf_nums:
        for n in case_nums:
            if len(n) >= 3 and n in pdf_nums:
                score += 5

    return score

wb = openpyxl.load_workbook(BASE_DIR / 'OPAS_translated_english_apr 2026.xlsx')
ws = wb['Sheet1']

NEW_COL = 27
ws.cell(row=1, column=NEW_COL, value='PDF Link')
ws.cell(row=1, column=NEW_COL).font = Font(bold=True)

matched = 0
unmatched_rows = []

for row in range(2, ws.max_row + 1):
    country = ws.cell(row=row, column=1).value
    case_num = ws.cell(row=row, column=6).value
    date_val = ws.cell(row=row, column=7).value

    if not country:
        continue

    folder_name = COUNTRY_FOLDER_MAP.get(country)
    if not folder_name:
        continue

    folder = BASE_DIR / folder_name
    if not folder.exists():
        continue

    if not case_num or str(case_num).strip().lower() in ('not found', 'procedural', 'procedural issue',
                                                           'not about pembrolizumab', 'out of scope',
                                                           'defective site / inaccessible'):
        continue

    pdfs = get_pdfs(folder)
    if not pdfs:
        continue

    scores = [(score_match(case_num, date_val, p), p) for p in pdfs]
    scores.sort(key=lambda x: x[0], reverse=True)

    best_score, best_pdf = scores[0]

    if best_score >= 8:
        pdf_url = 'file:///' + str(best_pdf).replace('\\', '/')
        cell = ws.cell(row=row, column=NEW_COL)
        cell.value = best_pdf.name
        cell.hyperlink = pdf_url
        cell.font = Font(color='0563C1', underline='single')
        matched += 1
        print(f'OK  Row {row:3d}: [{country}] {case_num} -> {best_pdf.name} (score={best_score})')
    else:
        unmatched_rows.append((row, country, case_num))
        print(f'--- Row {row:3d}: [{country}] {case_num} -> NO MATCH (best={best_score})')

print(f'\nMatched: {matched}')
print(f'Unmatched: {len(unmatched_rows)}')
if unmatched_rows:
    print('Rows without PDF:')
    for r, c, n in unmatched_rows:
        print(f'  Row {r}: [{c}] {n}')

output = BASE_DIR / 'OPAS_com_links_PDF.xlsx'
wb.save(output)
print(f'\nSalvo em: {output}')
