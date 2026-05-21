#!/usr/bin/env python3
"""
Version 2: Comprehensive PDF clinical data analysis with accurate extraction.
"""

import os
import re
import sys
import traceback
from datetime import datetime
import pdfplumber
import openpyxl

BASE_DIR = r"C:\Users\Claudio\Downloads\Processos - OPAS"
INPUT_XLSX = os.path.join(BASE_DIR, "OPAS_com_links_PDF_BACKUP_20260520_1713.xlsx")
OUTPUT_XLSX = os.path.join(BASE_DIR, "OPAS_ANALISE_ATUALIZADA_20260521.xlsx")
OUTPUT_REPORT = os.path.join(BASE_DIR, "RELATORIO_ANALISE_PDFS_20260521.txt")

COL_COUNTRY = 1
COL_CASE = 6
COL_DISEASE = 14
COL_DETAILS = 15
COL_PDF_LINK = 27
COL_AB = 28

# Known PDF mismatches
KNOWN_MISMATCHES = {
    3: {
        'row': 3,
        'case': '013326/2024',
        'pdf': 'CCF 000104_2024.pdf',
        'correct_case': 'CCF 000104/2024 (row 8)',
        'note': 'PDF belongs to a different case. The PDF CCF 000104_2024.pdf corresponds to case CCF 000104/2024 (row 8 - Triple-negative apocrine breast carcinoma). This row (013326/2024 - procedural issue) has no matching PDF available.'
    },
    4: {
        'row': 4,
        'case': 'CCF 011768/2025',
        'pdf': '14838_2025.pdf',
        'correct_case': '14838/2025 (row 2)',
        'note': 'PDF belongs to a different case. The PDF 14838_2025.pdf corresponds to case 14838/2025 (row 2 - Lung cancer). This row (CCF 011768/2025 - Not about pembrolizumab) has no matching PDF available.'
    },
    6: {
        'row': 6,
        'case': 'CCF 016494/2025',
        'pdf': '14838_2025.pdf',
        'correct_case': '14838/2025 (row 2)',
        'note': 'PDF belongs to a different case. The PDF 14838_2025.pdf corresponds to case 14838/2025 (row 2 - Lung cancer). This row (CCF 016494/2025 - Advanced-stage cervical cancer) has the wrong PDF assigned.'
    },
    11: {
        'row': 11,
        'case': 'CCF 000453/2016/CA001',
        'pdf': 'CCF 004030_2020_CA001.pdf',
        'correct_case': 'CCF 004030/2020/CA001 (row 10)',
        'note': 'PDF belongs to a different case. The PDF CCF 004030_2020_CA001.pdf corresponds to case CCF 004030/2020/CA001 (row 10 - Advanced-stage renal cancer). This row (CCF 000453/2016/CA001 - NI) has no matching PDF available.'
    },
    128: {
        'row': 128,
        'case': '7104-2025',
        'pdf': '1343-2025.pdf',
        'correct_case': '1343-2025 (row 122)',
        'note': 'PDF belongs to a different case. The PDF 1343-2025.pdf corresponds to case 1343-2025 (row 122 - Cervical Cancer). This row (7104-2025 - Not about pembrolizumab) has the wrong PDF assigned.'
    },
}

def build_pdf_map(base_dir):
    pdf_map = {}
    for root, dirs, files in os.walk(base_dir):
        for f in files:
            if f.lower().endswith('.pdf') or f.lower().endswith('.docx'):
                pdf_map[f] = os.path.join(root, f)
    return pdf_map

def extract_pdf_text(pdf_path):
    try:
        text_pages = []
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    text_pages.append(t)
        return "\n".join(text_pages)
    except Exception as e:
        return f"[ERROR: {e}]"

def extract_docx_text(docx_path):
    try:
        from docx import Document
        doc = Document(docx_path)
        return "\n".join([p.text for p in doc.paragraphs])
    except Exception as e:
        return f"[ERROR DOCX: {e}]"

def extract_text(file_path):
    if file_path.lower().endswith('.docx'):
        return extract_docx_text(file_path)
    return extract_pdf_text(file_path)

def clean_text_fragment(s, max_len=100):
    """Clean a text fragment: remove excess whitespace and truncate."""
    s = re.sub(r'\s+', ' ', s).strip()
    if len(s) > max_len:
        s = s[:max_len].rstrip() + "..."
    return s

def extract_age(text):
    """Extract patient age from text (Spanish/Portuguese/English)."""
    patterns = [
        r'(\d{2})\s*(?:años?|anos?)\s*de\s*edad',
        r'(\d{2})\s*(?:años?|anos?)\s*de\s*edad',
        r'paciente?\s+(?:de\s+)?(\d{2})\s*(?:años?|anos?)',
        r'actor\s+(?:es\s+un\s+)?(?:paciente\s+de\s+)?(\d{2})\s*a[ñn]os',
        r'(\d{2})\s*(?:years?\s*old|year\s*old)',
        r'age[:\s]+(\d{2})',
        r'aged?\s+(\d{2})',
        r'(\d{2})\s*(?:años?|anos?)',
        r'(?:tiene|tiene|tem)\s+(\d{2})\s*(?:años?|anos?)',
        r'nasceu.*?(\d{2})\s*anos?',
        r'(\d{2})\s*a[ñn]os?\s*de\s*edad',
    ]
    for pat in patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            age = int(m.group(1))
            if 18 <= age <= 100:
                return f"{age}-year-old patient"
    return None

def extract_stage(text):
    """Extract disease staging from text."""
    # Roman numeral stage patterns - be careful not to match "la" -> "IA"
    patterns = [
        # English explicit stage mentions
        r'(?:clinical\s+stage|stage)\s+(IV[ABC]?|III[ABC]?|II[ABC]?|I[ABC]?)\b',
        r'(?:FIGO|figo)\s+(IV[ABC]?|III[ABC]?|II[ABC]?|I[ABC]?)\b',
        r'pT\d+[a-d]?\s*pN\d+[a-c]?\s*[pM](\d+)',  # TNM staging
        # Spanish/Portuguese explicit stage
        r'(?:estadio|estadío|etapa|EC)\s+(IV[ABC]?|III[ABC]?|II[ABC]?|I[ABC]?)\b',
        r'(?:estadio|estadío|etapa|EC)\.?\s*(IV|III|II|I[VAB]?)',
        r'estadiaje\s+(?:cl[íi]nico\s+)?(?:cercano\s+a\s+)?(T\d)',
        # Direct stage numbers with explicit context
        r'\b(IV[ABC]?)\b(?=\s*(?:de\s+)?(?:la\s+)?(?:enfermedad|disease|cancer|cáncer|carcinoma))',
        r'(?:b)\s+(IV[ABC]?)\s*;',
    ]

    for pat in patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            stage_raw = m.group(1).strip().upper()
            # Validate - must look like a stage
            if re.match(r'^(IV[ABC]?|III[ABC]?|II[ABC]?|I[VABC]?)$', stage_raw):
                return f"Stage {stage_raw}"
            elif re.match(r'^T\d', stage_raw):
                return f"Clinical staging: {stage_raw}"

    # Try more careful extraction - look for "estadio" followed by Roman numerals
    m = re.search(r'(?:estadio|stage|etapa|EC)[:\s.]+([IVX]{1,4}[ABC]?)\b', text, re.IGNORECASE)
    if m:
        stage_raw = m.group(1).upper()
        if re.match(r'^(IV[ABC]?|III[ABC]?|II[ABC]?|I[VABC]?)$', stage_raw):
            return f"Stage {stage_raw}"

    return None

def extract_pdl1(text):
    """Extract PD-L1 expression from text."""
    patterns = [
        r'PD-?L1[^.]{0,60}(\d{1,3}(?:\.\d+)?)\s*%',
        r'(\d{1,3}(?:\.\d+)?)\s*%[^.]{0,40}PD-?L1',
        r'(?:CPS|TPS)[:\s]+(\d{1,3}(?:\.\d+)?)',
        r'PD-?L1[^.]{0,50}(?:positivo|positive|negativo|negative|alto|high|bajo|low)',
        r'inmunoreactividad\s+para\s+PD-?L1\s+(?:Positivo|Negativo)[^.]{0,50}TPS\s+(\d+)%',
        r'PD-?L[1l]\s+(?:Positivo|Negativo)',
    ]
    for pat in patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            if m.lastindex and m.group(1):
                val = m.group(1)
                # Determine if it's CPS or TPS
                before = text[max(0,m.start()-20):m.start()]
                if 'CPS' in before.upper():
                    return f"PD-L1 CPS: {val}%"
                elif 'TPS' in before.upper():
                    return f"PD-L1 TPS: {val}%"
                else:
                    return f"PD-L1: {val}%"
            else:
                context = clean_text_fragment(m.group(0), 80)
                return f"PD-L1: {context}"
    return None

def extract_msi(text):
    """Extract microsatellite status."""
    patterns = [
        (r'\bMSI-H\b', 'High microsatellite instability (MSI-H)'),
        (r'\bMSI-L\b', 'Low microsatellite instability (MSI-L)'),
        (r'\bMSS\b', 'Microsatellite stable (MSS)'),
        (r'\bdMMR\b', 'Deficient mismatch repair (dMMR)'),
        (r'\bpMMR\b', 'Proficient mismatch repair (pMMR)'),
        (r'alta\s+inestabilidad\s+(?:de\s+)?microsatélites', 'High microsatellite instability (MSI-H)'),
        (r'microsatellite\s+instability.{0,30}high', 'High microsatellite instability (MSI-H)'),
        (r'high\s+microsatellite\s+instability', 'High microsatellite instability (MSI-H)'),
        (r'instabilidade\s+(?:de\s+)?microssatélites', 'Microsatellite instability'),
    ]
    for pat, label in patterns:
        if re.search(pat, text, re.IGNORECASE):
            return label
    return None

def extract_ecog(text):
    """Extract ECOG performance status."""
    patterns = [
        r'ECOG[:\s]+(\d)',
        r'ECOG[^.]{0,20}(\d)',
        r'performance\s+status[:\s]+(\d)',
        r'estado\s+(?:funcional|de\s+desempe[ñn]o)[:\s]+(\d)',
        r'\bPS[:\s]+(\d)',
        r'ECOG\s*PS[:\s]+(\d)',
    ]
    for pat in patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            val = m.group(1)
            if val in '01234':
                return f"ECOG PS: {val}"
    return None

def extract_prior_treatments(text):
    """Extract prior treatment history."""
    found = []

    # Chemotherapy drugs/regimens
    chemo_drugs = [
        'carboplatino', 'carboplatin', 'cisplatino', 'cisplatin',
        'paclitaxel', 'docetaxel', 'gemcitabina', 'gemcitabine',
        'oxaliplatina', 'oxaliplatin', 'pemetrexed', 'bevacizumab',
        'cetuximab', 'trastuzumab', 'pertuzumab', 'FOLFOX', 'FOLFIRI',
        'CAPOX', 'XELOX', 'capecitabine', 'capecitabina',
        'quimioterapia', 'chemotherapy', 'quimioterapia',
        'nivolumab', 'atezolizumab', 'durvalumab',
    ]

    chemo_found = set()
    for drug in chemo_drugs:
        if re.search(r'\b' + drug + r'\b', text, re.IGNORECASE):
            # Get context
            m = re.search(r'\b' + drug + r'\b', text, re.IGNORECASE)
            if m:
                canonical = drug.lower()
                if canonical in ['quimioterapia', 'chemotherapy', 'quimioterapia']:
                    chemo_found.add('Chemotherapy')
                else:
                    chemo_found.add(drug.capitalize() if drug[0].islower() else drug)

    if chemo_found:
        found.append("Prior chemotherapy: " + ", ".join(sorted(chemo_found)[:5]))

    # Surgery
    surgery_terms = [
        r'cirugía\s+(?:de\s+)?(?:tórax|toracoscopia|resección|\w+)',
        r'resección\s+\w+',
        r'nephrectomy|nefrectomía',
        r'mastectomy|mastectomía',
        r'hysterectomy|histerectomía',
        r'colectomy|colectomía',
        r'gastrectomy|gastrectomía',
        r'operación\s+(?:de\s+)?\w+',
        r'surgical\s+resection',
    ]
    surgery_found = []
    for pat in surgery_terms:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            surgery_found.append(clean_text_fragment(m.group(0), 50))
    if surgery_found:
        found.append("Prior surgery: " + "; ".join(surgery_found[:2]))

    # Radiotherapy
    if re.search(r'radioterapia|radiotherapy|radiation therapy|\bRT\b', text, re.IGNORECASE):
        found.append("Prior radiotherapy")

    return found

def extract_biomarkers(text):
    """Extract key biomarkers."""
    biomarkers = []

    patterns = [
        (r'EGFR[^.;]{0,50}(?:mutaci[oó]n|mutation|positiv|negativ|wild)', 'EGFR'),
        (r'EGFR\s+(?:Negativo|Positivo|mutado)', 'EGFR'),
        (r'mutaciones?\s+(?:en\s+el\s+)?(?:GEN\s+)?EGFR[^.;]{0,30}', 'EGFR'),
        (r'ALK[^.;]{0,40}(?:positiv|negativ|reordenam|rearrang|fusin|fusion)', 'ALK'),
        (r'(?:inmunoreactividad|immunoreactivity)\s+para\s+ALK[^.;]{0,30}', 'ALK'),
        (r'KRAS[^.;]{0,40}(?:mutaci[oó]n|mutation|positiv|negativ|wild|selvagem)', 'KRAS'),
        (r'HER2[^.;]{0,40}(?:positiv|negativ|amplif|sobreexpresi)', 'HER2'),
        (r'BRAF[^.;]{0,40}(?:V600|mutaci[oó]n|mutation|positiv|negativ)', 'BRAF'),
        (r'ROS1[^.;]{0,30}(?:positiv|negativ|fusin|fusion)', 'ROS1'),
        (r'TMB[^.;]{0,30}(?:alto|high|bajo|low|\d)', 'TMB'),
    ]

    seen_types = set()
    for pat, marker_type in patterns:
        if marker_type in seen_types:
            continue
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            fragment = clean_text_fragment(m.group(0), 70)
            biomarkers.append(fragment)
            seen_types.add(marker_type)

    return biomarkers

def extract_treatment_context(text):
    """Extract treatment context: neoadjuvant, adjuvant, palliative, line of treatment."""
    contexts = []

    if re.search(r'neoadjuvan[te]|neoadyuvan[te]', text, re.IGNORECASE):
        contexts.append("neoadjuvant")
    if re.search(r'\badjuvan[te]\b|adyuvan[te]', text, re.IGNORECASE):
        contexts.append("adjuvant")
    if re.search(r'paliativ[oa]|palliative', text, re.IGNORECASE):
        contexts.append("palliative intent")
    if re.search(r'primera\s+l[íi]nea|first[\s-]line|1[aª][\s-]l[íi]nea|primera\s+línea', text, re.IGNORECASE):
        contexts.append("first-line therapy")
    if re.search(r'segunda\s+l[íi]nea|second[\s-]line|2[aª][\s-]l[íi]nea|2da\s+l[íi]nea', text, re.IGNORECASE):
        contexts.append("second-line therapy")
    if re.search(r'tercera\s+l[íi]nea|third[\s-]line|3[aª][\s-]l[íi]nea', text, re.IGNORECASE):
        contexts.append("third-line therapy")

    return contexts

def extract_metastasis(text):
    """Extract metastasis information."""
    sites = []
    site_patterns = {
        'liver/hepatic': r'metástas[ie]s?\s+(?:hepáticas?|(?:en\s+)?hígado)|hepátic[ao]\s+metástas|liver\s+metastas|hepatic\s+metastas',
        'lung/pulmonary': r'metástas[ie]s?\s+(?:pulmonares?|(?:en\s+)?pulmón)|pulmonar\s+metástas|lung\s+metastas|pulmonary\s+metastas',
        'bone': r'metástas[ie]s?\s+(?:óseas?|(?:en\s+)?hueso)|bone\s+metastas|ósea\s+metástas',
        'brain': r'metástas[ie]s?\s+(?:cerebral|(?:en\s+)?cerebro)|brain\s+metastas|cerebral\s+metástas',
        'lymph node': r'metástas[ie]s?\s+(?:ganglionares?|(?:en\s+)?ganglios?)|lymph\s+node\s+metastas',
        'peritoneal': r'metástas[ie]s?\s+(?:peritoneales?)|peritoneal\s+metastas',
    }
    for site_label, pat in site_patterns.items():
        if re.search(pat, text, re.IGNORECASE):
            sites.append(site_label)

    # General metastasis mention
    if not sites:
        m = re.search(r'metástas[ie]s?[^.]{0,80}', text, re.IGNORECASE)
        if m:
            sites.append(clean_text_fragment(m.group(0), 80))

    return sites

def find_case_section(full_text, case_num, country, pdf_filename):
    """For multi-case PDFs, try to isolate the relevant case section."""
    if not case_num or not full_text:
        return full_text

    case_str = str(case_num).strip()

    # For Uruguay PDFs: case numbers like "150/2021", "47/2020"
    # The PDFs are named by date and may contain the case number directly
    if country and 'Uruguay' in str(country):
        # Extract numeric part
        m = re.match(r'^(\d+)/(\d{4})', case_str)
        if m:
            num = m.group(1)
            year = m.group(2)
            # Look for "Nro: NNN/YYYY" header pattern
            idx = full_text.find(f"{num}/{year}")
            if idx > -1:
                return full_text[max(0, idx-100):min(len(full_text), idx+8000)]

    # For Costa Rica PDFs: case numbers like "00087 - 2024"
    if country and 'Costa Rica' in str(country):
        # Extract the numeric part
        m = re.match(r'(\d+)\s*-\s*(\d{4})', case_str)
        if m:
            num = m.group(1).lstrip('0')
            year = m.group(2)
            # Try to find the case in text
            for pattern in [f"{num}-{year}", f"{num.zfill(5)}-{year}", case_str]:
                idx = full_text.find(pattern)
                if idx > -1:
                    return full_text[max(0, idx-100):min(len(full_text), idx+6000)]

    return full_text

def build_ab_entry(findings_dict):
    """Build a pipe-separated AB entry from findings."""
    parts = []

    if findings_dict.get('age'):
        parts.append(f"Patient age: {findings_dict['age']}")
    if findings_dict.get('stage'):
        parts.append(f"Disease staging: {findings_dict['stage']}")
    if findings_dict.get('ecog'):
        parts.append(findings_dict['ecog'])
    if findings_dict.get('pdl1'):
        parts.append(findings_dict['pdl1'])
    if findings_dict.get('msi'):
        parts.append(f"Microsatellite status: {findings_dict['msi']}")
    if findings_dict.get('biomarkers'):
        parts.append(f"Biomarkers: {' | '.join(findings_dict['biomarkers'][:4])}")
    if findings_dict.get('treatment_context'):
        parts.append(f"Treatment context: {', '.join(findings_dict['treatment_context'])}")
    if findings_dict.get('prior_treatments'):
        parts.append(" | ".join(findings_dict['prior_treatments']))
    if findings_dict.get('metastasis'):
        parts.append(f"Metastasis sites: {', '.join(findings_dict['metastasis'])}")
    if findings_dict.get('urgency'):
        parts.append(f"Clinical urgency/severity: {findings_dict['urgency']}")
    if findings_dict.get('notes'):
        parts.append(findings_dict['notes'])

    return " | ".join(parts) if parts else ""

def analyze_pdf_content(text, case_num, country, disease, details, pdf_filename):
    """
    Full analysis of PDF text. Returns:
    - findings dict
    - issues list
    - disease_verified bool
    """
    findings = {}
    issues = []

    if not text or text.startswith('[ERROR'):
        return findings, issues, False

    # Find case-specific section for multi-case PDFs
    section = find_case_section(text, case_num, country, pdf_filename)

    # Extract all fields
    age = extract_age(section)
    if age:
        findings['age'] = age

    stage = extract_stage(section)
    if stage:
        findings['stage'] = stage

    pdl1 = extract_pdl1(section)
    if pdl1:
        findings['pdl1'] = pdl1

    msi = extract_msi(section)
    if msi:
        findings['msi'] = msi

    ecog = extract_ecog(section)
    if ecog:
        findings['ecog'] = ecog

    prior_tx = extract_prior_treatments(section)
    if prior_tx:
        findings['prior_treatments'] = prior_tx

    biomarkers = extract_biomarkers(section)
    if biomarkers:
        findings['biomarkers'] = biomarkers

    treatment_ctx = extract_treatment_context(section)
    if treatment_ctx:
        findings['treatment_context'] = treatment_ctx

    mets = extract_metastasis(section)
    if mets:
        findings['metastasis'] = mets

    # Check urgency
    urgency_m = re.search(r'urgente|urgency|urgencia|emergencia|emergenc|crítico|critically', section, re.IGNORECASE)
    if urgency_m:
        urgency_ctx = clean_text_fragment(section[max(0,urgency_m.start()-20):urgency_m.end()+100], 120)
        findings['urgency'] = urgency_ctx

    # Disease verification - check if PDF content matches column N disease
    disease_str = str(disease or "").lower().strip()
    verified = True

    if disease_str and not any(t in disease_str for t in ['procedural', 'not about', 'ni', 'nao']):
        # Check for disease keyword presence in text
        disease_keywords = {
            'lung': ['pulmon', 'pulm?n', 'lung', 'bronquio', 'nsclc', 'adenocarcinoma de pulm'],
            'breast': ['mama', 'breast', 'mamario', 'mamaria'],
            'cervical': ['cervical', 'cuello', 'cérvix', 'cervix'],
            'renal': ['renal', 'riñón', 'kidney', 'células claras', 'clear cell'],
            'melanoma': ['melanoma'],
            'lymphoma': ['lymphoma', 'linfoma'],
            'colon': ['colon', 'colorectal', 'cólon', 'recto'],
            'endometrial': ['endometrial', 'endometrio', 'uterine'],
            'bladder': ['bladder', 'vejiga', 'urotelial', 'urothelial'],
            'gastric': ['gástrico', 'gastric', 'estómago', 'stomach'],
            'ovary': ['ovar', 'ovário', 'ovario'],
        }

        text_lower = section.lower()
        for disease_key, keywords in disease_keywords.items():
            if disease_key in disease_str:
                found_any = any(re.search(kw, text_lower) for kw in keywords)
                if not found_any:
                    issues.append(f"Disease '{disease}' not clearly confirmed in PDF text (expected keywords: {keywords[:3]})")
                    verified = False
                break

    return findings, issues, verified

# ─── Row-by-row processing ────────────────────────────────────────────────────

def process_row(row_num, row_data, pdf_map):
    """
    Process a single row.
    Returns: result dict with new_ab, issues, notes, etc.
    """
    country = row_data['country']
    case_num = row_data['case']
    disease = row_data['disease']
    details = row_data['details']
    pdf_filename = row_data['pdf_link']
    existing_ab = row_data['existing_ab']

    result = {
        'new_ab': existing_ab or "",
        'new_details': details or "",
        'issues': [],
        'notes': [],
        'ab_filled': False,
        'has_mismatch': False,
        'has_disease_warning': False,
    }

    # Check for known mismatches
    if row_num in KNOWN_MISMATCHES:
        mm = KNOWN_MISMATCHES[row_num]
        result['issues'].append(f"PDF MISMATCH: {mm['note']}")
        result['has_mismatch'] = True
        # Don't extract from wrong PDF
        return result

    if not pdf_filename:
        result['notes'].append("No PDF link in spreadsheet")
        return result

    pdf_filename = pdf_filename.strip()
    if pdf_filename not in pdf_map:
        result['issues'].append(f"PDF file not found on disk: '{pdf_filename}'")
        return result

    file_path = pdf_map[pdf_filename]

    # Extract text
    full_text = extract_text(file_path)
    if full_text.startswith('[ERROR'):
        result['issues'].append(f"Text extraction error: {full_text}")
        return result

    if not full_text.strip():
        result['issues'].append(f"Empty text extracted from PDF '{pdf_filename}'")
        return result

    # Analyze content
    findings, issues, disease_verified = analyze_pdf_content(
        full_text, case_num, country, disease, details, pdf_filename
    )

    for issue in issues:
        result['issues'].append(f"WARNING: {issue}")
        result['has_disease_warning'] = True

    # Build AB entry if empty
    disease_lower = str(disease or "").lower().strip()
    is_procedural = any(t in disease_lower for t in ['procedural', 'not about pembrolizumab'])
    is_ni = disease_lower in ['ni', '']

    if not existing_ab or existing_ab.strip() == "":
        if is_procedural:
            # For procedural/not-about cases, still extract and note
            ab_content = build_ab_entry(findings)
            if ab_content:
                result['new_ab'] = f"[PDF checked - classified as '{disease}'] {ab_content}"
                result['notes'].append(f"AB filled with PDF content despite '{disease}' classification")
                result['ab_filled'] = True
            else:
                result['notes'].append(f"PDF found but no extractable clinical data for '{disease}' case")
        elif is_ni:
            ab_content = build_ab_entry(findings)
            if ab_content:
                result['new_ab'] = ab_content
                result['notes'].append("AB filled from PDF (disease was NI/empty)")
                result['ab_filled'] = True
            else:
                result['notes'].append("PDF found but could not extract clinical data")
        else:
            # Normal case with real disease
            ab_content = build_ab_entry(findings)
            if ab_content:
                result['new_ab'] = ab_content
                result['notes'].append("AB filled from PDF clinical data")
                result['ab_filled'] = True
            else:
                result['notes'].append("PDF found but could not extract clinical data")
    else:
        result['notes'].append("AB already populated - no change")

    return result

# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("=" * 70)
    print("COMPREHENSIVE PDF CLINICAL DATA ANALYSIS v2")
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)

    # Build PDF map
    print("\nScanning PDF files...")
    pdf_map = build_pdf_map(BASE_DIR)
    print(f"Found {len(pdf_map)} files")

    # Load spreadsheet
    print(f"\nLoading: {INPUT_XLSX}")
    wb = openpyxl.load_workbook(INPUT_XLSX)
    ws = wb.active

    # Collect data rows
    data_rows = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, COL_CASE).value:
            data_rows.append(r)
    print(f"Data rows: {len(data_rows)}")

    # Process each row
    all_results = []
    rows_filled = 0
    rows_mismatches = 0
    rows_warnings = 0

    for r in data_rows:
        row_data = {
            'country': ws.cell(r, COL_COUNTRY).value or "",
            'case': ws.cell(r, COL_CASE).value or "",
            'disease': ws.cell(r, COL_DISEASE).value or "",
            'details': ws.cell(r, COL_DETAILS).value or "",
            'pdf_link': ws.cell(r, COL_PDF_LINK).value,
            'existing_ab': ws.cell(r, COL_AB).value or "",
        }

        result = process_row(r, row_data, pdf_map)
        result['row'] = r
        result['row_data'] = row_data
        all_results.append(result)

        # Apply changes to spreadsheet (only AB column)
        if result['new_ab'] != (row_data['existing_ab'] or ""):
            ws.cell(r, COL_AB).value = result['new_ab']

        if result['ab_filled']:
            rows_filled += 1
        if result['has_mismatch']:
            rows_mismatches += 1
        if result['has_disease_warning']:
            rows_warnings += 1

        print(f"  Row {r:3d}: {str(row_data['case'])[:25]:<25} | "
              f"{'FILLED' if result['ab_filled'] else 'MISMATCH' if result['has_mismatch'] else 'WARN' if result['has_disease_warning'] else 'OK'}")

    print(f"\nSummary: {rows_filled} AB filled, {rows_mismatches} mismatches, {rows_warnings} warnings")

    # Save spreadsheet
    print(f"\nSaving: {OUTPUT_XLSX}")
    wb.save(OUTPUT_XLSX)
    print("Saved.")

    # Generate report
    print(f"\nGenerating report: {OUTPUT_REPORT}")
    generate_report(all_results, OUTPUT_REPORT)
    print("Report saved.")

    return all_results

def generate_report(all_results, report_path):
    lines = []

    def L(s=""):
        lines.append(str(s))

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    L("=" * 80)
    L("REPORT: COMPREHENSIVE PDF CLINICAL DATA ANALYSIS")
    L(f"Generated: {now}")
    L(f"Input file: OPAS_com_links_PDF_BACKUP_20260520_1713.xlsx")
    L(f"Output file: OPAS_ANALISE_ATUALIZADA_20260521.xlsx")
    L("=" * 80)
    L()

    # Summary counts
    rows_with_pdf = [r for r in all_results if r['row_data']['pdf_link']]
    rows_filled = [r for r in all_results if r['ab_filled']]
    rows_mismatches = [r for r in all_results if r['has_mismatch']]
    rows_warnings = [r for r in all_results if r['has_disease_warning']]
    rows_errors = [r for r in all_results if any('error' in i.lower() for i in r['issues'])]

    L("EXECUTIVE SUMMARY")
    L("-" * 40)
    L(f"Total rows with PDF links:                    {len(rows_with_pdf)}")
    L(f"Rows where column AB was empty and filled:    {len(rows_filled)}")
    L(f"Rows with confirmed PDF mismatch errors:      {len(rows_mismatches)}")
    L(f"Rows with disease/content discrepancies:      {len(rows_warnings)}")
    L(f"Rows with PDF extraction errors:              {len(rows_errors)}")
    L()

    # ── Section 1: AB Filled Rows ─────────────────────────────────────────────
    L("=" * 80)
    L("SECTION 1: ROWS WHERE COLUMN AB WAS EMPTY AND WAS FILLED IN")
    L("=" * 80)
    L()

    if not rows_filled:
        L("No rows had empty AB that could be filled.")
    else:
        L(f"Total rows filled: {len(rows_filled)}")
        L()
        for res in rows_filled:
            r = res['row']
            d = res['row_data']
            L(f"Row {r} | Country: {d['country']} | Case: {d['case']}")
            L(f"  Disease (N):    {d['disease']}")
            L(f"  Details (O):    {str(d['details'])[:100]}")
            L(f"  PDF file:       {d['pdf_link']}")
            L(f"  NEW AB value:   {res['new_ab']}")
            if res['notes']:
                L(f"  Notes:          {' | '.join(res['notes'])}")
            L()

    # ── Section 2: PDF Mismatches ─────────────────────────────────────────────
    L("=" * 80)
    L("SECTION 2: CONFIRMED PDF MISMATCH ERRORS")
    L("(Rows where the PDF assigned does not correspond to this case)")
    L("=" * 80)
    L()

    if not rows_mismatches:
        L("No PDF mismatches detected.")
    else:
        L(f"Total mismatch rows: {len(rows_mismatches)}")
        L()
        for res in rows_mismatches:
            r = res['row']
            d = res['row_data']
            L(f"Row {r} | Country: {d['country']} | Case: {d['case']}")
            L(f"  Disease (N):  {d['disease']}")
            L(f"  PDF assigned: {d['pdf_link']}")
            for issue in res['issues']:
                L(f"  [MISMATCH] {issue}")
            L()

    # ── Section 3: Disease discrepancies ─────────────────────────────────────
    L("=" * 80)
    L("SECTION 3: ROWS WHERE DISEASE OR CLINICAL DATA MAY BE INCORRECT")
    L("(Cases where column N disease was not confirmed in PDF content)")
    L("=" * 80)
    L()

    if not rows_warnings:
        L("No disease discrepancies detected.")
    else:
        L(f"Total warning rows: {len(rows_warnings)}")
        L()
        for res in rows_warnings:
            r = res['row']
            d = res['row_data']
            L(f"Row {r} | Country: {d['country']} | Case: {d['case']}")
            L(f"  Disease (N):  {d['disease']}")
            L(f"  Details (O):  {str(d['details'])[:100]}")
            L(f"  PDF file:     {d['pdf_link']}")
            for issue in res['issues']:
                if 'WARNING' in issue.upper():
                    L(f"  [WARNING] {issue}")
            if res['notes']:
                L(f"  Notes:        {' | '.join(res['notes'])}")
            L()

    # ── Section 4: Procedural/Not-About Pembrolizumab checks ─────────────────
    L("=" * 80)
    L("SECTION 4: 'NOT ABOUT PEMBROLIZUMAB' / 'PROCEDURAL ISSUE' ROWS WITH PDFs")
    L("(Verified against PDF content to confirm classification or flag if wrong)")
    L("=" * 80)
    L()

    procedural_rows = [res for res in all_results
                       if res['row_data']['pdf_link']
                       and any(t in str(res['row_data']['disease']).lower()
                               for t in ['procedural', 'not about pembrolizumab'])]

    if not procedural_rows:
        L("No such rows with PDF links.")
    else:
        L(f"Total: {len(procedural_rows)} rows")
        L()
        for res in procedural_rows:
            r = res['row']
            d = res['row_data']
            is_mismatch = res['has_mismatch']
            L(f"Row {r} | Country: {d['country']} | Case: {d['case']}")
            L(f"  Classification: {d['disease']}")
            L(f"  PDF file:       {d['pdf_link']}")
            if is_mismatch:
                L(f"  STATUS: PDF IS MISMATCHED - cannot verify classification")
            elif res['new_ab']:
                L(f"  PDF findings:   {res['new_ab'][:200]}")
                L(f"  STATUS: Classification may need review - see PDF findings above")
            else:
                L(f"  STATUS: No extractable clinical data from PDF to verify classification")
            if res['notes']:
                L(f"  Notes: {' | '.join(res['notes'])}")
            L()

    # ── Section 5: Special case notes ────────────────────────────────────────
    L("=" * 80)
    L("SECTION 5: SPECIAL FINDINGS AND OBSERVATIONS")
    L("=" * 80)
    L()

    L("5.1 ROW 101 (Guatemala, Case 6313-2023) - DISEASE MISMATCH IN COLUMN N")
    L("  Column N says: 'Pulmonary and lymph node involvement'")
    L("  PDF reveals: The case is about Acral invasive Melanoma (Breslow 2mm, Clark IV)")
    L("  Patient: Nery Jeronimo Solis Requena, IGSS affiliate")
    L("  Actual diagnosis: 'Melanoma Acral invasor Breslow 2mm Clark IV'")
    L("  RECOMMENDATION: Update column N from 'Pulmonary and lymph node involvement'")
    L("  to 'Acral invasive melanoma (Breslow 2mm, Clark IV)'")
    L()

    L("5.2 ROW 51 (Colombia, Case 81511442) - DOCX FILE")
    L("  This case uses a .docx file instead of PDF.")
    L("  Patient: Anibal Trujillo Parra, 65 years old")
    L("  Diagnosis: Lung cancer (malignant tumor of the main bronchus) clinical staging")
    L("  close to T4")
    L("  EGFR: Negative | ALK: Negative | PD-L1 (TPS): 95% Positive")
    L("  Prior treatment: 3 cycles of chemotherapy (no favorable response)")
    L("  Treatment requested: Pembrolizumab 200mg, 6 cycles, palliative intent")
    L("  Treatment context: Second-line, palliative intent")
    L("  Note: Column N already has 'LUNG CANCER with clinical staging close to T4'")
    L("  AB was empty - filled with extracted clinical data.")
    L()

    L("5.3 ROW 53 (Ecuador, Case 2603/21-EP) - NI DISEASE")
    L("  Column N says: 'NI'")
    L("  PDF: The document is a Constitutional Court admission ruling (inadmissibility)")
    L("  The patient (Rosibel Alexandra Pena Narvaez) had a catastrophic illness")
    L("  and received Pembrolizumab for emergency treatment.")
    L("  The exact disease diagnosis is not specified in this court document.")
    L("  Classification as 'NI' (Not Identified) is appropriate for this admission ruling.")
    L()

    L("5.4 ROW 28 (Brazil, Case ARE 1572092) - OVARY CANCER")
    L("  Column N says: 'MALIGNANT NEOPLASM OF THE OVARY'")
    L("  PDF confirms: Treatment of malignant ovarian neoplasia with Pembrolizumab")
    L("  Patient: Edineia de Jesus Viana")
    L("  The PDF is a procedural/admissibility ruling - limited clinical detail.")
    L("  AB was empty - minimal clinical data available in this court document.")
    L()

    L("5.5 URUGUAY MULTI-CASE PDFs")
    L("  The following Uruguay PDFs contain only ONE case despite being linked to")
    L("  multiple rows (the other rows are 'Not about pembrolizumab' cases that")
    L("  appear to lack dedicated PDFs):")
    L("  - 2020-07-14.pdf: Only contains case 47/2020 (Row 89)")
    L("    Rows 88 (194/2020), 90 (91/2020), 91 (34/2020) are linked to this PDF")
    L("    but their cases are not found in the document.")
    L("  - 2021-02-22.pdf: Only contains case 19/2021 (Row 87)")
    L("    Row 86 (22/2021) is linked to this PDF but case not found in document.")
    L()

    # ── Section 6: Extraction errors ─────────────────────────────────────────
    L("=" * 80)
    L("SECTION 6: EXTRACTION ERRORS")
    L("=" * 80)
    L()

    if not rows_errors:
        L("No extraction errors.")
    else:
        for res in rows_errors:
            r = res['row']
            d = res['row_data']
            L(f"Row {r} | Case: {d['case']} | PDF: {d['pdf_link']}")
            for issue in res['issues']:
                if 'error' in issue.lower():
                    L(f"  [ERROR] {issue}")
            L()

    # ── Section 7: Full per-row log ───────────────────────────────────────────
    L("=" * 80)
    L("SECTION 7: FULL PER-ROW ANALYSIS LOG")
    L("(Only rows with PDF links)")
    L("=" * 80)
    L()

    for res in all_results:
        if not res['row_data']['pdf_link']:
            continue

        r = res['row']
        d = res['row_data']
        status_parts = []
        if res['ab_filled']:
            status_parts.append("AB FILLED")
        if res['has_mismatch']:
            status_parts.append("PDF MISMATCH")
        if res['has_disease_warning']:
            status_parts.append("DISEASE WARNING")
        status = " | ".join(status_parts) if status_parts else "OK"

        L(f"--- Row {r} [{status}] | {d['country']} | Case: {d['case']} ---")
        L(f"  Disease (N):   {d['disease']}")
        L(f"  Details (O):   {str(d['details'])[:120]}")
        L(f"  PDF file:      {d['pdf_link']}")
        L(f"  AB was empty:  {not bool(d['existing_ab'])}")

        if res['ab_filled']:
            L(f"  NEW AB:        {res['new_ab']}")
        elif d['existing_ab']:
            L(f"  Existing AB:   {str(d['existing_ab'])[:150]}")

        for issue in res['issues']:
            L(f"  [ISSUE] {issue}")
        for note in res['notes']:
            L(f"  [NOTE] {note}")
        L()

    # Write report
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write("\n".join(lines))

if __name__ == "__main__":
    main()
