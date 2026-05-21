# -*- coding: utf-8 -*-
"""
Análise clínica dos processos OPAS — Pembrolizumab
Filtro: apenas linhas onde coluna T (Pembrolizumab Granted) = 0 ou 1
Saída : relatório TXT + nova coluna no Excel com achados clínicos (em inglês)
"""

import openpyxl
import pdfplumber
import os
import re
from datetime import date, timedelta

BASE_DIR   = r"C:\Users\Claudio\Downloads\Processos - OPAS"
XLSX_IN    = os.path.join(BASE_DIR, "OPAS_com_links_PDF_BACKUP.xlsx")
XLSX_OUT   = os.path.join(BASE_DIR, "OPAS_com_links_PDF.xlsx")
REPORT_OUT = os.path.join(BASE_DIR, "RELATORIO_ACHADOS_CLINICOS.txt")

COUNTRY_FOLDER = {
    "Argentina":  "Argentina",
    "Brazil":     "Brasil",
    "Chile":      "Chile",
    "Colombia":   "Colombia",
    "Ecuador":    "Equador",
    "Guatemala":  "Guatemala",
    "Costa Rica": "Costa Rica",
    "Uruguay":    "Uruguai",
    "Paraguay":   "Paraguay",
    "Peru":       "Peru",
    "Bolivia":    "Bolivia",
    "Venezuela":  "Venezuela",
    "Guyana":     "Guyana",
    "Canada":     "Canada",
}

# ── helpers ─────────────────────────────────────────────────────────────────

def find_pdf(pdf_name, country):
    if not pdf_name or str(pdf_name).strip() in ("", "None"):
        return None
    pdf_name = str(pdf_name).strip()
    folder   = COUNTRY_FOLDER.get(country, country)
    base     = os.path.join(BASE_DIR, folder)
    if os.path.exists(base):
        for root, _, files in os.walk(base):
            for f in files:
                if f.strip() == pdf_name:
                    return os.path.join(root, f)
    direct = os.path.join(BASE_DIR, pdf_name)
    return direct if os.path.exists(direct) else None


def read_pdf(path, max_pages=20):
    try:
        pages = []
        with pdfplumber.open(path) as pdf:
            for p in pdf.pages[:max_pages]:
                t = p.extract_text()
                if t:
                    pages.append(t)
        return "\n".join(pages)
    except Exception as e:
        return f"[PDF ERROR: {e}]"


def val(v):
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() in ("none", "nan") else s


def fmt_date(d):
    if d is None:
        return "not found"
    s = str(d)
    if s.isdigit() and int(s) > 10000:
        try:
            return (date(1899, 12, 30) + timedelta(days=int(s))).strftime("%Y-%m-%d")
        except:
            return s
    return s[:10] if "00:00:00" in s else s


def is_granted_01(v):
    """Return True only if the cell contains exactly 0 or 1 (numeric)."""
    if v is None:
        return False
    if isinstance(v, (int, float)):
        return v in (0, 1)
    s = str(v).strip()
    return s in ("0", "1", "0.0", "1.0")


# ── PDF clinical extractors ──────────────────────────────────────────────────

def first_match(text, patterns):
    for pat in patterns:
        m = re.search(pat, text, re.IGNORECASE | re.DOTALL)
        if m:
            return m.group(0)[:250].strip()
    return ""


def extract_age(text):
    for pat in [
        r"(\d{1,3})\s*(?:years?\s*old|años?\s*de\s*edad|anos?\s*de\s*idade)",
        r"(?:age|edad|idade)\s*[:\-]?\s*(\d{1,3})",
        r"paciente\s+de\s+(\d{1,3})\s+años?",
        r"(\d{1,3})\s*-year-old",
        r"\((\d{2,3})\s*años?\)",
    ]:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            gs = [g for g in m.groups() if g and g.isdigit()]
            if gs and 1 <= int(gs[0]) <= 120:
                return gs[0] + " years old"
    return ""


def extract_cid(text):
    found = []
    for pat in [
        r"(?:CID|ICD|CIE)[\s\-:]*([A-Z]\d{2}(?:\.\d{1,2})?)",
        r"\b([C][0-9]{2}\.[0-9]{1,2})\b",
    ]:
        for m in re.finditer(pat, text, re.IGNORECASE):
            c = (m.group(1) if m.lastindex >= 1 else m.group(0)).upper()
            if re.match(r"[A-Z]\d{2}", c):
                found.append(c)
    return ", ".join(sorted(set(found))) if found else ""


def extract_staging(text):
    return first_match(text, [
        r"(?:stage|estadio|etapa|staging)[:\s]*([IVX]{1,4}[A-D]?(?:\s*[A-D])?)\b",
        r"\b(pT\d[a-d]?\s*(?:pN|N)\d[a-d]?\s*(?:M|pM)\d[a-d]?)\b",
        r"(?:Stage|Estadio|Etapa)\s+([IVX]{1,4}[A-C]?)",
    ])


def extract_ecog(text):
    m = re.search(r"ECOG\s*[:\-]?\s*(\d)", text, re.IGNORECASE)
    return "ECOG " + m.group(1) if m else ""


def extract_pdl1(text):
    return first_match(text, [
        r"PD[\-\s]?L1[^\n\.]{0,120}",
        r"PDL[\-\s]?1[^\n\.]{0,120}",
        r"expresi[oó]n\s+de\s+PD[\-\s]?L1[^\n\.]{0,120}",
    ])


def extract_msi(text):
    return first_match(text, [
        r"(?:MSI|microsatellite\s+instabilit|inestabilidad\s+de\s+microsatélites)[^\n\.]{0,150}",
        r"(?:dMMR|mismatch\s+repair)[^\n\.]{0,150}",
        r"(?:MLH1|MSH2|MSH6|PMS2)[^\n\.]{0,150}",
    ])


def extract_braf(text):
    return first_match(text, [r"BRAF[^\n\.]{0,120}"])


def extract_metastasis(text):
    return first_match(text, [
        r"metásta[^\n\.]{0,200}",
        r"metastati[^\n\.]{0,200}",
        r"met[aá]stasis[^\n\.]{0,200}",
    ])


def extract_prior_tx(text):
    snippets = []
    for kw in ["chemotherapy", "quimioterapia", "radiotherapy", "radioterapia",
                "surgery", "cirugía", "nephrectomy", "nefrectomía",
                "ABVD", "FOLFOX", "XELOX", "carboplatino", "paclitaxel",
                "first.?line", "primera.?línea", "second.?line", "segunda.?línea",
                "ciclos de", "cycles of"]:
        m = re.search(r"[^.]*" + kw + r"[^.]*\.", text, re.IGNORECASE)
        if m:
            s = m.group(0)[:180].strip()
            if s not in snippets:
                snippets.append(s)
        if len(snippets) >= 3:
            break
    return " | ".join(snippets) if snippets else ""


def extract_biomarkers(text):
    """Collect any biomarkers not covered by specific extractors."""
    bio = first_match(text, [
        r"(?:TMB|tumor\s+mutational\s+burden)[^\n\.]{0,120}",
        r"(?:HER2|EGFR|ALK|ROS1|KRAS|IDH1)[^\n\.]{0,120}",
        r"(?:BRCA\d?)[^\n\.]{0,120}",
    ])
    return bio


def extract_urgency(text):
    return first_match(text, [
        r"(?:riesgo\s+de\s+muerte|risk\s+of\s+death|risco\s+de\s+morte)[^\n\.]{0,150}",
        r"(?:urgente|urgencia|urgency|urgente)[^\n\.]{0,150}",
        r"(?:grave|gravidade|gravedad|grave\s+risk)[^\n\.]{0,150}",
        r"(?:progression|progresi[oó]n\s+de\s+la\s+enfermedad|progresión)[^\n\.]{0,150}",
        r"(?:paliativ[ao]|palliative)[^\n\.]{0,100}",
    ])


# ── comparison logic ────────────────────────────────────────────────────────

def compare_and_find_gaps(col_n, col_o, pdf_data):
    """
    Compare what is in col_n + col_o against what was extracted from the PDF.
    Return (gaps_dict, summary_str) where gaps_dict maps field → extracted value
    for fields NOT already captured in col_n / col_o.
    """
    combined = (col_n + " " + col_o).lower()
    gaps = {}

    def missing(value, keywords):
        """Value is non-empty AND none of keywords appear in combined text."""
        if not value:
            return False
        return not any(kw.lower() in combined for kw in keywords)

    age    = pdf_data.get("age", "")
    cid    = pdf_data.get("cid", "")
    stage  = pdf_data.get("staging", "")
    ecog   = pdf_data.get("ecog", "")
    pdl1   = pdf_data.get("pdl1", "")
    msi    = pdf_data.get("msi", "")
    braf   = pdf_data.get("braf", "")
    meta   = pdf_data.get("metastasis", "")
    prior  = pdf_data.get("prior_tx", "")
    bio    = pdf_data.get("biomarkers", "")
    urgency = pdf_data.get("urgency", "")

    if age and not re.search(r"\d{2}\s*year", combined) and \
               not re.search(r"\d{2}\s*año", combined) and \
               not re.search(r"\d{2}\s*ano", combined):
        gaps["Patient age"] = age

    if cid and not any(c.lower() in combined for c in cid.split(",")):
        gaps["ICD code"] = cid

    if stage and not any(kw in combined for kw in
                         ["stage", "estadio", "etapa", "pt", "pn"]):
        gaps["Disease staging"] = stage[:150]

    if ecog and "ecog" not in combined:
        gaps["ECOG status"] = ecog

    if pdl1 and "pdl" not in combined and "pd-l1" not in combined and "pd l1" not in combined:
        gaps["PD-L1 expression"] = pdl1[:200]

    if msi and not any(kw in combined for kw in ["msi", "mismatch", "dmmr", "mlh", "msh"]):
        gaps["Microsatellite instability"] = msi[:200]

    if braf and "braf" not in combined:
        gaps["BRAF mutation"] = braf[:150]

    if meta and not any(kw in combined for kw in
                        ["metast", "metas", "secondary", "secondar"]):
        gaps["Metastasis details"] = meta[:200]

    if prior and not any(kw in combined for kw in
                         ["chemo", "quimio", "radio", "surgery", "cirugí",
                          "nephrectom", "line of treatment", "línea"]):
        gaps["Prior treatments"] = prior[:200]

    if bio and not any(kw in combined for kw in
                       ["tmb", "her2", "egfr", "alk", "ros1", "kras", "idh", "brca"]):
        gaps["Biomarkers"] = bio[:200]

    if urgency and not any(kw in combined for kw in
                           ["urgenc", "grave", "progression", "progresión",
                            "palliativ", "paliativ", "risk of"]):
        gaps["Clinical urgency/severity"] = urgency[:200]

    return gaps


def build_additional_column(gaps):
    """Build a concise English string for the new Excel column."""
    if not gaps:
        return ""
    parts = []
    for field, value in gaps.items():
        # Clean up value: remove newlines, collapse whitespace
        v = re.sub(r"\s+", " ", value).strip()
        parts.append(f"{field}: {v}")
    return " | ".join(parts)


# ── main ────────────────────────────────────────────────────────────────────

def load_workbook_data():
    wb = openpyxl.load_workbook(XLSX_IN, data_only=True)
    ws = wb["Sheet1"]
    headers = []
    rows = []
    row_indices = []  # 1-based excel row numbers
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c) if c is not None else "" for c in row]
            continue
        row_dict = {headers[j]: row[j] for j in range(min(len(headers), len(row)))}
        rows.append(row_dict)
        row_indices.append(i + 1)  # excel row = i+1 (1-based, header is row 1)
    return wb, ws, headers, rows, row_indices


def main():
    wb, ws, headers, rows, row_indices = load_workbook_data()

    # Find or create the new column
    NEW_COL_HEADER = "Additional Clinical Findings (from PDF)"
    if NEW_COL_HEADER in headers:
        new_col_idx = headers.index(NEW_COL_HEADER) + 1  # 1-based
    else:
        new_col_idx = len(headers) + 1
        ws.cell(row=1, column=new_col_idx, value=NEW_COL_HEADER)

    report_lines = [
        "CLINICAL FINDINGS REPORT — OPAS Pembrolizumab",
        f"Generated: {date.today().strftime('%Y-%m-%d')}",
        f"Filter: Column T (Pembrolizumab Granted) = 0 or 1 only",
        "=" * 70,
        "",
    ]

    analyzed = skipped_filter = no_pdf = 0

    for i, row in enumerate(rows):
        excel_row = row_indices[i]
        granted   = row.get("Pembrolizumab Granted")

        if not is_granted_01(granted):
            skipped_filter += 1
            continue

        country  = val(row.get("Country"))
        case_num = val(row.get("Case Number"))
        court    = val(row.get("Court"))
        dt       = fmt_date(row.get("Date"))
        col_n    = val(row.get("Disease"))
        col_o    = val(row.get("Further clinical detalis"))
        pdf_link = val(row.get("PDF Link"))
        result   = "GRANTED" if str(granted).strip() in ("1", "1.0") else "DENIED"

        # Read PDF
        pdf_path = find_pdf(pdf_link, country)
        pdf_text = ""
        if pdf_path:
            pdf_text = read_pdf(pdf_path)
        else:
            no_pdf += 1

        # Extract clinical fields from PDF
        pdf_data = {}
        if pdf_text and not pdf_text.startswith("[PDF ERROR"):
            pdf_data = {
                "age":        extract_age(pdf_text),
                "cid":        extract_cid(pdf_text),
                "staging":    extract_staging(pdf_text),
                "ecog":       extract_ecog(pdf_text),
                "pdl1":       extract_pdl1(pdf_text),
                "msi":        extract_msi(pdf_text),
                "braf":       extract_braf(pdf_text),
                "metastasis": extract_metastasis(pdf_text),
                "prior_tx":   extract_prior_tx(pdf_text),
                "biomarkers": extract_biomarkers(pdf_text),
                "urgency":    extract_urgency(pdf_text),
            }

        gaps = compare_and_find_gaps(col_n, col_o, pdf_data)
        additional_col_value = build_additional_column(gaps)

        # Write to Excel
        ws.cell(row=excel_row, column=new_col_idx, value=additional_col_value if additional_col_value else "")

        # Build report section
        report_lines.append(f"{'=' * 70}")
        report_lines.append(f"CASE #{i+1} | {country} — {case_num}")
        report_lines.append(f"Court: {court}")
        report_lines.append(f"Date: {dt} | Decision: {result}")
        report_lines.append(f"PDF: {pdf_link or 'not available'}")
        report_lines.append("")
        report_lines.append(f"COLUMN N (Disease):              {col_n or '[empty]'}")
        report_lines.append(f"COLUMN O (Further clinical):     {col_o or '[empty]'}")
        report_lines.append("")

        if not pdf_text or pdf_text.startswith("[PDF ERROR"):
            report_lines.append("PDF STATUS: Not available — comparison not possible.")
        elif not gaps:
            report_lines.append("CLINICAL GAP ANALYSIS: No additional clinical data found in the PDF beyond what is already recorded in columns N and O.")
        else:
            report_lines.append("CLINICAL DATA FOUND IN PDF BUT NOT IN COLUMNS N/O:")
            for field, value in gaps.items():
                clean = re.sub(r"\s+", " ", value).strip()
                report_lines.append(f"  [{field}]")
                report_lines.append(f"    {clean[:300]}")

        report_lines.append("")
        analyzed += 1

    # Summary
    total_eligible = analyzed
    report_lines += [
        "=" * 70,
        "SUMMARY",
        f"  Total rows in spreadsheet:          {len(rows)}",
        f"  Eligible (T = 0 or 1):              {total_eligible}",
        f"  Skipped (T not 0/1):                {skipped_filter}",
        f"  Of eligible — no PDF available:     {no_pdf}",
        f"  Of eligible — PDF read + compared:  {total_eligible - no_pdf}",
        "=" * 70,
    ]

    # Save Excel
    wb.save(XLSX_OUT)

    # Save report
    with open(REPORT_OUT, "w", encoding="utf-8") as f:
        f.write("\n".join(report_lines))

    print(f"Report: {REPORT_OUT}")
    print(f"Excel:  {XLSX_OUT}")
    print(f"Eligible cases (T=0/1): {total_eligible}")
    print(f"Cases with PDF data:    {total_eligible - no_pdf}")


if __name__ == "__main__":
    main()
