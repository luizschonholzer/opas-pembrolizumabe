"""Force Sheet1 as the visible active tab when the file opens in Excel."""
import openpyxl, zipfile, shutil
from pathlib import Path

BASE_DIR    = Path(__file__).parent
OUTPUT_FILE = BASE_DIR / "OPAS_PDF_CLAUDE_ANALISE.xlsx"
TEMP_FILE   = BASE_DIR / "OPAS_PDF_CLAUDE_ANALISE_tmp.xlsx"

# Load, move Sheet1 to front, set active index, save
wb = openpyxl.load_workbook(str(OUTPUT_FILE))

# Ensure Sheet1 is index 0
if "Sheet1" in wb.sheetnames and wb.sheetnames[0] != "Sheet1":
    idx = wb.sheetnames.index("Sheet1")
    wb.move_sheet("Sheet1", offset=-idx)

# Force active sheet index to 0 (Sheet1)
wb._active_sheet_index = 0

wb.save(str(TEMP_FILE))
wb.close()

# Now patch the workbook.xml inside the xlsx (zip) to set activeTab=0
import zipfile, re
with zipfile.ZipFile(str(TEMP_FILE), 'r') as zin:
    names = zin.namelist()
    contents = {}
    for name in names:
        contents[name] = zin.read(name)

# Patch workbook.xml: set activeTab="0"
wb_xml = contents.get("xl/workbook.xml", b"").decode("utf-8")
# Replace any activeTab="N" with activeTab="0"
wb_xml_fixed = re.sub(r'activeTab="\d+"', 'activeTab="0"', wb_xml)
if 'activeTab=' not in wb_xml_fixed:
    # Insert activeTab into workbookView if not present
    wb_xml_fixed = wb_xml_fixed.replace('<workbookView ', '<workbookView activeTab="0" ')
contents["xl/workbook.xml"] = wb_xml_fixed.encode("utf-8")

# Write patched zip
with zipfile.ZipFile(str(OUTPUT_FILE), 'w', zipfile.ZIP_DEFLATED) as zout:
    for name, data in contents.items():
        zout.writestr(name, data)

TEMP_FILE.unlink(missing_ok=True)

# Verify
wb2 = openpyxl.load_workbook(str(OUTPUT_FILE))
print("Sheets:", wb2.sheetnames)
print("Active index:", wb2._active_sheet_index)
print("Active sheet:", wb2.active.title)
cell = wb2["Sheet1"].cell(row=2, column=28)
print("Row 2 AB (first 100 chars):", str(cell.value)[:100] if cell.value else "EMPTY")
wb2.close()
print("\nFile fixed successfully.")
