import openpyxl
import pdfplumber
import os
import sys
import json

BASE_DIR = r"C:\Users\Claudio\Downloads\Processos - OPAS"
XLSX_PATH = os.path.join(BASE_DIR, "OPAS_com_links_PDF_BACKUP.xlsx")

# Map country names to folder names
COUNTRY_FOLDER = {
    "Argentina": "Argentina",
    "Brazil": "Brasil",
    "Chile": "Chile",
    "Colombia": "Colombia",
    "Ecuador": "Equador",
    "Guatemala": "Guatemala",
    "Costa Rica": "Costa Rica",
    "Uruguay": "Uruguai",
    "Paraguay": "Paraguay",
    "Peru": "Peru",
    "Bolivia": "Bolivia",
    "Venezuela": "Venezuela",
    "Guyana": "Guyana",
}

def find_pdf(pdf_name, country):
    if not pdf_name or str(pdf_name).strip() == "":
        return None
    folder = COUNTRY_FOLDER.get(country, country)
    # Search recursively in the country folder
    country_dir = os.path.join(BASE_DIR, folder)
    if os.path.exists(country_dir):
        for root, dirs, files in os.walk(country_dir):
            for f in files:
                if f.strip() == str(pdf_name).strip():
                    return os.path.join(root, f)
    # Try base dir
    direct = os.path.join(BASE_DIR, str(pdf_name).strip())
    if os.path.exists(direct):
        return direct
    return None

def extract_pdf_text(pdf_path, max_chars=8000):
    try:
        text = ""
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    text += t + "\n"
                if len(text) > max_chars:
                    break
        return text[:max_chars]
    except Exception as e:
        return f"[ERRO AO LER PDF: {e}]"

def load_spreadsheet():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    # Find the main data sheet
    sheet_names = wb.sheetnames
    print(f"Sheets: {sheet_names}", file=sys.stderr)

    # Try Sheet1
    ws = wb["Sheet1"] if "Sheet1" in sheet_names else wb[sheet_names[0]]

    rows = []
    headers = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c) if c is not None else "" for c in row]
            print(f"Headers ({len(headers)}): {headers[:10]}", file=sys.stderr)
            continue
        # Skip empty rows
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        row_dict = {headers[j]: row[j] for j in range(min(len(headers), len(row)))}
        rows.append(row_dict)

    return headers, rows

if __name__ == "__main__":
    headers, rows = load_spreadsheet()

    print(f"\nTotal rows: {len(rows)}")
    print(f"Columns: {headers}\n")

    # Print first 3 rows as sample
    for i, r in enumerate(rows[:3]):
        print(f"\n=== ROW {i+1} ===")
        for k, v in r.items():
            if v is not None and str(v).strip():
                print(f"  {k}: {v}")
