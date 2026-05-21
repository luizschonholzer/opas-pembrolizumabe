
import os
import re
import pdfplumber
import openpyxl
from copy import copy
from datetime import datetime

BASE = r'C:\Users\Claudio\Downloads\Processos - OPAS'
XLSX_IN  = os.path.join(BASE, 'OPAS_com_links_PDF_BACKUP_20260520_1713.xlsx')
XLSX_OUT = os.path.join(BASE, 'OPAS_ANALISE_ATUALIZADA_20260521.xlsx')
REPORT   = os.path.join(BASE, 'RELATORIO_ANALISE_PDFS_20260521.txt')

# ── Build filename → full path map ─────────────────────────────────────────
pdf_map = {}
for root, dirs, files in os.walk(BASE):
    for f in files:
        if f.endswith('.pdf') or f.endswith('.docx'):
            pdf_map[f] = os.path.join(root, f)

# ── Extract text from PDF ──────────────────────────────────────────────────
def extract_text(pdf_filename):
    path = pdf_map.get(pdf_filename)
    if not path:
        # try fuzzy match (strip spaces/dashes)
        norm = pdf_filename.replace(' ', '').replace('-', '').lower()
        for k, v in pdf_map.items():
            if k.replace(' ', '').replace('-', '').lower() == norm:
                path = v
                break
    if not path:
        return None, None
    try:
        if path.endswith('.docx'):
            try:
                from docx import Document
                doc = Document(path)
                text = '\n'.join([p.text for p in doc.paragraphs])
            except Exception as e:
                text = f'[DOCX read error: {e}]'
        else:
            with pdfplumber.open(path) as pdf:
                text = ''
                for page in pdf.pages:
                    t = page.extract_text()
                    if t:
                        text += t + '\n'
        return path, text
    except Exception as e:
        return path, f'[READ ERROR: {e}]'

# ── Clinical data extraction helpers ──────────────────────────────────────
def find_age(text):
    patterns = [
        r'(\d{1,3})\s*a[ñn]os?\s+de\s+edad',
        r'paciente\s+de\s+(\d{1,3})\s*a[ñn]os?',
        r'(\d{1,3})\s*years?\s+old',
        r'age[d]?\s*[:\-]?\s*(\d{1,3})',
        r'(\d{1,3})\s*a[ñn]os?\b',
        r'paciente\s+.*?(\d{1,3})\s*a[ñn]os?',
    ]
    for pat in patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            age = int(m.group(1))
            if 10 <= age <= 100:
                return age
    return None

def find_staging(text):
    m = re.search(r'[Ee]tapa\s+(IV|III|II|I)\b|[Ss]tage\s+(IV|III|II|I)\b|[Ee]stadio\s+(IV|III|II|I)\b|[Ee]st[aá]dium\s+(IV|III|II|I)\b|EC\s+(IV|III|II|I)', text)
    if m:
        stage = m.group(1) or m.group(2) or m.group(3) or m.group(4) or m.group(5)
        return stage
    return None

def find_pdl1(text):
    m = re.search(r'PD[L\-]?1\s*[:\-]?\s*([\d<>]+\s*%?|positiv\w*|negativ\w*|alto|bajo|high|low)', text, re.IGNORECASE)
    if m:
        return m.group(1).strip()
    return None

def find_msi(text):
    if re.search(r'MSI[- ]?H|inestabilidad\s+de\s+microsatélites?\s+(alta|elevada)|alta\s+inestabilidad|high\s+microsatellite|dMMR|deficiente\s+en\s+reparaci[oó]n', text, re.IGNORECASE):
        return 'Alta instabilidade de microssatélites (MSI-H / dMMR)'
    if re.search(r'MSI[- ]?L|MSS|microsatélites?\s+estables?|estabilidad\s+de\s+microsatélites?|pMMR|proficient\s+mismatch', text, re.IGNORECASE):
        return 'Microssatélites estáveis (MSS / pMMR)'
    return None

def find_ecog(text):
    m = re.search(r'ECOG\s*[:\-]?\s*([0-4])', text, re.IGNORECASE)
    if m:
        return m.group(1)
    return None

def find_braf(text):
    m = re.search(r'BRAF\s*[:\-]?\s*(V600[EK]?|positiv\w*|negativ\w*|mutad\w*|no\s+mutad\w*|wild\s+type|selvagem)', text, re.IGNORECASE)
    if m:
        return m.group(0).strip()
    return None

def find_egfr(text):
    m = re.search(r'EGFR\s*[:\-]?\s*(positiv\w*|negativ\w*|mutad\w*|no\s+mutad\w*|wild\s+type|ausent\w*)', text, re.IGNORECASE)
    if m:
        return m.group(0).strip()
    return None

def find_alk(text):
    m = re.search(r'ALK\s*[:\-]?\s*(positiv\w*|negativ\w*|transloca[cç][aã]o|rearranjo|no\s+transloca|ausent\w*)', text, re.IGNORECASE)
    if m:
        return m.group(0).strip()
    return None

def find_her2(text):
    m = re.search(r'HER[- ]?2\s*[:\-]?\s*(positiv\w*|negativ\w*|\d\+|sobreexpress\w*|amplificad\w*)', text, re.IGNORECASE)
    if m:
        return m.group(0).strip()
    return None

def find_kras(text):
    m = re.search(r'KRAS\s*[:\-]?\s*(mutad\w*|positiv\w*|negativ\w*|selvagem|wild\s*type|no\s+mutad\w*)', text, re.IGNORECASE)
    if m:
        return m.group(0).strip()
    return None

def find_metastasis(text):
    sites = []
    organ_map = {
        'pulmón|pulmonar|pulmonary|lung': 'pulmão',
        'hígado|hepático|hepáticas?|liver|hepatic': 'fígado',
        'hueso|óseo|ósseo|osso|bone': 'osso',
        'peritoneo|peritoneal': 'peritônio',
        'cerebro|cerebral|brain|encéfalo': 'cérebro',
        r'ganglios?|linfonodos?|linfáticos?|lymph\s*nodes?|linfonodal': 'linfonodos',
        'piel|cutáneo|cutânea|skin': 'pele',
        'suprarrenal|adrenal': 'suprarrenal',
    }
    for pat, label in organ_map.items():
        if re.search(r'metástasis\s+(?:a|en|de|para)?\s*(?:\w+\s+)*' + pat, text, re.IGNORECASE) or \
           re.search(r'metastasi[sc]\s+(?:to|in)?\s*(?:\w+\s+)*' + pat, text, re.IGNORECASE) or \
           re.search(r'comprometim\w+\s+(?:\w+\s+)*' + pat, text, re.IGNORECASE):
            sites.append(label)
    return sites

DRUG_KEYWORDS = [
    'carboplatino|carboplatin', 'paclitaxel|taxol', 'cisplatino|cisplatin',
    'pemetrexed', 'docetaxel', 'gemcitabina|gemcitabine', 'oxaliplatino|oxaliplatin',
    'bevacizumab', 'cetuximab', 'rituximab', 'axitinib', 'sunitinib', 'sorafenib',
    'lenvatinib', 'capecitabina|capecitabine', 'ciclofosfamida|cyclophosphamide',
    'doxorubicina|doxorubicin', 'epirrubicina|epirubicin', 'vincristina|vincristine',
    'brentuximab', 'ABVD', 'CHOP', 'DHAP', 'FOLFOX', 'FOLFIRI', 'XELOX|CAPOX',
    'XELIRI|CAPIRI', 'trastuzumab', 'pertuzumab', 'nivolumab', 'ipilimumab',
    'atezolizumab', 'durvalumab', 'abemaciclib', 'letrozol|letrozole',
    'metotrexato|methotrexate', 'fluorouracilo|fluorouracil|5-FU',
    'irinotecan', 'temozolomida|temozolomide',
]

def find_prior_treatments(text):
    found = []
    for drug_pat in DRUG_KEYWORDS:
        if re.search(drug_pat, text, re.IGNORECASE):
            found.append(drug_pat.split('|')[0].capitalize())
    # surgeries
    if re.search(r'nefrectom[íi]a|nephrectomy', text, re.IGNORECASE):
        found.append('Nefrectomia')
    if re.search(r'mastectom[íi]a|mastectomy', text, re.IGNORECASE):
        found.append('Mastectomia')
    if re.search(r'glossectom[íi]a|glossectomy', text, re.IGNORECASE):
        found.append('Glossectomia')
    if re.search(r'trasplante\s+de\s+médula|bone\s+marrow\s+transplant|autólogo|autólogo', text, re.IGNORECASE):
        found.append('Transplante de medula óssea')
    if re.search(r'radioterapia|radiotherapy|radiation', text, re.IGNORECASE):
        found.append('Radioterapia')
    if re.search(r'cirugía|cirug[ií]a\s+\w+|surgery|procedimiento\s+quir[uú]rgico', text, re.IGNORECASE):
        found.append('Procedimento cirúrgico')
    return list(dict.fromkeys(found))  # deduplicate preserving order

def find_urgency(text):
    if re.search(r'riesgo\s+de\s+vida|risco\s+de\s+vida|risk\s+of\s+death|urgente|urgência|life[- ]threatening|peligro\s+de\s+muerte', text, re.IGNORECASE):
        return 'Condição com risco de vida — tratamento urgentemente necessário'
    if re.search(r'estado\s+grave|estado\s+crítico|gravidade|condição\s+grave|serious\s+condition|critical\s+condition', text, re.IGNORECASE):
        return 'Paciente em condição clínica grave'
    if re.search(r'urgente|urgente\s+por\s+el\s+médico|médico\s+tratante\s+indicó\s+urgencia|tratamiento\s+urgente', text, re.IGNORECASE):
        return 'Tratamento marcado como urgente pelo médico assistente'
    if re.search(r'progresi[oó]n|progressão\s+de\s+doença|disease\s+progression|agravamiento', text, re.IGNORECASE):
        return 'Progressão ativa da doença — condição piorando apesar do tratamento'
    return None

def find_treatment_line(text):
    if re.search(r'neoadyuvante|neoadjuvante|neoadjuvant', text, re.IGNORECASE):
        return 'Tratamento neoadjuvante (pré-cirúrgico)'
    if re.search(r'adyuvante|adjuvante|adjuvant', text, re.IGNORECASE):
        return 'Tratamento adjuvante (pós-cirúrgico)'
    if re.search(r'paliativo|paliativa|palliative', text, re.IGNORECASE):
        return 'Tratamento paliativo (doença incurável; objetivo é qualidade de vida)'
    if re.search(r'primera\s+l[íi]nea|primera\s+linea|first[- ]line', text, re.IGNORECASE):
        return 'Primeira linha de tratamento'
    if re.search(r'segunda\s+l[íi]nea|segunda\s+linea|second[- ]line', text, re.IGNORECASE):
        return 'Segunda linha de tratamento'
    return None

# ── Build AB content from extracted info ──────────────────────────────────
def build_ab(age, staging, pdl1, msi, ecog, braf, egfr, alk, her2, kras, metastasis, treatments, urgency, treatment_line):
    parts = []
    if age:
        parts.append(f'Idade do paciente: {age} anos')
    if staging:
        labels = {'IV': 'Estádio IV (doença metastática/avançada)', 'III': 'Estádio III (regionalmente avançado)',
                  'II': 'Estádio II (localmente avançado)', 'I': 'Estádio I (doença localizada)'}
        parts.append(f'Estadiamento da doença: {labels.get(staging, "Estádio " + staging)}')
    if ecog:
        ecog_labels = {'0': 'ECOG 0 — totalmente ativo, sem sintomas', '1': 'ECOG 1 — sintomático mas totalmente ambulatorial',
                       '2': 'ECOG 2 — ambulatorial; incapaz para atividades pesadas', '3': 'ECOG 3 — confinado à cama > 50% do dia', '4': 'ECOG 4 — totalmente incapacitado'}
        parts.append(f'Estado funcional (ECOG): {ecog_labels.get(ecog, "ECOG " + ecog)}')
    if pdl1:
        parts.append(f'Expressão de PD-L1: {pdl1}')
    if msi:
        parts.append(f'Estado de microssatélites: {msi}')
    if braf:
        parts.append(f'Biomarcador BRAF: {braf}')
    if egfr:
        parts.append(f'Biomarcador EGFR: {egfr}')
    if alk:
        parts.append(f'Biomarcador ALK: {alk}')
    if her2:
        parts.append(f'Biomarcador HER2: {her2}')
    if kras:
        parts.append(f'Biomarcador KRAS: {kras}')
    if metastasis:
        parts.append(f'Locais de metástase: {", ".join(metastasis)}')
    if treatments:
        parts.append(f'Tratamentos anteriores: {", ".join(treatments)}')
    if treatment_line:
        parts.append(f'Contexto do tratamento: {treatment_line}')
    if urgency:
        parts.append(f'Urgência clínica: {urgency}')
    return ' | '.join(parts) if parts else None

# ── Main processing ────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(XLSX_IN)
ws = wb.active

report_lines = []
report_lines.append('RELATÓRIO DE ANÁLISE DOS PDFs — ' + datetime.now().strftime('%d/%m/%Y %H:%M'))
report_lines.append('=' * 80)
report_lines.append('')

errors_found = []
filled_rows = []
skipped_rows = []
pdf_not_found = []

for row in range(2, ws.max_row + 1):
    country = ws.cell(row, 1).value
    if not country:
        continue

    aa_val = ws.cell(row, 27).value  # PDF Link
    if not aa_val or not str(aa_val).strip():
        continue

    pdf_filename = str(aa_val).strip()
    case_num = str(ws.cell(row, 6).value or '').strip()
    n_val = str(ws.cell(row, 14).value or '').strip()
    o_val = str(ws.cell(row, 15).value or '').strip()
    ab_val = str(ws.cell(row, 28).value or '').strip()

    path, text = extract_text(pdf_filename)

    report_lines.append(f'LINHA {row} | País: {country} | Processo: {case_num} | PDF: {pdf_filename}')

    if not path:
        report_lines.append(f'  ❌ PDF NÃO ENCONTRADO: {pdf_filename}')
        pdf_not_found.append((row, country, case_num, pdf_filename))
        report_lines.append('')
        continue

    if '[READ ERROR' in (text or '') or '[DOCX read error' in (text or ''):
        report_lines.append(f'  ❌ ERRO AO LER O ARQUIVO: {text}')
        report_lines.append('')
        continue

    if not text or len(text.strip()) < 50:
        report_lines.append(f'  ⚠️  PDF com conteúdo insuficiente para extração ({len(text or "")} chars)')
        report_lines.append('')
        continue

    # Skip rows classified as procedural/not relevant
    skip_flags = ['procedural issue', 'not about pembrolizumab', 'NI']
    is_skip = any(sf.lower() in n_val.lower() for sf in skip_flags)

    # Extract clinical data
    age = find_age(text)
    staging = find_staging(text)
    pdl1 = find_pdl1(text)
    msi = find_msi(text)
    ecog = find_ecog(text)
    braf = find_braf(text)
    egfr = find_egfr(text)
    alk = find_alk(text)
    her2 = find_her2(text)
    kras = find_kras(text)
    metastasis_sites = find_metastasis(text)
    treatments = find_prior_treatments(text)
    urgency = find_urgency(text)
    treatment_line = find_treatment_line(text)

    new_ab = build_ab(age, staging, pdl1, msi, ecog, braf, egfr, alk, her2, kras,
                      metastasis_sites, treatments, urgency, treatment_line)

    # Check for PDF mismatch (case number in PDF vs case in spreadsheet)
    mismatch_flag = False
    if case_num and len(case_num) > 3:
        # Look for case number in PDF text
        case_clean = re.sub(r'[^0-9]', '', case_num)
        if case_clean and len(case_clean) >= 4:
            if case_clean not in re.sub(r'[^0-9]', '', text[:3000]):
                mismatch_flag = True
                report_lines.append(f'  ⚠️  POSSÍVEL PDF INCORRETO: número do processo {case_num} não encontrado no PDF')
                errors_found.append((row, country, case_num, pdf_filename, 'PDF pode não corresponder ao processo'))

    changes_made = []

    # Update AB if empty or significantly shorter than new content
    if new_ab:
        if not ab_val or ab_val in ['None', '']:
            ws.cell(row, 28).value = new_ab
            changes_made.append(f'AB preenchido: {new_ab[:120]}...')
            filled_rows.append((row, country, case_num))
        elif len(new_ab) > len(ab_val) + 50:
            # New content is substantially richer — append what's missing
            existing_lower = ab_val.lower()
            additions = []
            for part in new_ab.split(' | '):
                key = part.split(':')[0].strip().lower()
                if key not in existing_lower:
                    additions.append(part)
            if additions:
                updated_ab = ab_val + ' | ' + ' | '.join(additions)
                ws.cell(row, 28).value = updated_ab
                changes_made.append(f'AB complementado com: {" | ".join(additions)[:120]}')
                filled_rows.append((row, country, case_num))

    # Report what was found
    findings = []
    if age: findings.append(f'Idade: {age} anos')
    if staging: findings.append(f'Estadiamento: {staging}')
    if pdl1: findings.append(f'PD-L1: {pdl1}')
    if msi: findings.append(f'MSI: {msi[:40]}')
    if ecog: findings.append(f'ECOG: {ecog}')
    if treatments: findings.append(f'Tratamentos anteriores: {", ".join(treatments[:4])}')
    if urgency: findings.append(f'Urgência: {urgency[:50]}')

    if findings:
        report_lines.append(f'  ✅ Dados extraídos: {" | ".join(findings)}')
    if changes_made:
        for c in changes_made:
            report_lines.append(f'  📝 Alteração: {c}')
    else:
        report_lines.append(f'  ℹ️  Sem alterações necessárias na coluna AB')

    report_lines.append('')

# Summary
report_lines.append('=' * 80)
report_lines.append('RESUMO')
report_lines.append('=' * 80)
report_lines.append(f'Total de linhas com PDF analisadas: {len(filled_rows) + len(skipped_rows) + len(pdf_not_found)}')
report_lines.append(f'Linhas com dados complementados (AB): {len(filled_rows)}')
report_lines.append(f'PDFs não encontrados: {len(pdf_not_found)}')
report_lines.append(f'Erros / alertas de PDF incorreto: {len(errors_found)}')
report_lines.append('')

if pdf_not_found:
    report_lines.append('PDFs NÃO ENCONTRADOS:')
    for row, country, case, pdf in pdf_not_found:
        report_lines.append(f'  Linha {row} | {country} | {case} | {pdf}')
    report_lines.append('')

if errors_found:
    report_lines.append('ALERTAS DE PDF POSSIVELMENTE INCORRETO:')
    for row, country, case, pdf, msg in errors_found:
        report_lines.append(f'  Linha {row} | {country} | {case} | {pdf} — {msg}')
    report_lines.append('')

# Save
wb.save(XLSX_OUT)
print(f'Planilha salva: {XLSX_OUT}')

with open(REPORT, 'w', encoding='utf-8') as f:
    f.write('\n'.join(report_lines))
print(f'Relatório salvo: {REPORT}')
print(f'Linhas com AB complementado: {len(filled_rows)}')
print(f'Erros/alertas: {len(errors_found)}')
print(f'PDFs não encontrados: {len(pdf_not_found)}')
