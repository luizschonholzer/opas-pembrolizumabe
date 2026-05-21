# -*- coding: utf-8 -*-
"""
OPAS Pembrolizumab — v4
- Filters: column T = 0 or 1 only
- Compares PDF clinical data against columns N and O
- Writes findings in English to new column
- Removes worksheet protection so the file stays editable
"""

import openpyxl
import pdfplumber
import os
import re
from datetime import date, timedelta

BASE_DIR   = r"C:\Users\Claudio\Downloads\Processos - OPAS"
XLSX_IN    = os.path.join(BASE_DIR, "OPAS_com_links_PDF_BACKUP.xlsx")
XLSX_OUT   = os.path.join(BASE_DIR, "OPAS_com_links_PDF.xlsx")
REPORT_OUT = os.path.join(BASE_DIR, "RELATORIO_ACHADOS_CLINICOS.txt")

COUNTRY_FOLDER = {
    "Argentina":  "Argentina",
    "Brazil":     "Brasil",
    "Chile":      "Chile",
    "Colombia":   "Colombia",
    "Ecuador":    "Equador",
    "Guatemala":  "Guatemala",
    "Costa Rica": "Costa Rica",
    "Uruguay":    "Uruguai",
    "Paraguay":   "Paraguay",
    "Peru":       "Peru",
    "Bolivia":    "Bolivia",
    "Venezuela":  "Venezuela",
    "Guyana":     "Guyana",
    "Canada":     "Canada",
}

# ── basic term translation map ───────────────────────────────────────────────
ORGAN_MAP = {
    "pulmón": "lung", "pulmon": "lung", "pulmão": "lung",
    "hígado": "liver", "higado": "liver", "fígado": "liver",
    "hígados": "liver",
    "riñón": "kidney", "rinon": "kidney", "rim": "kidney",
    "hueso": "bone", "huesos": "bone", "osso": "bone",
    "ganglios": "lymph nodes", "ganglionar": "lymph node",
    "peritoneo": "peritoneum", "peritoneal": "peritoneal",
    "cerebro": "brain", "cerebral": "brain",
    "médula": "bone marrow", "medula": "bone marrow",
    "vejiga": "bladder", "bexiga": "bladder",
    "colon": "colon", "cólon": "colon",
    "mama": "breast", "seno": "breast",
    "cuello uterino": "cervix", "cérvix": "cervix", "colo uterino": "cervix",
    "útero": "uterus", "endometrio": "endometrium",
    "ovario": "ovary", "ovários": "ovaries",
    "próstata": "prostate",
    "melanoma": "melanoma",
    "gástrico": "gastric", "estómago": "stomach",
    "piel": "skin",
    "lengua": "tongue", "lengua móvil": "mobile tongue",
    "tonsila": "tonsil", "amígdala": "tonsil",
    "sacro": "sacrum", "fémur": "femur", "femur": "femur",
    "mediastinal": "mediastinal", "mediastino": "mediastinum",
    "inguinal": "inguinal",
    "retroperitoneal": "retroperitoneal",
    "pleural": "pleural",
    "hepático": "hepatic", "hepatica": "hepatic",
}

URGENCY_TERMS = {
    "riesgo de muerte": "risk of death",
    "risco de morte": "risk of death",
    "riesgo de fallecer": "risk of death",
    "urgente": "urgent",
    "urgencia": "urgency — treatment needed immediately",
    "gravemente": "serious/severe condition",
    "grave": "serious/severe condition",
    "gravidad": "serious condition",
    "gravidade": "serious condition",
    "progresión de la enfermedad": "disease progression",
    "progressão da doença": "disease progression",
    "progresión": "disease progression",
    "progresion": "disease progression",
    "paliativo": "palliative treatment",
    "paliativa": "palliative treatment",
    "palliative": "palliative treatment",
    "irreversible": "irreversible damage",
    "irreversível": "irreversible damage",
}

STAGE_MAP = {
    "i ": "Stage I", "ii ": "Stage II", "iii ": "Stage III", "iv ": "Stage IV",
    "etapa i": "Stage I", "etapa ii": "Stage II", "etapa iii": "Stage III",
    "etapa iv": "Stage IV", "estadio i": "Stage I", "estadio ii": "Stage II",
    "estadio iii": "Stage III", "estadio iv": "Stage IV",
    "stage i": "Stage I", "stage ii": "Stage II",
    "stage iii": "Stage III", "stage iv": "Stage IV",
}


def translate_snippet(text):
    """Best-effort translation of short clinical snippets to English."""
    if not text:
        return ""
    out = text
    for es, en in ORGAN_MAP.items():
        out = re.sub(r"\b" + re.escape(es) + r"\b", en, out, flags=re.IGNORECASE)
    # Remove common filler phrases
    for filler in [
        r"que\s+padece[a-z]*", r"la\s+cual\s+padece[a-z]*",
        r"de\s+la\s+cual\s+es\s+portador[a]?",
        r"corte\s+de\s+constitucionalidad.*",
        r"república\s+de\s+guatemala.*",
        r"\(cid:\d+\)\s*",
        r"pagina:\d+.*hora:\d+.*",
        r"sistema\s+de\s+jurisprudencia.*",
        r"fecha\s+emision:.*",
    ]:
        out = re.sub(filler, "", out, flags=re.IGNORECASE)
    out = re.sub(r"\s{2,}", " ", out).strip(" |.,;")
    return out


# ── PDF helpers ──────────────────────────────────────────────────────────────

def find_pdf(pdf_name, country):
    if not pdf_name or str(pdf_name).strip() in ("", "None"):
        return None
    pdf_name = str(pdf_name).strip()
    folder = COUNTRY_FOLDER.get(country, country)
    base = os.path.join(BASE_DIR, folder)
    if os.path.exists(base):
        for root, _, files in os.walk(base):
            for f in files:
                if f.strip() == pdf_name:
                    return os.path.join(root, f)
    direct = os.path.join(BASE_DIR, pdf_name)
    return direct if os.path.exists(direct) else None


def read_pdf(path, max_pages=20):
    try:
        pages = []
        with pdfplumber.open(path) as pdf:
            for p in pdf.pages[:max_pages]:
                t = p.extract_text()
                if t:
                    pages.append(t)
        return "\n".join(pages)
    except Exception as e:
        return f"[PDF ERROR: {e}]"


# ── structured extractors (output always in English) ─────────────────────────

def extract_age(text):
    for pat in [
        r"(\d{1,3})\s*(?:years?\s*old|años?\s*de\s*edad|anos?\s*de\s*idade)",
        r"(?:age|edad|idade)\s*[:\-]?\s*(\d{1,3})\b",
        r"paciente\s+de\s+(\d{1,3})\s+años?",
        r"(\d{1,3})\s*-year-old",
        r"\((\d{2,3})\s*años?\)",
        r"\b(\d{2,3})\s+años?\b",
    ]:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            gs = [g for g in m.groups() if g and g.isdigit()]
            if gs and 1 <= int(gs[0]) <= 120:
                return f"{gs[0]}-year-old patient"
    return ""


def extract_cid(text):
    found = []
    for pat in [
        r"(?:CID|ICD|CIE|ICD[\s-]?10)[\s\-:]*([A-Z]\d{2}(?:\.\d{1,2})?)",
        r"\b([C][0-9]{2}\.[0-9]{1,2})\b",
    ]:
        for m in re.finditer(pat, text, re.IGNORECASE):
            c = (m.group(1) if m.lastindex >= 1 else m.group(0)).upper()
            if re.match(r"[A-Z]\d{2}", c):
                found.append(c)
    return "ICD-10 code: " + ", ".join(sorted(set(found))) if found else ""


STAGE_DESCRIPTIONS = {
    "I": "Stage I (localized disease)",
    "IA": "Stage IA (localized disease)",
    "IB": "Stage IB (localized disease)",
    "II": "Stage II (locally advanced)",
    "IIA": "Stage IIA (locally advanced)",
    "IIB": "Stage IIB (locally advanced)",
    "III": "Stage III (regionally advanced)",
    "IIIA": "Stage IIIA (regionally advanced)",
    "IIIB": "Stage IIIB (regionally advanced)",
    "IIIC": "Stage IIIC (regionally advanced)",
    "IV": "Stage IV (metastatic/advanced disease)",
    "IVA": "Stage IVA (metastatic/advanced disease)",
    "IVB": "Stage IVB (metastatic/advanced disease)",
    "IVC": "Stage IVC (metastatic/advanced disease)",
}

def extract_staging(text):
    m = re.search(r"\b(pT\d[a-d]?\s*(?:pN|N)\d[a-d]?\s*(?:M|pM)\d[a-d]?)\b", text, re.IGNORECASE)
    if m:
        return f"Pathological staging: {m.group(1).upper()}"
    m = re.search(
        r"(?:stage|estadio|etapa|staging)[:\s]*([IVX]{1,4}[A-D]?(?:\s*[A-D])?)\b",
        text, re.IGNORECASE)
    if m:
        stage = m.group(1).strip().upper().replace(" ", "")
        return STAGE_DESCRIPTIONS.get(stage, f"Stage {stage}")
    return ""


ECOG_DESCRIPTIONS = {
    "0": "ECOG 0 — fully active, no symptoms",
    "1": "ECOG 1 — symptomatic but fully ambulatory",
    "2": "ECOG 2 — ambulatory but unable to work; up >50% of waking hours",
    "3": "ECOG 3 — limited self-care; confined to bed >50% of waking hours",
    "4": "ECOG 4 — completely disabled, no self-care",
}

def extract_ecog(text):
    m = re.search(r"ECOG\s*[:\-]?\s*(\d)", text, re.IGNORECASE)
    if m:
        n = m.group(1)
        return "Performance status: " + ECOG_DESCRIPTIONS.get(n, f"ECOG {n}")
    return ""


def extract_pdl1(text):
    m = re.search(r"PD[\-\s]?L[\-\s]?1[^\n\.]{0,100}?(\d+\s*%)", text, re.IGNORECASE)
    if m:
        return f"PD-L1 protein expression (tumor marker for immunotherapy response): {m.group(1)}"
    m = re.search(r"PD[\-\s]?L[\-\s]?1[^\n\.]{0,80}?(?:positive|negativo|negative|pos|neg)",
                  text, re.IGNORECASE)
    if m:
        snippet = m.group(0)[:120]
        if re.search(r"negativ", snippet, re.IGNORECASE):
            return "PD-L1 protein expression: negative (low predicted response to immunotherapy)"
        if re.search(r"positiv|pos\b", snippet, re.IGNORECASE):
            pct = re.search(r"(\d+)\s*%", snippet)
            pct_str = f" ({pct.group(1)}%)" if pct else ""
            return f"PD-L1 protein expression: positive{pct_str}"
    m = re.search(r"PDL[\-\s]?1[^\n\.]{0,80}", text, re.IGNORECASE)
    if m:
        return "PD-L1 expression noted: " + translate_snippet(m.group(0)[:120])
    return ""


def extract_msi(text):
    if re.search(r"MSI[\s\-]?H\b|microsatellite\s+instabilit|inestabilidad\s+de\s+microsatélites|MSI\s+alta",
                 text, re.IGNORECASE):
        return ("High microsatellite instability (tumor has many DNA repair errors — "
                "strongly predicts response to Pembrolizumab)")
    if re.search(r"dMMR|mismatch\s+repair\s+deficien|deficiencia\s+de\s+MMR|defici[eê]ncia\s+de\s+proteínas\s+MMR",
                 text, re.IGNORECASE):
        return ("Deficient mismatch repair system (DNA repair deficiency — "
                "strongly predicts response to Pembrolizumab)")
    if re.search(r"MSI[\s\-]?L\b|microsatelite\s+estable|microsatellite\s+stable|MSS\b|pMMR",
                 text, re.IGNORECASE):
        return "Microsatellite stable / proficient mismatch repair (lower predicted response to immunotherapy)"
    m = re.search(r"(?:MLH1|MSH2|MSH6|PMS2)[^\n\.]{0,120}", text, re.IGNORECASE)
    if m:
        snippet = m.group(0)[:150]
        if re.search(r"pérdida|loss|ausente|absent|no\s+expres", snippet, re.IGNORECASE):
            gene = re.search(r"(MLH1|MSH2|MSH6|PMS2)", snippet, re.IGNORECASE).group(1)
            return (f"Loss of {gene} protein expression (DNA repair gene) — "
                    "indicates deficient mismatch repair system")
        return "DNA mismatch repair gene alteration: " + translate_snippet(snippet)
    return ""


def extract_braf(text):
    m = re.search(r"BRAF[^\n\.]{0,100}", text, re.IGNORECASE)
    if not m:
        return ""
    snippet = m.group(0)
    if re.search(r"V600E|V600", snippet, re.IGNORECASE):
        if re.search(r"negativ|no\s+se\s+detect|ausente|not\s+detect", snippet, re.IGNORECASE):
            return "BRAF V600E gene mutation: negative (not detected)"
        return "BRAF V600E gene mutation: positive (targetable oncogenic mutation)"
    if re.search(r"negativ|no\s+mutad|sin\s+mutación|wild.?type", snippet, re.IGNORECASE):
        return "BRAF gene: no mutation detected (wild-type)"
    if re.search(r"mutad|mutación|mutation|positiv", snippet, re.IGNORECASE):
        return "BRAF gene mutation detected"
    return ""


def extract_metastasis_sites(text):
    sites = []
    site_patterns = {
        "lung": r"met[aá]stasis[^\n\.]{0,60}?pulm[oó]n|pulmonar[^\n\.]{0,30}?met[aá]|lung\s+metast",
        "liver": r"met[aá]stasis[^\n\.]{0,60}?h[íi]gado|hep[aá]tic[ao][^\n\.]{0,30}?met[aá]|liver\s+metast",
        "bone": r"met[aá]stasis[^\n\.]{0,60}?(?:hueso|óseas?|oseas?)|bone\s+metast",
        "lymph nodes": r"met[aá]stasis[^\n\.]{0,60}?gangli|lymph\s+node\s+metast|ganglionar[^\n\.]{0,30}?met",
        "peritoneum": r"carcinomatosis\s+peritoneal|peritoneal\s+metast",
        "brain": r"met[aá]stasis[^\n\.]{0,60}?cerebr|brain\s+metast|CNS\s+metast",
        "skin": r"met[aá]stasis[^\n\.]{0,60}?cut[aá]nea|skin\s+metast",
        "ovary": r"met[aá]stasis[^\n\.]{0,60}?ovari",
        "pleura": r"met[aá]stasis[^\n\.]{0,60}?pleural|pleural\s+metast",
        "adrenal glands": r"met[aá]stasis[^\n\.]{0,60}?supraren|adrenal\s+metast",
    }
    for site, pat in site_patterns.items():
        if re.search(pat, text, re.IGNORECASE):
            sites.append(site)
    if sites:
        return "Metastatic spread to: " + ", ".join(sites)
    m = re.search(r"met[aá]stasis[^\n\.]{0,120}", text, re.IGNORECASE)
    if m:
        return "Metastatic disease present: " + translate_snippet(m.group(0)[:150])
    return ""


DRUG_FULL_NAMES = {
    r"\bcarboplatino\b|\bcarboplatin\b": "Carboplatin (platinum-based chemotherapy)",
    r"\bpaclitaxel\b|\btaxol\b": "Paclitaxel (taxane chemotherapy)",
    r"\bcisp?latino?\b|\bcisplatin\b": "Cisplatin (platinum-based chemotherapy)",
    r"\bgemcitabina\b|\bgemcitabine\b|\bgemzar\b": "Gemcitabine (antimetabolite chemotherapy)",
    r"\bpemetrexed\b|\balimta\b": "Pemetrexed (antimetabolite chemotherapy)",
    r"\bdocetaxel\b|\btaxotere\b": "Docetaxel (taxane chemotherapy)",
    r"\boxaliplatino\b|\boxaliplatin\b": "Oxaliplatin (platinum-based chemotherapy)",
    r"\bcapecitabina\b|\bcapecitabine\b|\bxeloda\b": "Capecitabine (oral fluoropyrimidine chemotherapy)",
    r"\bbevacizumab\b|\bavastin\b": "Bevacizumab (anti-angiogenic targeted therapy)",
    r"\bcetuximab\b|\berbitux\b": "Cetuximab (anti-EGFR monoclonal antibody)",
    r"\baxitinib\b|\binlyta\b": "Axitinib (tyrosine kinase inhibitor — targeted therapy)",
    r"\blenvatinib\b|\blenvima\b": "Lenvatinib (tyrosine kinase inhibitor — targeted therapy)",
    r"\bdoxorrubicina\b|\bdoxorubicin\b": "Doxorubicin (anthracycline chemotherapy)",
    r"\bciclofosf?amida\b|\bcyclophosphamide\b": "Cyclophosphamide (alkylating agent chemotherapy)",
    r"\bABVD\b": "ABVD regimen (Doxorubicin + Bleomycin + Vinblastine + Dacarbazine — standard Hodgkin lymphoma chemotherapy)",
    r"\bDHAP\b": "DHAP regimen (Dexamethasone + Cytarabine + Cisplatin — salvage chemotherapy for lymphoma)",
    r"\bR-GEMOX\b|\bGEMOX\b": "R-GEMOX regimen (Rituximab + Gemcitabine + Oxaliplatin — salvage chemotherapy)",
    r"\bFOLFOX\b": "FOLFOX regimen (Folinic acid + Fluorouracil + Oxaliplatin — colorectal cancer chemotherapy)",
    r"\bCAPOX\b|\bXELOX\b": "CAPOX/XELOX regimen (Capecitabine + Oxaliplatin — colorectal cancer chemotherapy)",
    r"\bXELIRI\b|\bCAPIRI\b": "XELIRI/CAPIRI regimen (Capecitabine + Irinotecan — colorectal cancer chemotherapy)",
    r"\bICE\b\s+(?:quimio|chemo|esquema)": "ICE regimen (Ifosfamide + Carboplatin + Etoposide — salvage chemotherapy)",
    r"\brituximab\b": "Rituximab (anti-CD20 monoclonal antibody — targeted therapy for lymphoma)",
    r"\btrastuzumab\b|\bherceptin\b": "Trastuzumab (anti-HER2 monoclonal antibody — targeted therapy)",
}

def extract_prior_treatments(text):
    drugs = []
    for pat, full_name in DRUG_FULL_NAMES.items():
        if re.search(pat, text, re.IGNORECASE) and full_name not in drugs:
            drugs.append(full_name)

    procedures = []
    if re.search(r"nefrectom[íi]a|nephrectomy", text, re.IGNORECASE):
        procedures.append("kidney removal surgery (nephrectomy)")
    if re.search(r"hepatectom[íi]a|hepatectomy", text, re.IGNORECASE):
        procedures.append("partial liver removal surgery (hepatectomy)")
    if re.search(r"radioterapia|radiotherapy|radiation\s+therapy", text, re.IGNORECASE):
        procedures.append("radiotherapy (radiation treatment)")
    if re.search(r"trasplante\s+de\s+médula|bone\s+marrow\s+transplant|trasplante\s+autólogo|autologous\s+transplant",
                 text, re.IGNORECASE):
        procedures.append("autologous bone marrow transplant")
    if re.search(r"glosectom[íi]a|glossectomy", text, re.IGNORECASE):
        procedures.append("partial tongue removal surgery (glossectomy)")
    if re.search(r"cirug[íi]a|surgery|operaci[oó]n", text, re.IGNORECASE) \
            and "kidney removal" not in str(procedures) \
            and "liver removal" not in str(procedures):
        procedures.append("surgical procedure")

    parts = []
    if drugs:
        parts.append("Prior chemotherapy/targeted therapy: " + "; ".join(drugs))
    if procedures:
        parts.append("Prior procedures: " + ", ".join(procedures))
    return " | ".join(parts) if parts else ""


def extract_biomarkers(text):
    results = []
    m = re.search(r"TMB[^\n\.]{0,80}", text, re.IGNORECASE)
    if m:
        snippet = m.group(0)[:100]
        pct = re.search(r"(\d+\.?\d*)\s*mut", snippet, re.IGNORECASE)
        if pct:
            results.append(f"Tumor mutational burden (TMB): {pct.group(1)} mutations/Mb "
                           f"(high TMB predicts response to immunotherapy)")
        else:
            results.append("Tumor mutational burden (TMB) noted: " + translate_snippet(snippet))
    for gene, desc in [
        ("HER2", "HER2 receptor overexpression (targeted therapy marker)"),
        ("EGFR", "EGFR gene mutation (lung cancer targeted therapy marker)"),
        ("ALK", "ALK gene rearrangement (lung cancer targeted therapy marker)"),
        ("ROS1", "ROS1 gene rearrangement (lung cancer targeted therapy marker)"),
        ("KRAS", "KRAS gene mutation (resistance to anti-EGFR therapy)"),
        ("IDH1", "IDH1 gene mutation (targetable mutation in bile duct/brain cancers)"),
        ("IDH2", "IDH2 gene mutation (targetable mutation)"),
    ]:
        m = re.search(gene + r"[^\n\.]{0,80}", text, re.IGNORECASE)
        if m:
            snippet = m.group(0)[:100]
            if re.search(r"negativ|no\s+express|absent|wild.?type|sin\s+mutación", snippet, re.IGNORECASE):
                results.append(f"{desc}: negative")
            elif re.search(r"positiv|express|mutad|mutation|amplif", snippet, re.IGNORECASE):
                results.append(f"{desc}: positive/detected")
            else:
                results.append(f"{desc}: " + translate_snippet(snippet))
    m = re.search(r"BRCA\d?[^\n\.]{0,80}", text, re.IGNORECASE)
    if m:
        snippet = m.group(0)[:100]
        gene = re.search(r"BRCA\d?", snippet, re.IGNORECASE).group(0).upper()
        if re.search(r"positiv|mutad|mutation|germinal|germline", snippet, re.IGNORECASE):
            results.append(f"{gene} gene mutation: positive (hereditary cancer predisposition)")
        else:
            results.append(f"{gene} gene status: " + translate_snippet(snippet))
    return " | ".join(results) if results else ""


def extract_urgency_en(text):
    if re.search(r"riesgo\s+de\s+muerte|risco\s+de\s+morte|risk\s+of\s+death", text, re.IGNORECASE):
        return "Clinically urgent: risk of death if treatment is not provided"
    if re.search(r"life[\s-]threatening|terminal", text, re.IGNORECASE):
        return "Life-threatening condition — treatment urgently needed"
    if re.search(r"urgente|urgencia|urgency", text, re.IGNORECASE):
        return "Treatment marked as urgent by the treating physician"
    if re.search(r"gravemente|gravedad|gravidade|grave\b", text, re.IGNORECASE):
        return "Patient in serious/severe clinical condition"
    if re.search(r"progresión\s+de\s+la\s+enfermedad|progressão\s+da\s+doença|disease\s+progression|progresion",
                 text, re.IGNORECASE):
        return "Active disease progression — condition worsening despite treatment"
    if re.search(r"paliativ[ao]|palliative", text, re.IGNORECASE):
        return "Treatment context: palliative (disease is incurable; goal is quality of life)"
    if re.search(r"irreversible|irreversível", text, re.IGNORECASE):
        return "Risk of irreversible harm if treatment is delayed"
    return ""


# ── gap detection ─────────────────────────────────────────────────────────────

def find_gaps(col_n, col_o, pdf_data):
    """
    Return dict of fields present in PDF but NOT specifically captured in N+O.
    Uses granular matching: generic terms in N/O (e.g. 'treatment', 'metastasis')
    do NOT suppress specific details found in the PDF.
    """
    combined = (col_n + " " + col_o).lower()
    gaps = {}

    # ── age ──────────────────────────────────────────────────────────────────
    age = pdf_data.get("age", "")
    if age and not re.search(r"\d{2}\s*year", combined):
        gaps["Patient age"] = age

    # ── ICD code ─────────────────────────────────────────────────────────────
    cid = pdf_data.get("cid", "")
    if cid and not re.search(r"\b[A-Z]\d{2}(?:\.\d)?\b", combined, re.IGNORECASE):
        gaps["ICD code"] = cid

    # ── staging: only skip if a concrete roman-numeral stage is already there ─
    staging = pdf_data.get("staging", "")
    if staging:
        has_stage = bool(re.search(
            r"\bstage\s+[IVX]+|estadio\s+[IVX]+|etapa\s+[IVX]+|"
            r"pT\d|stage\s+ii|stage\s+iii|stage\s+iv|"
            r"\biib\b|\biiib\b|\biiic\b|\biva\b|\bivb\b",
            combined, re.IGNORECASE))
        if not has_stage:
            gaps["Disease staging"] = staging

    # ── ECOG ─────────────────────────────────────────────────────────────────
    ecog = pdf_data.get("ecog", "")
    if ecog and "ecog" not in combined:
        gaps["ECOG status"] = ecog

    # ── PD-L1: skip only if a % or explicit pos/neg is already in N/O ────────
    pdl1 = pdf_data.get("pdl1", "")
    if pdl1:
        has_pdl1 = bool(re.search(
            r"pd[\-\s]?l1|pdl1", combined, re.IGNORECASE))
        if not has_pdl1:
            gaps["PD-L1 expression"] = pdl1

    # ── MSI / dMMR ───────────────────────────────────────────────────────────
    msi = pdf_data.get("msi", "")
    if msi and not re.search(
            r"msi|dmmr|mmr|mlh|msh|pms|microsatellite|instabilid",
            combined, re.IGNORECASE):
        gaps["Microsatellite status"] = msi

    # ── BRAF ─────────────────────────────────────────────────────────────────
    braf = pdf_data.get("braf", "")
    if braf and "braf" not in combined:
        gaps["BRAF mutation status"] = braf

    # ── Metastasis: only skip if SPECIFIC SITES already appear ───────────────
    metastasis = pdf_data.get("metastasis", "")
    if metastasis:
        # Extract sites from the PDF finding
        pdf_sites = set(re.findall(
            r"\b(lung|liver|bone|lymph node|brain|skin|peritoneum|pleural|"
            r"adrenal|ovary|kidney|bladder|mediastinal|retroperitoneal|inguinal)\b",
            metastasis, re.IGNORECASE))
        # Check which sites are NOT already in N/O
        missing_sites = [s for s in pdf_sites
                         if s.lower() not in combined]
        if missing_sites:
            gaps["Metastasis sites"] = "Metastasis to: " + ", ".join(sorted(missing_sites))
        elif not re.search(r"metast|secondary|secondar", combined, re.IGNORECASE):
            gaps["Metastasis details"] = metastasis

    # ── Prior treatments: only skip if SPECIFIC DRUG NAMES appear in N/O ─────
    prior_tx = pdf_data.get("prior_tx", "")
    if prior_tx:
        # Extract specific drug/procedure names from the PDF finding
        drug_names = re.findall(
            r"Carboplatin|Paclitaxel|Cisplatin|Gemcitabine|Pemetrexed|Docetaxel|"
            r"Oxaliplatin|Capecitabine|Bevacizumab|Cetuximab|Axitinib|Lenvatinib|"
            r"Doxorubicin|Cyclophosphamide|ABVD|DHAP|R-GEMOX|FOLFOX|CAPOX|XELOX|"
            r"XELIRI|Rituximab|Trastuzumab|nephrectomy|hepatectomy|"
            r"bone marrow transplant|glossectomy|radiotherapy",
            prior_tx, re.IGNORECASE)
        missing_drugs = [d for d in drug_names
                         if d.lower() not in combined]
        if missing_drugs:
            gaps["Prior treatments (specific)"] = (
                "Prior medications/procedures not in columns N/O: "
                + ", ".join(dict.fromkeys(missing_drugs)))  # deduplicated
        elif not re.search(
                r"carboplatin|paclitaxel|cisplatin|gemcitabine|folfox|"
                r"docetaxel|oxaliplatin|abvd|radiother|nephrectomy|"
                r"transplant|surgery|chemotherapy",
                combined, re.IGNORECASE):
            gaps["Prior treatments"] = prior_tx

    # ── Additional biomarkers ─────────────────────────────────────────────────
    bio = pdf_data.get("biomarkers", "")
    if bio and not re.search(
            r"tmb|her2|egfr|alk|ros1|kras|idh|brca", combined, re.IGNORECASE):
        gaps["Additional biomarkers"] = bio

    # ── Urgency: add specific nature even if generic urgency is mentioned ─────
    urgency = pdf_data.get("urgency", "")
    if urgency:
        urgency_already = re.search(
            r"urgent|progression|palliativ|risk of death|serious|grave|"
            r"progresión|progresion",
            combined, re.IGNORECASE)
        # Add only if the PDF urgency is MORE SPECIFIC than what's in N/O
        # e.g. "palliative treatment" in O is vague — still add if PDF says
        # "risk of death" or "disease progression"
        if not urgency_already:
            gaps["Clinical urgency/severity"] = urgency
        else:
            # Check if the specific urgency category adds something not stated
            if "risk of death" in urgency and "risk of death" not in combined:
                gaps["Clinical urgency/severity"] = urgency
            elif "disease progression" in urgency and \
                    not re.search(r"progression|progresión", combined, re.IGNORECASE):
                gaps["Clinical urgency/severity"] = urgency

    return gaps


def build_column_value(gaps):
    """Build concise English string for the new Excel column."""
    if not gaps:
        return ""
    parts = [f"{k}: {re.sub(r'\\s+', ' ', v).strip()}" for k, v in gaps.items()]
    return " | ".join(parts)


# ── main ──────────────────────────────────────────────────────────────────────

def val(v):
    if v is None: return ""
    s = str(v).strip()
    return "" if s.lower() in ("none","nan") else s

def fmt_date(d):
    if d is None: return "not found"
    s = str(d)
    if s.isdigit() and int(s) > 10000:
        try: return (date(1899,12,30) + timedelta(days=int(s))).strftime("%Y-%m-%d")
        except: return s
    return s[:10] if "00:00:00" in s else s

def is_01(v):
    if isinstance(v,(int,float)): return v in (0,1)
    return str(v).strip() in ("0","1","0.0","1.0")


def main():
    wb = openpyxl.load_workbook(XLSX_IN, data_only=True)
    ws = wb["Sheet1"]

    # Remove sheet protection so the file stays fully editable
    ws.protection.sheet = False

    headers = []
    rows, row_idxs = [], []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c) if c is not None else "" for c in row]
            continue
        rows.append({headers[j]: row[j] for j in range(min(len(headers), len(row)))})
        row_idxs.append(i + 1)

    NEW_COL = "Additional Clinical Findings (from PDF)"
    if NEW_COL in headers:
        new_col_idx = headers.index(NEW_COL) + 1
    else:
        new_col_idx = len(headers) + 1
        ws.cell(row=1, column=new_col_idx, value=NEW_COL)

    report = [
        "CLINICAL FINDINGS REPORT — OPAS Pembrolizumab",
        f"Generated: {date.today().strftime('%Y-%m-%d')}",
        "Filter: Column T (Pembrolizumab Granted) = 0 or 1 only",
        "=" * 72, "",
    ]

    analyzed = skipped = no_pdf = 0

    for i, row in enumerate(rows):
        granted = row.get("Pembrolizumab Granted")
        if not is_01(granted):
            skipped += 1
            continue

        country  = val(row.get("Country"))
        case_num = val(row.get("Case Number"))
        court    = val(row.get("Court"))
        col_n    = val(row.get("Disease"))
        col_o    = val(row.get("Further clinical detalis"))
        pdf_link = val(row.get("PDF Link"))
        result   = "GRANTED" if str(granted).strip() in ("1","1.0") else "DENIED"

        pdf_path = find_pdf(pdf_link, country)
        pdf_text = read_pdf(pdf_path) if pdf_path else ""
        if not pdf_path:
            no_pdf += 1

        pdf_data = {}
        if pdf_text and not pdf_text.startswith("[PDF ERROR"):
            pdf_data = {
                "age":        extract_age(pdf_text),
                "cid":        extract_cid(pdf_text),
                "staging":    extract_staging(pdf_text),
                "ecog":       extract_ecog(pdf_text),
                "pdl1":       extract_pdl1(pdf_text),
                "msi":        extract_msi(pdf_text),
                "braf":       extract_braf(pdf_text),
                "metastasis": extract_metastasis_sites(pdf_text),
                "prior_tx":   extract_prior_treatments(pdf_text),
                "biomarkers": extract_biomarkers(pdf_text),
                "urgency":    extract_urgency_en(pdf_text),
            }

        gaps = find_gaps(col_n, col_o, pdf_data)
        col_value = build_column_value(gaps)

        ws.cell(row=row_idxs[i], column=new_col_idx, value=col_value)

        # Report section
        report += [
            "=" * 72,
            f"CASE | {country} — {case_num}",
            f"Court: {court}",
            f"Date: {fmt_date(row.get('Date'))} | Decision: {result}",
            f"PDF: {pdf_link or 'not available'}",
            "",
            f"Column N (Disease):           {col_n or '[empty]'}",
            f"Column O (Further clinical):  {col_o or '[empty]'}",
            "",
        ]

        if not pdf_text or pdf_text.startswith("[PDF ERROR"):
            report.append("PDF STATUS: Not available — comparison not possible.")
        elif not gaps:
            report.append("GAP ANALYSIS: No additional clinical data found in PDF beyond columns N and O.")
        else:
            report.append("CLINICAL DATA IN PDF NOT CAPTURED IN COLUMNS N/O:")
            for field, value in gaps.items():
                clean = re.sub(r"\s+", " ", value).strip()
                report.append(f"  [{field}]: {clean}")

        report.append("")
        analyzed += 1

    report += [
        "=" * 72,
        "SUMMARY",
        f"  Total rows in spreadsheet:         {len(rows)}",
        f"  Eligible (T = 0 or 1):             {analyzed}",
        f"  Skipped (T not 0/1):               {skipped}",
        f"  No PDF available:                  {no_pdf}",
        f"  PDF read and compared:             {analyzed - no_pdf}",
        "=" * 72,
    ]

    wb.save(XLSX_OUT)
    with open(REPORT_OUT, "w", encoding="utf-8") as f:
        f.write("\n".join(report))

    print(f"Report : {REPORT_OUT}")
    print(f"Excel  : {XLSX_OUT}")
    print(f"Cases analyzed (T=0/1): {analyzed}")
    print(f"Cases with PDF:         {analyzed - no_pdf}")


if __name__ == "__main__":
    main()
