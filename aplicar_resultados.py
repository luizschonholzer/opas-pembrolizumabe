"""
Aplica os resultados do JSON de análise à planilha,
usando mapeamento linha→melhor_PDF (não PDF→linhas).
Isso evita que um PDF seja atribuído a múltiplas linhas incorretas.
"""

import sys, json, re
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Side, Border
from openpyxl.utils import get_column_letter

BASE_DIR = Path(r"C:\Users\Claudio\Downloads\Processos - OPAS")
XLSX_IN  = BASE_DIR / "OPAS_translated_english_apr 2026.xlsx"
XLSX_OUT = BASE_DIR / "OPAS_analise_clinica.xlsx"
JSON_LOG = BASE_DIR / "resultados_analise.json"

PAISES_COM_PASTA = {
    "Argentina":  "Argentina",
    "Brazil":     "Brasil",
    "Chile":      "Chile",
    "Colombia":   "Colombia",
    "Ecuador":    "Equador",
    "Guatemala":  "Guatemala",
    "Uruguay":    "Uruguai",
    "Costa Rica": "Costa Rica",
}

COL_N    = 14   # Disease
COL_O    = 15   # Further clinical details
COL_FLAG = 27   # Verificação PDF

# ── Carregamento ────────────────────────────────────────────────────────────

with open(JSON_LOG, encoding="utf-8") as f:
    resultados: dict[str, dict] = json.load(f)

wb = openpyxl.load_workbook(XLSX_IN)
ws = wb["Sheet1"]

# Adicionar/limpar cabeçalho da coluna de verificação
ws.cell(row=1, column=COL_FLAG).value = "PDF — Verificação"
ws.cell(row=1, column=COL_FLAG).font = Font(bold=True)

# ── Funções de pontuação (mesmo critério do add_pdf_links.py) ──────────────

def normalizar(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s).lower())

def numeros(s: str) -> list[str]:
    return re.findall(r"\d+", str(s))

def get_date_str(date_val) -> str | None:
    from datetime import datetime, date
    if date_val is None:
        return None
    if isinstance(date_val, (datetime, date)):
        return date_val.strftime("%Y-%m-%d")
    s = str(date_val).strip()
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        return s[:10]
    return None

def pontuar(caso_str: str, date_val, pdf_path: Path) -> int:
    """Mesma lógica de scoring do script original."""
    pdf_nome = pdf_path.name
    caso_str = str(caso_str) if caso_str else ""
    score = 0

    caso_nums = numeros(caso_str)
    pdf_nums  = numeros(pdf_nome)

    # Correspondência pela data (útil para Uruguai)
    date_str = get_date_str(date_val)
    if date_str and date_str in pdf_nome:
        score += 20
        date_parts = set(numeros(date_str))
        extras = [n for n in pdf_nums if n not in date_parts]
        for n in extras:
            score += 4 if n in caso_nums else -2

    # Correspondência textual normalizada
    c_norm = normalizar(caso_str)
    p_norm = normalizar(pdf_nome.replace(".pdf", "").replace(".docx", ""))
    if c_norm and len(c_norm) >= 4:
        if c_norm == p_norm:
            score += 15
        elif c_norm in p_norm:
            score += 8
        elif p_norm in c_norm:
            score += 5

    # Correspondência numérica (>= 4 dígitos para evitar falsos positivos com anos)
    for n in caso_nums:
        if len(n) >= 4 and n in pdf_nums:
            score += 5

    return score

# ── Mapeamento: para cada linha, encontrar o melhor PDF ──────────────────────

SKIP_VALORES = {
    "not found", "procedural", "procedural issue", "not about pembrolizumab",
    "out of scope", "defective site / inaccessible", "none", "ni", "",
}

FILL_AMARELO = PatternFill("solid", fgColor="FFFF00")
FILL_VERMELHO = PatternFill("solid", fgColor="FFCCCC")

total_atualizadas = 0
total_flags = 0

for row in range(2, ws.max_row + 1):
    pais    = ws.cell(row=row, column=1).value
    caso    = ws.cell(row=row, column=COL_N - 8).value    # coluna 6: Case Number
    date_v  = ws.cell(row=row, column=7).value
    n_atual = str(ws.cell(row=row, column=COL_N).value or "")
    o_atual = str(ws.cell(row=row, column=COL_O).value or "")

    if not pais or not PAISES_COM_PASTA.get(pais):
        continue
    if not caso or str(caso).strip().lower() in SKIP_VALORES:
        continue

    pasta = BASE_DIR / PAISES_COM_PASTA[pais]
    if not pasta.exists():
        continue

    # Listar candidatos de PDF
    candidatos = list(pasta.rglob("*.pdf")) + list(pasta.rglob("*.docx"))
    if not candidatos:
        continue

    # Escolher melhor PDF
    melhor_score = 0
    melhor_pdf   = None
    for pdf in candidatos:
        s = pontuar(caso, date_v, pdf)
        if s > melhor_score:
            melhor_score = s
            melhor_pdf   = pdf

    THRESHOLD = 8
    if melhor_score < THRESHOLD or melhor_pdf is None:
        continue

    chave = str(melhor_pdf)
    dados = resultados.get(chave, {})
    if not dados or "erro" in dados:
        continue

    # Verificar consistência do PDF
    inconsistente = (
        dados.get("sobre_pembrolizumabe") is False
        or dados.get("sobre_cancer") is False
    )
    inconsistencia_msg = dados.get("inconsistencia") or ""

    if inconsistente:
        # Destacar linha com aviso
        for col in range(1, COL_FLAG):
            ws.cell(row=row, column=col).fill = FILL_AMARELO
        cell_flag = ws.cell(row=row, column=COL_FLAG)
        cell_flag.value = f"PDF INCONSISTENTE: {inconsistencia_msg}"
        cell_flag.fill  = FILL_VERMELHO
        cell_flag.font  = Font(bold=True, color="990000")
        cell_flag.alignment = Alignment(wrap_text=True)
        total_flags += 1
        print(f"INCONSISTENTE  Linha {row:3d}: {pais} | {caso} → {melhor_pdf.name}")
        print(f"               Motivo: {inconsistencia_msg[:120]}")
        continue

    # ── Montar texto para coluna O (apenas quadro clínico) ──────────────────
    partes = []

    resumo = dados.get("quadro_clinico_resumo", "") or ""
    if resumo and "null" not in resumo.lower():
        partes.append(resumo)

    # Biomarcadores
    bm = []
    for campo, rotulo in [
        ("pdl1", "PD-L1"), ("msi", "MSI"), ("braf", "BRAF"),
        ("her2", "HER2"), ("ecog", "ECOG"), ("ki67", "Ki67"), ("brca", "BRCA"),
    ]:
        val = dados.get(campo)
        if val and str(val).lower() not in ("null", "none", ""):
            bm.append(f"{rotulo}: {val}")
    if bm:
        partes.append("Biomarcadores — " + "; ".join(bm))

    trat = dados.get("tratamentos_anteriores") or ""
    if trat and str(trat).lower() not in ("null", "none", ""):
        partes.append(f"Tratamentos anteriores: {trat}")

    linha_pemb = dados.get("linha_pembrolizumabe") or ""
    if linha_pemb and str(linha_pemb).lower() not in ("null", "none", ""):
        partes.append(f"Pembrolizumabe: {linha_pemb}")

    texto_o = " | ".join(p for p in partes if p.strip())

    # ── Atualizar coluna N se vazio/NI ──────────────────────────────────────
    diag = dados.get("diagnostico") or ""
    if (
        n_atual.strip().upper() in ("", "NI", "NONE", "NULL")
        and diag
        and str(diag).lower() not in ("null", "none", "")
    ):
        ws.cell(row=row, column=COL_N).value = diag

    # ── Atualizar coluna O se vazio/NI ──────────────────────────────────────
    if (
        o_atual.strip().upper() in ("", "NI", "NONE", "NULL")
        and texto_o
    ):
        cell_o = ws.cell(row=row, column=COL_O)
        cell_o.value = texto_o
        cell_o.alignment = Alignment(wrap_text=True)
        total_atualizadas += 1

    # Marcar coluna de verificação
    cell_vf = ws.cell(row=row, column=COL_FLAG)
    cell_vf.value = f"PDF verificado: {melhor_pdf.name}"
    cell_vf.font  = Font(color="006400", italic=True)

# Ajustar largura das colunas N, O e FLAG
for col in [COL_N, COL_O, COL_FLAG]:
    ws.column_dimensions[get_column_letter(col)].width = 55

wb.save(XLSX_OUT)
print(f"\nLinhas com coluna O atualizada : {total_atualizadas}")
print(f"PDFs inconsistentes sinalizados: {total_flags}")
print(f"Planilha salva em             : {XLSX_OUT.name}")
